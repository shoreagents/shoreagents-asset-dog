"""
Inventory API router
"""
from fastapi import APIRouter, HTTPException, Query, Depends, Path, Body
from typing import Optional
import logging
from decimal import Decimal
import re
from datetime import datetime
from fastapi.responses import StreamingResponse
import io
from models.inventory import (
    InventoryItem,
    InventoryItemCreate,
    InventoryItemUpdate,
    InventoryItemsResponse,
    InventoryItemResponse,
    InventoryTransaction,
    InventoryTransactionCreate,
    InventoryTransactionsResponse,
    InventoryTransactionResponse,
    BulkDeleteTransactionsRequest,
    BulkDeleteTransactionsResponse,
    RestoreResponse,
    BulkRestoreRequest,
    BulkRestoreResponse,
    BulkDeleteItemsRequest,
    BulkDeleteItemsResponse,
    PaginationInfo,
    GenerateCodeResponse,
    EmptyTrashResponse,
    CheckItemCodesRequest,
    CheckItemCodesResponse,
)
from auth import verify_auth
from database import prisma

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/inventory", tags=["inventory"])

async def check_permission(user_id: str, permission: str) -> bool:
    """Check if user has a specific permission"""
    try:
        asset_user = await prisma.assetuser.find_unique(
            where={"userId": user_id}
        )
        if not asset_user or not asset_user.isActive:
            return False
        return getattr(asset_user, permission, False)
    except Exception:
        return False

def is_uuid(value: str) -> bool:
    """Check if a string is a UUID"""
    uuid_pattern = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.IGNORECASE)
    return bool(uuid_pattern.match(value))


def get_company_initials(company_name: Optional[str]) -> str:
    """
    Extract company initials from company name
    Handles:
    - Multiple words: "Shore Agents" -> "SA"
    - CamelCase/combined words: "ShoreAgents" -> "SA", "ABCCompany" -> "AC"
    - Single word: "XYZ" -> "XY"
    """
    if not company_name or not company_name.strip():
        return 'SA'  # Default fallback
    
    trimmed = company_name.strip()
    
    # First, try splitting by spaces (multiple words)
    words = [w for w in trimmed.split() if w]
    
    if len(words) >= 2:
        # Multiple words: take first letter of first two words
        first = words[0][0].upper()
        second = words[1][0].upper()
        return f"{first}{second}"
    elif len(words) == 1:
        word = words[0]
        
        # Check for camelCase pattern - handles both "ShoreAgents" and "shoreAgents"
        # Pattern 1: Uppercase letter followed by lowercase, then uppercase (e.g., "ShoreAgents")
        match1 = re.match(r'^([A-Z][a-z]+)([A-Z][a-z]*)', word)
        if match1:
            first_part = match1.group(1)
            second_part = match1.group(2)
            return f"{first_part[0].upper()}{second_part[0].upper()}"
        
        # Pattern 2: Lowercase followed by uppercase (e.g., "shoreAgents")
        match2 = re.match(r'^([a-z]+)([A-Z][a-z]*)', word)
        if match2:
            first_part = match2.group(1)
            second_part = match2.group(2)
            return f"{first_part[0].upper()}{second_part[0].upper()}"
        
        # Check for all caps with word boundaries (e.g., "ABCCOMPANY" -> "AC")
        if word == word.upper() and len(word) > 2:
            first = word[0]
            for i in range(1, len(word)):
                if word[i].isalpha():
                    return f"{first}{word[i]}"
        
        # No camelCase detected: take first 2 letters
        return trimmed[:2].upper().ljust(2, 'X')
    
    return 'SA'

@router.get("/generate-code", response_model=GenerateCodeResponse)
async def generate_item_code(auth: dict = Depends(verify_auth)):
    """Generate a unique item code for a new inventory item"""
    try:
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Get company info to extract initials
        company_info = await prisma.companyinfo.find_first(
            order={"createdAt": "desc"}
        )
        
        # Get company initials (e.g., "Shore Agents" -> "SA")
        company_suffix = get_company_initials(company_info.companyName if company_info else None)
        
        # Get all item codes that match the pattern INV-XXX-[SUFFIX]
        items = await prisma.inventoryitem.find_many(
            where={
                "itemCode": {"startswith": "INV-"},
                "isDeleted": False,
            },
            order={"itemCode": "desc"}
        )
        
        # Extract the highest sequential number for this suffix
        next_number = 1
        pattern = re.compile(rf'^INV-(\d+)-{company_suffix}$')
        
        for item in items:
            match = pattern.match(item.itemCode)
            if match:
                num = int(match.group(1))
                if num >= next_number:
                    next_number = num + 1
        
        # Format: INV-001-[SUFFIX] (3 digits, zero-padded)
        next_code = f"INV-{str(next_number).zfill(3)}-{company_suffix}"
        
        # Check if the generated code already exists (safety check)
        exists = await prisma.inventoryitem.find_unique(
            where={"itemCode": next_code}
        )
        
        if exists:
            # If exists, try next number
            next_number += 1
            next_code = f"INV-{str(next_number).zfill(3)}-{company_suffix}"
        
        return GenerateCodeResponse(itemCode=next_code)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating item code: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to generate item code")


@router.post("/check-codes", response_model=CheckItemCodesResponse)
async def check_item_codes(
    request: CheckItemCodesRequest = Body(...),
    auth: dict = Depends(verify_auth)
):
    """Check which item codes already exist (including soft-deleted)"""
    try:
        if not request.itemCodes or len(request.itemCodes) == 0:
            return CheckItemCodesResponse(existingCodes=[])
        
        # Find all items with matching item codes (including soft-deleted)
        existing_items = await prisma.inventoryitem.find_many(
            where={
                "itemCode": {"in": request.itemCodes}
            }
        )
        
        existing_codes = [item.itemCode for item in existing_items]
        
        return CheckItemCodesResponse(existingCodes=existing_codes)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error checking item codes: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to check item codes")


@router.get("/export")
async def export_inventory(
    format: str = Query("excel", description="Export format"),
    search: Optional[str] = Query(None, description="Search term"),
    category: Optional[str] = Query(None, description="Filter by category"),
    lowStock: bool = Query(False, description="Filter low stock items only"),
    includeSummary: bool = Query(False, description="Include summary sheet"),
    includeByCategory: bool = Query(False, description="Include by category sheet"),
    includeByStatus: bool = Query(False, description="Include by status sheet"),
    includeTotalCost: bool = Query(False, description="Include total cost sheet"),
    includeLowStock: bool = Query(False, description="Include low stock items sheet"),
    includeItemList: bool = Query(False, description="Include item list sheet"),
    itemFields: Optional[str] = Query(None, description="Comma-separated item fields to include"),
    auth: dict = Depends(verify_auth)
):
    """Export inventory data to Excel"""
    try:
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Import openpyxl for Excel generation
        try:
            from openpyxl import Workbook  # type: ignore
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="Excel export not available - openpyxl not installed"
            )
        
        # Build where clause
        where_clause = {"isDeleted": False}
        
        # Search filter
        if search:
            where_clause["OR"] = [
                {"itemCode": {"contains": search, "mode": "insensitive"}},
                {"name": {"contains": search, "mode": "insensitive"}},
                {"description": {"contains": search, "mode": "insensitive"}},
                {"sku": {"contains": search, "mode": "insensitive"}},
                {"barcode": {"contains": search, "mode": "insensitive"}},
            ]
        
        # Category filter
        if category:
            where_clause["category"] = category
        
        # Fetch all items
        items = await prisma.inventoryitem.find_many(
            where=where_clause,
            order={"createdAt": "desc"}
        )
        
        # Filter low stock items if requested
        filtered_items = items
        if lowStock:
            filtered_items = [
                item for item in items
                if item.minStockLevel is not None 
                and float(item.currentStock) <= float(item.minStockLevel)
            ]
        
        # Helper function to format numbers
        def format_number(value) -> str:
            if value is None:
                return ''
            return f"{float(value):,.2f}"
        
        # Calculate summary data
        total_items = len(filtered_items)
        total_stock = sum(float(item.currentStock) for item in filtered_items)
        total_cost = sum(
            float(item.currentStock) * (float(item.unitCost) if item.unitCost else 0)
            for item in filtered_items
        )
        
        # Group by category
        by_category = {}
        for item in filtered_items:
            cat = item.category or 'Uncategorized'
            if cat not in by_category:
                by_category[cat] = {'count': 0, 'totalStock': 0, 'totalCost': 0}
            by_category[cat]['count'] += 1
            by_category[cat]['totalStock'] += float(item.currentStock)
            cost = float(item.unitCost) if item.unitCost else 0
            by_category[cat]['totalCost'] += float(item.currentStock) * cost
        
        # Group by status
        by_status = {}
        for item in filtered_items:
            stock = float(item.currentStock)
            min_level = float(item.minStockLevel) if item.minStockLevel else None
            if stock == 0:
                status = 'Out of Stock'
            elif min_level is not None and stock <= min_level:
                status = 'Low Stock'
            else:
                status = 'In Stock'
            
            if status not in by_status:
                by_status[status] = {'count': 0, 'totalStock': 0, 'totalCost': 0}
            by_status[status]['count'] += 1
            by_status[status]['totalStock'] += stock
            cost = float(item.unitCost) if item.unitCost else 0
            by_status[status]['totalCost'] += stock * cost
        
        # Low stock items
        low_stock_items = [
            item for item in filtered_items
            if item.minStockLevel is not None 
            and float(item.currentStock) <= float(item.minStockLevel)
        ]
        
        # Create workbook
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Summary sheet
        if includeSummary:
            ws = wb.create_sheet("Summary")
            ws.append(["Metric", "Value"])
            ws.append(["Total Items", total_items])
            ws.append(["Total Stock", int(total_stock)])
            ws.append(["Total Cost", format_number(total_cost)])
        
        # Create By Category sheet
        if includeByCategory and by_category:
            ws = wb.create_sheet("By Category")
            ws.append(["Category", "Item Count", "Total Stock", "Total Cost"])
            for cat, data in by_category.items():
                ws.append([cat, data['count'], int(data['totalStock']), format_number(data['totalCost'])])
        
        # Create By Status sheet
        if includeByStatus and by_status:
            ws = wb.create_sheet("By Status")
            ws.append(["Status", "Item Count", "Total Stock", "Total Cost"])
            for status, data in by_status.items():
                ws.append([status, data['count'], int(data['totalStock']), format_number(data['totalCost'])])
        
        # Create Total Cost sheet
        if includeTotalCost:
            ws = wb.create_sheet("Total Cost")
            ws.append(["Description", "Amount"])
            ws.append(["Total Inventory Value", format_number(total_cost)])
        
        # Create Low Stock Items sheet
        if includeLowStock and low_stock_items:
            ws = wb.create_sheet("Low Stock Items")
            ws.append(["Item Code", "Name", "Current Stock", "Min Level"])
            for item in low_stock_items:
                ws.append([
                    item.itemCode,
                    item.name,
                    int(float(item.currentStock)),
                    int(float(item.minStockLevel)) if item.minStockLevel else ''
                ])
        
        # Create Item List sheet
        if includeItemList and itemFields:
            field_list = [f.strip() for f in itemFields.split(',') if f.strip()]
            if field_list:
                field_labels = {
                    'itemCode': 'Item Code',
                    'name': 'Name',
                    'description': 'Description',
                    'category': 'Category',
                    'unit': 'Unit',
                    'currentStock': 'Current Stock',
                    'minStockLevel': 'Min Stock Level',
                    'maxStockLevel': 'Max Stock Level',
                    'unitCost': 'Unit Cost',
                    'location': 'Location',
                    'supplier': 'Supplier',
                    'brand': 'Brand',
                    'model': 'Model',
                    'sku': 'SKU',
                    'barcode': 'Barcode',
                    'remarks': 'Remarks',
                }
                
                ws = wb.create_sheet("Item List")
                # Header row
                headers = [field_labels.get(f, f) for f in field_list]
                ws.append(headers)
                
                # Data rows
                for item in filtered_items:
                    row = []
                    for field in field_list:
                        if field == 'itemCode':
                            row.append(item.itemCode)
                        elif field == 'name':
                            row.append(item.name)
                        elif field == 'description':
                            row.append(item.description or '')
                        elif field == 'category':
                            row.append(item.category or '')
                        elif field == 'unit':
                            row.append(item.unit or '')
                        elif field == 'currentStock':
                            row.append(int(float(item.currentStock)))
                        elif field == 'minStockLevel':
                            row.append(int(float(item.minStockLevel)) if item.minStockLevel else '')
                        elif field == 'maxStockLevel':
                            row.append(int(float(item.maxStockLevel)) if item.maxStockLevel else '')
                        elif field == 'unitCost':
                            row.append(format_number(item.unitCost) if item.unitCost else '')
                        elif field == 'location':
                            row.append(item.location or '')
                        elif field == 'supplier':
                            row.append(item.supplier or '')
                        elif field == 'brand':
                            row.append(item.brand or '')
                        elif field == 'model':
                            row.append(item.model or '')
                        elif field == 'sku':
                            row.append(item.sku or '')
                        elif field == 'barcode':
                            row.append(item.barcode or '')
                        elif field == 'remarks':
                            row.append(item.remarks or '')
                        else:
                            row.append('')
                    ws.append(row)
        
        # If no sheets were created, add a default empty sheet
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet("Inventory")
            ws.append(["No data selected for export"])
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"inventory-export-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting inventory: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export inventory")


@router.delete("/trash/empty", response_model=EmptyTrashResponse)
async def empty_inventory_trash(auth: dict = Depends(verify_auth)):
    """Permanently delete all soft-deleted inventory items"""
    try:
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
    
        
        # Find all soft-deleted items
        deleted_items = await prisma.inventoryitem.find_many(
            where={"isDeleted": True}
        )
        
        if len(deleted_items) == 0:
            return EmptyTrashResponse(
                success=True,
                message="Trash is already empty",
                deletedCount=0
            )
        
        # Get IDs of deleted items
        deleted_ids = [item.id for item in deleted_items]
        
        # Delete related transactions first (to avoid foreign key constraints)
        await prisma.inventorytransaction.delete_many(
            where={"inventoryItemId": {"in": deleted_ids}}
        )
        
        # Permanently delete all soft-deleted items
        result = await prisma.inventoryitem.delete_many(
            where={"isDeleted": True}
        )
        
        return EmptyTrashResponse(
            success=True,
            message=f"Permanently deleted {result} item(s)",
            deletedCount=result
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error emptying inventory trash: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to empty trash")


@router.delete("/bulk-delete", response_model=BulkDeleteItemsResponse)
async def bulk_delete_inventory_items(
    request: BulkDeleteItemsRequest = Body(...),
    auth: dict = Depends(verify_auth)
):
    """Bulk delete multiple inventory items (permanently or soft delete)"""
    try:
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        if not request.ids or len(request.ids) == 0:
            raise HTTPException(
                status_code=400,
                detail="Invalid request. Expected an array of item IDs."
            )
        
        if request.permanent:
            # Delete related transactions first (to avoid foreign key constraints)
            await prisma.inventorytransaction.delete_many(
                where={"inventoryItemId": {"in": request.ids}}
            )
            
            # Permanently delete items
            result = await prisma.inventoryitem.delete_many(
                where={"id": {"in": request.ids}}
            )
            
            return BulkDeleteItemsResponse(
                success=True,
                deletedCount=result,
                message=f"Permanently deleted {result} item(s)"
            )
        else:
            # Soft delete items
            result = await prisma.inventoryitem.update_many(
                where={"id": {"in": request.ids}},
                data={
                    "isDeleted": True,
                    "deletedAt": datetime.now()
                }
            )
            
            return BulkDeleteItemsResponse(
                success=True,
                deletedCount=result,
                message=f"Archived {result} item(s). They will be permanently deleted after 30 days."
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error bulk deleting inventory items: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to delete inventory items")


@router.get("", response_model=InventoryItemsResponse)
async def get_inventory_items(
    search: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    includeDeleted: Optional[bool] = Query(False),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=10000),
    lowStock: Optional[bool] = Query(False),
    auth: dict = Depends(verify_auth)
):
    """Get all inventory items with optional filters and pagination"""
    try:
        # Build where clause
        where_clause = {}
        
        if not includeDeleted:
            where_clause["isDeleted"] = False
        
        # Search filter
        if search:
            where_clause["OR"] = [
                {"itemCode": {"contains": search, "mode": "insensitive"}},
                {"name": {"contains": search, "mode": "insensitive"}},
                {"description": {"contains": search, "mode": "insensitive"}},
                {"sku": {"contains": search, "mode": "insensitive"}},
                {"barcode": {"contains": search, "mode": "insensitive"}},
            ]
        
        # Category filter
        if category:
            where_clause["category"] = category
        
        # Get total count and items
        # Note: Low stock filtering is done in memory since Prisma doesn't support field comparison in where clause
        total_count = await prisma.inventoryitem.count(where=where_clause)
        
        items_data = await prisma.inventoryitem.find_many(
            where=where_clause,
            order={"createdAt": "desc"},
            skip=0 if lowStock else (page - 1) * pageSize,
            take=None if lowStock else pageSize,
        )
        
        # Filter low stock items if requested
        if lowStock:
            filtered_items = []
            for item in items_data:
                if item.minStockLevel:
                    current = float(item.currentStock)
                    min_level = float(item.minStockLevel)
                    if current <= min_level:
                        filtered_items.append(item)
            items_data = filtered_items
            total_count = len(filtered_items)
            # Apply pagination after filtering
            items_data = items_data[(page - 1) * pageSize:page * pageSize]
        
        # Fetch transaction counts for all items in one query
        item_ids = [str(item.id) for item in items_data]
        transaction_counts = {}
        if item_ids:
            # Fetch all transactions for these items (Prisma Python doesn't support select)
            transactions_data = await prisma.inventorytransaction.find_many(
                where={
                    "inventoryItemId": {"in": item_ids}
                }
            )
            # Count transactions per item
            for transaction in transactions_data:
                item_id = str(transaction.inventoryItemId)
                transaction_counts[item_id] = transaction_counts.get(item_id, 0) + 1
        
        # Convert to Pydantic models
        items = []
        for item_data in items_data:
            try:
                # Get transaction count for this item
                item_id = str(item_data.id)
                transaction_count = transaction_counts.get(item_id, 0)
                count_dict = {"transactions": transaction_count}
                
                item = InventoryItem(
                    id=item_id,
                    itemCode=str(item_data.itemCode),
                    name=str(item_data.name),
                    description=item_data.description if item_data.description else None,
                    category=item_data.category if item_data.category else None,
                    unit=item_data.unit if item_data.unit else None,
                    currentStock=item_data.currentStock,
                    minStockLevel=item_data.minStockLevel if item_data.minStockLevel else None,
                    maxStockLevel=item_data.maxStockLevel if item_data.maxStockLevel else None,
                    unitCost=item_data.unitCost if item_data.unitCost else None,
                    location=item_data.location if item_data.location else None,
                    supplier=item_data.supplier if item_data.supplier else None,
                    brand=item_data.brand if item_data.brand else None,
                    model=item_data.model if item_data.model else None,
                    sku=item_data.sku if item_data.sku else None,
                    barcode=item_data.barcode if item_data.barcode else None,
                    remarks=item_data.remarks if item_data.remarks else None,
                    createdAt=item_data.createdAt,
                    updatedAt=item_data.updatedAt,
                    _count=count_dict,
                    isDeleted=item_data.isDeleted if hasattr(item_data, 'isDeleted') else False,
                    deletedAt=item_data.deletedAt if hasattr(item_data, 'deletedAt') and item_data.deletedAt else None
                )
                items.append(item)
            except Exception as e:
                logger.error(f"Error creating InventoryItem model: {type(e).__name__}: {str(e)}", exc_info=True)
                continue
        
        total_pages = (total_count + pageSize - 1) // pageSize if total_count > 0 else 1
        
        pagination = PaginationInfo(
            total=total_count,
            page=page,
            pageSize=pageSize,
            totalPages=total_pages,
            hasNextPage=page < total_pages,
            hasPreviousPage=page > 1
        )
        
        return InventoryItemsResponse(items=items, pagination=pagination)
    
    except Exception as e:
        logger.error(f"Error fetching inventory items: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch inventory items")

@router.post("", response_model=InventoryItemResponse, status_code=201)
async def create_inventory_item(
    item_data: InventoryItemCreate,
    auth: dict = Depends(verify_auth)
):
    """Create a new inventory item"""
    try:
        # Check permission
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Validation
        if not item_data.itemCode or not item_data.name:
            raise HTTPException(
                status_code=400,
                detail="Item code and name are required"
            )
        
        # Check if item code already exists
        existing_item = await prisma.inventoryitem.find_unique(
            where={"itemCode": item_data.itemCode}
        )
        
        if existing_item:
            raise HTTPException(
                status_code=400,
                detail="Item code already exists"
            )
        
        # Get user info
        user = auth.get("user", {})
        user_name = (
            user.get("user_metadata", {}).get("name") or
            user.get("user_metadata", {}).get("full_name") or
            user.get("email", "").split("@")[0] if user.get("email") else None or
            user.get("email") or
            user.get("id")
        )
        
        # Create inventory item
        created_item = await prisma.inventoryitem.create(
            data={
                "itemCode": item_data.itemCode,
                "name": item_data.name,
                "description": item_data.description,
                "category": item_data.category,
                "unit": item_data.unit,
                "currentStock": item_data.currentStock or Decimal("0"),
                "minStockLevel": item_data.minStockLevel,
                "maxStockLevel": item_data.maxStockLevel,
                "unitCost": item_data.unitCost,
                "location": item_data.location,
                "supplier": item_data.supplier,
                "brand": item_data.brand,
                "model": item_data.model,
                "sku": item_data.sku,
                "barcode": item_data.barcode,
                "remarks": item_data.remarks,
            }
        )
        
        # Create initial transaction if stock is provided
        if item_data.currentStock and float(item_data.currentStock) > 0:
            await prisma.inventorytransaction.create(
                data={
                    "inventoryItemId": created_item.id,
                    "transactionType": "IN",
                    "quantity": item_data.currentStock,
                    "unitCost": item_data.unitCost,
                    "notes": "Initial stock",
                    "actionBy": user_name,
                }
            )
        
        # Fetch the created item
        item_with_count = await prisma.inventoryitem.find_unique(
            where={"id": created_item.id}
        )
        
        # Get transaction count for this item
        transaction_count = await prisma.inventorytransaction.count(
            where={"inventoryItemId": created_item.id}
        )
        count_dict = {"transactions": transaction_count}
        
        item = InventoryItem(
            id=str(item_with_count.id),
            itemCode=str(item_with_count.itemCode),
            name=str(item_with_count.name),
            description=item_with_count.description if item_with_count.description else None,
            category=item_with_count.category if item_with_count.category else None,
            unit=item_with_count.unit if item_with_count.unit else None,
            currentStock=item_with_count.currentStock,
            minStockLevel=item_with_count.minStockLevel if item_with_count.minStockLevel else None,
            maxStockLevel=item_with_count.maxStockLevel if item_with_count.maxStockLevel else None,
            unitCost=item_with_count.unitCost if item_with_count.unitCost else None,
            location=item_with_count.location if item_with_count.location else None,
            supplier=item_with_count.supplier if item_with_count.supplier else None,
            brand=item_with_count.brand if item_with_count.brand else None,
            model=item_with_count.model if item_with_count.model else None,
            sku=item_with_count.sku if item_with_count.sku else None,
            barcode=item_with_count.barcode if item_with_count.barcode else None,
            remarks=item_with_count.remarks if item_with_count.remarks else None,
            createdAt=item_with_count.createdAt,
            updatedAt=item_with_count.updatedAt,
            _count=count_dict
        )
        
        return InventoryItemResponse(item=item)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating inventory item: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to create inventory item")

@router.get("/{item_id}", response_model=InventoryItemResponse)
async def get_inventory_item(
    item_id: str = Path(..., description="Inventory item ID or itemCode"),
    auth: dict = Depends(verify_auth)
):
    """Get a single inventory item by ID or itemCode"""
    try:
        # Check if it's a UUID or itemCode
        is_id_uuid = is_uuid(item_id)
        
        where_clause = {}
        if is_id_uuid:
            where_clause["id"] = item_id
        else:
            where_clause["itemCode"] = item_id
        where_clause["isDeleted"] = False
        
        item_data = await prisma.inventoryitem.find_first(where=where_clause)
        
        if not item_data:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        
        # Get transaction count
        transaction_count = await prisma.inventorytransaction.count(
            where={"inventoryItemId": item_data.id}
        )
        count_dict = {"transactions": transaction_count}
        
        # Get last 50 transactions
        transactions_data = await prisma.inventorytransaction.find_many(
            where={"inventoryItemId": item_data.id},
            order={"transactionDate": "desc"},
            take=50
        )
        
        transactions = []
        for trans in transactions_data:
            # Get related transaction if exists
            related_trans = None
            related_item_info = None
            if trans.relatedTransactionId:
                related_trans_data = await prisma.inventorytransaction.find_unique(
                    where={"id": trans.relatedTransactionId}
                )
                if related_trans_data:
                    # Get related item info
                    related_item = await prisma.inventoryitem.find_unique(
                        where={"id": related_trans_data.inventoryItemId}
                    )
                    if related_item:
                        related_item_info = {
                            "id": str(related_item.id),
                            "itemCode": str(related_item.itemCode),
                            "name": str(related_item.name)
                        }
                        related_trans = {
                            "id": str(related_trans_data.id),
                            "inventoryItemId": str(related_trans_data.inventoryItemId),
                            "transactionType": str(related_trans_data.transactionType),
                            "quantity": related_trans_data.quantity,
                            "unitCost": related_trans_data.unitCost,
                            "reference": related_trans_data.reference,
                            "notes": related_trans_data.notes,
                            "actionBy": str(related_trans_data.actionBy),
                            "transactionDate": related_trans_data.transactionDate,
                            "createdAt": related_trans_data.createdAt,
                            "updatedAt": related_trans_data.updatedAt,
                            "inventoryItem": related_item_info
                        }
            
            transaction = InventoryTransaction(
                id=str(trans.id),
                inventoryItemId=str(trans.inventoryItemId),
                transactionType=str(trans.transactionType),
                quantity=trans.quantity,
                unitCost=trans.unitCost if trans.unitCost else None,
                reference=trans.reference if trans.reference else None,
                notes=trans.notes if trans.notes else None,
                actionBy=str(trans.actionBy),
                transactionDate=trans.transactionDate,
                relatedTransactionId=str(trans.relatedTransactionId) if trans.relatedTransactionId else None,
                createdAt=trans.createdAt,
                updatedAt=trans.updatedAt,
                relatedTransaction=related_trans,
                inventoryItem=related_item_info
            )
            transactions.append(transaction)
        
        item = InventoryItem(
            id=str(item_data.id),
            itemCode=str(item_data.itemCode),
            name=str(item_data.name),
            description=item_data.description if item_data.description else None,
            category=item_data.category if item_data.category else None,
            unit=item_data.unit if item_data.unit else None,
            currentStock=item_data.currentStock,
            minStockLevel=item_data.minStockLevel if item_data.minStockLevel else None,
            maxStockLevel=item_data.maxStockLevel if item_data.maxStockLevel else None,
            unitCost=item_data.unitCost if item_data.unitCost else None,
            location=item_data.location if item_data.location else None,
            supplier=item_data.supplier if item_data.supplier else None,
            brand=item_data.brand if item_data.brand else None,
            model=item_data.model if item_data.model else None,
            sku=item_data.sku if item_data.sku else None,
            barcode=item_data.barcode if item_data.barcode else None,
            remarks=item_data.remarks if item_data.remarks else None,
            createdAt=item_data.createdAt,
            updatedAt=item_data.updatedAt,
            _count=count_dict
        )
        
        # Add transactions to item (for compatibility with frontend)
        item_dict = item.model_dump()
        item_dict["transactions"] = [t.model_dump() for t in transactions]
        
        return InventoryItemResponse(item=item_dict)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching inventory item: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch inventory item")

@router.put("/{item_id}", response_model=InventoryItemResponse)
async def update_inventory_item(
    item_id: str = Path(..., description="Inventory item ID"),
    item_data: InventoryItemUpdate = None,
    auth: dict = Depends(verify_auth)
):
    """Update an existing inventory item"""
    try:
        # Check permission
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Check if item exists
        existing_item = await prisma.inventoryitem.find_unique(where={"id": item_id})
        
        if not existing_item:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        
        # Check if itemCode is being changed and if it already exists
        if item_data.itemCode and item_data.itemCode != existing_item.itemCode:
            duplicate_item = await prisma.inventoryitem.find_unique(
                where={"itemCode": item_data.itemCode}
            )
            
            if duplicate_item:
                raise HTTPException(
                    status_code=400,
                    detail="Item code already exists"
                )
        
        # Prepare update data
        update_data = {}
        if item_data.itemCode is not None:
            update_data["itemCode"] = item_data.itemCode
        if item_data.name is not None:
            update_data["name"] = item_data.name
        if item_data.description is not None:
            update_data["description"] = item_data.description
        if item_data.category is not None:
            update_data["category"] = item_data.category
        if item_data.unit is not None:
            update_data["unit"] = item_data.unit
        if item_data.minStockLevel is not None:
            update_data["minStockLevel"] = item_data.minStockLevel
        if item_data.maxStockLevel is not None:
            update_data["maxStockLevel"] = item_data.maxStockLevel
        if item_data.unitCost is not None:
            update_data["unitCost"] = item_data.unitCost
        if item_data.location is not None:
            update_data["location"] = item_data.location
        if item_data.supplier is not None:
            update_data["supplier"] = item_data.supplier
        if item_data.brand is not None:
            update_data["brand"] = item_data.brand
        if item_data.model is not None:
            update_data["model"] = item_data.model
        if item_data.sku is not None:
            update_data["sku"] = item_data.sku
        if item_data.barcode is not None:
            update_data["barcode"] = item_data.barcode
        if item_data.remarks is not None:
            update_data["remarks"] = item_data.remarks
        
        updated_item = await prisma.inventoryitem.update(
            where={"id": item_id},
            data=update_data
        )
        
        # Get transaction count
        transaction_count = await prisma.inventorytransaction.count(
            where={"inventoryItemId": updated_item.id}
        )
        count_dict = {"transactions": transaction_count}
        
        item = InventoryItem(
            id=str(updated_item.id),
            itemCode=str(updated_item.itemCode),
            name=str(updated_item.name),
            description=updated_item.description if updated_item.description else None,
            category=updated_item.category if updated_item.category else None,
            unit=updated_item.unit if updated_item.unit else None,
            currentStock=updated_item.currentStock,
            minStockLevel=updated_item.minStockLevel if updated_item.minStockLevel else None,
            maxStockLevel=updated_item.maxStockLevel if updated_item.maxStockLevel else None,
            unitCost=updated_item.unitCost if updated_item.unitCost else None,
            location=updated_item.location if updated_item.location else None,
            supplier=updated_item.supplier if updated_item.supplier else None,
            brand=updated_item.brand if updated_item.brand else None,
            model=updated_item.model if updated_item.model else None,
            sku=updated_item.sku if updated_item.sku else None,
            barcode=updated_item.barcode if updated_item.barcode else None,
            remarks=updated_item.remarks if updated_item.remarks else None,
            createdAt=updated_item.createdAt,
            updatedAt=updated_item.updatedAt,
            _count=count_dict
        )
        
        return InventoryItemResponse(item=item)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating inventory item: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to update inventory item")

@router.delete("/{item_id}")
async def delete_inventory_item(
    item_id: str = Path(..., description="Inventory item ID"),
    permanent: bool = Query(False, description="Permanently delete the item"),
    auth: dict = Depends(verify_auth)
):
    """Delete an inventory item (soft delete by default, permanent if specified)"""
    try:
        # Check permission
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        if permanent:
            # Hard delete
            await prisma.inventoryitem.delete(where={"id": item_id})
            return {"success": True, "message": "Item permanently deleted"}
        else:
            # Soft delete
            await prisma.inventoryitem.update(
                where={"id": item_id},
                data={
                    "deletedAt": datetime.now(),
                    "isDeleted": True,
                }
            )
            return {
                "success": True,
                "message": "Item archived. It will be permanently deleted after 30 days."
            }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting inventory item: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to delete inventory item")

@router.post("/{item_id}/restore", response_model=RestoreResponse)
async def restore_inventory_item(
    item_id: str = Path(..., description="Inventory item ID"),
    auth: dict = Depends(verify_auth)
):
    """Restore a soft-deleted inventory item"""
    try:
        # Check permission
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Check if item exists
        item_data = await prisma.inventoryitem.find_unique(where={"id": item_id})
        
        if not item_data:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        
        if not item_data.isDeleted:
            raise HTTPException(status_code=400, detail="Item is not deleted")
        
        # Restore the item
        restored_item = await prisma.inventoryitem.update(
            where={"id": item_id},
            data={
                "isDeleted": False,
                "deletedAt": None,
            }
        )
        
        # Get transaction count
        transaction_count = await prisma.inventorytransaction.count(
            where={"inventoryItemId": restored_item.id}
        )
        count_dict = {"transactions": transaction_count}
        
        item = InventoryItem(
            id=str(restored_item.id),
            itemCode=str(restored_item.itemCode),
            name=str(restored_item.name),
            description=restored_item.description if restored_item.description else None,
            category=restored_item.category if restored_item.category else None,
            unit=restored_item.unit if restored_item.unit else None,
            currentStock=restored_item.currentStock,
            minStockLevel=restored_item.minStockLevel if restored_item.minStockLevel else None,
            maxStockLevel=restored_item.maxStockLevel if restored_item.maxStockLevel else None,
            unitCost=restored_item.unitCost if restored_item.unitCost else None,
            location=restored_item.location if restored_item.location else None,
            supplier=restored_item.supplier if restored_item.supplier else None,
            brand=restored_item.brand if restored_item.brand else None,
            model=restored_item.model if restored_item.model else None,
            sku=restored_item.sku if restored_item.sku else None,
            barcode=restored_item.barcode if restored_item.barcode else None,
            remarks=restored_item.remarks if restored_item.remarks else None,
            createdAt=restored_item.createdAt,
            updatedAt=restored_item.updatedAt,
            _count=count_dict
        )
        
        return RestoreResponse(
            success=True,
            message="Item restored successfully",
            item=item
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error restoring inventory item: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to restore inventory item")

@router.post("/bulk-restore", response_model=BulkRestoreResponse)
async def bulk_restore_inventory_items(
    request: BulkRestoreRequest,
    auth: dict = Depends(verify_auth)
):
    """Bulk restore multiple soft-deleted inventory items"""
    try:
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        if not request.ids or len(request.ids) == 0:
            raise HTTPException(
                status_code=400,
                detail="Invalid request. Expected an array of item IDs."
            )
        
        # Restore all items in a transaction
        async with prisma.tx() as transaction:
            # Update all items to restore them
            result = await transaction.inventoryitem.update_many(
                where={
                    "id": {"in": request.ids},
                    "isDeleted": True  # Only restore items that are actually deleted
                },
                data={
                    "deletedAt": None,
                    "isDeleted": False
                }
            )
        
        return BulkRestoreResponse(
            success=True,
            restoredCount=result,
            message=f"{result} item(s) restored successfully"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error bulk restoring inventory items: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to restore inventory items")

@router.get("/{item_id}/transactions", response_model=InventoryTransactionsResponse)
async def get_inventory_transactions(
    item_id: str = Path(..., description="Inventory item ID or itemCode"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=10000),
    type: Optional[str] = Query(None, description="Filter by transaction type"),
    auth: dict = Depends(verify_auth)
):
    """Get transactions for an inventory item"""
    try:
        # Check if it's a UUID or itemCode
        is_id_uuid = is_uuid(item_id)
        
        # First, get the inventory item to get its ID
        where_clause = {}
        if is_id_uuid:
            where_clause["id"] = item_id
        else:
            where_clause["itemCode"] = item_id
        where_clause["isDeleted"] = False
        
        inventory_item = await prisma.inventoryitem.find_first(
            where=where_clause
        )
        
        if not inventory_item:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        
        # Build where clause for transactions
        transaction_where = {
            "inventoryItemId": inventory_item.id
        }
        
        if type:
            transaction_where["transactionType"] = type
        
        # Get total count and transactions
        total_count = await prisma.inventorytransaction.count(where=transaction_where)
        
        transactions_data = await prisma.inventorytransaction.find_many(
            where=transaction_where,
            order={"transactionDate": "desc"},
            skip=(page - 1) * pageSize,
            take=pageSize,
        )
        
        # Build transactions with related transaction info
        transactions = []
        for trans in transactions_data:
            related_trans = None
            related_item_info = None
            
            if trans.relatedTransactionId:
                related_trans_data = await prisma.inventorytransaction.find_unique(
                    where={"id": trans.relatedTransactionId}
                )
                if related_trans_data:
                    related_item = await prisma.inventoryitem.find_unique(
                        where={"id": related_trans_data.inventoryItemId}
                    )
                    if related_item:
                        related_item_info = {
                            "id": str(related_item.id),
                            "itemCode": str(related_item.itemCode),
                            "name": str(related_item.name)
                        }
                        related_trans = {
                            "id": str(related_trans_data.id),
                            "inventoryItemId": str(related_trans_data.inventoryItemId),
                            "transactionType": str(related_trans_data.transactionType),
                            "quantity": related_trans_data.quantity,
                            "unitCost": related_trans_data.unitCost,
                            "reference": related_trans_data.reference,
                            "notes": related_trans_data.notes,
                            "actionBy": str(related_trans_data.actionBy),
                            "transactionDate": related_trans_data.transactionDate,
                            "createdAt": related_trans_data.createdAt,
                            "updatedAt": related_trans_data.updatedAt,
                            "inventoryItem": related_item_info
                        }
            
            transaction = InventoryTransaction(
                id=str(trans.id),
                inventoryItemId=str(trans.inventoryItemId),
                transactionType=str(trans.transactionType),
                quantity=trans.quantity,
                unitCost=trans.unitCost if trans.unitCost else None,
                reference=trans.reference if trans.reference else None,
                notes=trans.notes if trans.notes else None,
                actionBy=str(trans.actionBy),
                transactionDate=trans.transactionDate,
                relatedTransactionId=str(trans.relatedTransactionId) if trans.relatedTransactionId else None,
                createdAt=trans.createdAt,
                updatedAt=trans.updatedAt,
                relatedTransaction=related_trans,
                inventoryItem=related_item_info
            )
            transactions.append(transaction)
        
        total_pages = (total_count + pageSize - 1) // pageSize if total_count > 0 else 1
        
        pagination = PaginationInfo(
            total=total_count,
            page=page,
            pageSize=pageSize,
            totalPages=total_pages,
            hasNextPage=page < total_pages,
            hasPreviousPage=page > 1
        )
        
        return InventoryTransactionsResponse(transactions=transactions, pagination=pagination)
    
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logging.error(f"Error fetching transactions: {type(e).__name__}: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail="Failed to fetch transactions")

@router.post("/{item_id}/transactions", response_model=InventoryTransactionResponse, status_code=201)
async def create_inventory_transaction(
    item_id: str = Path(..., description="Inventory item ID or itemCode"),
    transaction_data: InventoryTransactionCreate = None,
    auth: dict = Depends(verify_auth)
):
    """Create a new inventory transaction"""
    try:
        # Check permission
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Validation
        if not transaction_data.transactionType or not transaction_data.quantity:
            raise HTTPException(
                status_code=400,
                detail="Transaction type and quantity are required"
            )
        
        if transaction_data.transactionType not in ['IN', 'OUT', 'ADJUSTMENT', 'TRANSFER']:
            raise HTTPException(
                status_code=400,
                detail="Invalid transaction type"
            )
        
        # Check if it's a UUID or itemCode
        is_id_uuid = is_uuid(item_id)
        
        # Get source item
        where_clause = {}
        if is_id_uuid:
            where_clause["id"] = item_id
        else:
            where_clause["itemCode"] = item_id
        where_clause["isDeleted"] = False
        
        source_item = await prisma.inventoryitem.find_first(where=where_clause)
        
        if not source_item:
            raise HTTPException(status_code=404, detail="Source inventory item not found")
        
        # For TRANSFER, destinationItemId is required
        if transaction_data.transactionType == 'TRANSFER' and not transaction_data.destinationItemId:
            raise HTTPException(
                status_code=400,
                detail="Destination item is required for transfer transactions"
            )
        
        # Cannot transfer to the same item
        if transaction_data.transactionType == 'TRANSFER' and transaction_data.destinationItemId == item_id:
            raise HTTPException(
                status_code=400,
                detail="Cannot transfer to the same item"
            )
        
        # Get destination item if TRANSFER
        destination_item = None
        if transaction_data.transactionType == 'TRANSFER':
            is_dest_uuid = is_uuid(transaction_data.destinationItemId)
            dest_where = {}
            if is_dest_uuid:
                dest_where["id"] = transaction_data.destinationItemId
            else:
                dest_where["itemCode"] = transaction_data.destinationItemId
            dest_where["isDeleted"] = False
            
            destination_item = await prisma.inventoryitem.find_first(where=dest_where)
            
            if not destination_item:
                raise HTTPException(status_code=404, detail="Destination inventory item not found")
        
        # Get user info
        user = auth.get("user", {})
        user_name = (
            user.get("user_metadata", {}).get("name") or
            user.get("user_metadata", {}).get("full_name") or
            user.get("email", "").split("@")[0] if user.get("email") else None or
            user.get("email") or
            user.get("id")
        )
        
        # Calculate new stock
        new_source_stock = float(source_item.currentStock)
        qty = float(transaction_data.quantity)
        
        # Calculate weighted average cost for IN transactions
        new_unit_cost = float(source_item.unitCost) if source_item.unitCost else None
        if transaction_data.transactionType == 'IN' and transaction_data.unitCost:
            current_cost = float(source_item.unitCost) if source_item.unitCost else 0
            current_stock = float(source_item.currentStock)
            new_cost = float(transaction_data.unitCost)
            
            if current_stock > 0 and current_cost > 0:
                # Weighted average
                total_value = (current_stock * current_cost) + (qty * new_cost)
                total_stock = current_stock + qty
                new_unit_cost = total_value / total_stock
            else:
                new_unit_cost = new_cost
        
        # Calculate stock changes
        if transaction_data.transactionType == 'IN' or transaction_data.transactionType == 'ADJUSTMENT':
            new_source_stock += qty
        elif transaction_data.transactionType == 'OUT':
            new_source_stock -= qty
            if new_source_stock < 0:
                raise HTTPException(status_code=400, detail="Insufficient stock")
        elif transaction_data.transactionType == 'TRANSFER':
            new_source_stock -= qty
            if new_source_stock < 0:
                raise HTTPException(status_code=400, detail="Insufficient stock")
        
        # Create transaction(s) and update stock in a transaction
        async with prisma.tx() as tx:
            if transaction_data.transactionType == 'TRANSFER':
                # Create OUT transaction for source item
                source_transaction = await tx.inventorytransaction.create(
                    data={
                        "inventoryItemId": source_item.id,
                        "transactionType": "TRANSFER",
                        "quantity": Decimal(str(qty)),
                        "unitCost": Decimal(str(transaction_data.unitCost)) if transaction_data.unitCost else None,
                        "reference": transaction_data.reference,
                        "notes": transaction_data.notes or f"Transfer to {destination_item.name if destination_item else transaction_data.destinationItemId}",
                        "actionBy": user_name,
                    }
                )
                
                # Create IN transaction for destination item
                destination_transaction = await tx.inventorytransaction.create(
                    data={
                        "inventoryItemId": destination_item.id,
                        "transactionType": "IN",
                        "quantity": Decimal(str(qty)),
                        "unitCost": Decimal(str(transaction_data.unitCost)) if transaction_data.unitCost else None,
                        "reference": transaction_data.reference,
                        "notes": transaction_data.notes or f"Transfer from {source_item.name}",
                        "actionBy": user_name,
                        "relatedTransactionId": source_transaction.id,
                    }
                )
                
                # Link source transaction to destination transaction
                await tx.inventorytransaction.update(
                    where={"id": source_transaction.id},
                    data={
                        "relatedTransactionId": destination_transaction.id,
                    }
                )
                
                # Update source item stock
                await tx.inventoryitem.update(
                    where={"id": source_item.id},
                    data={
                        "currentStock": Decimal(str(new_source_stock)),
                    }
                )
                
                # Update destination item stock and calculate weighted average cost
                new_destination_stock = float(destination_item.currentStock) + qty
                new_destination_unit_cost = float(destination_item.unitCost) if destination_item.unitCost else None
                if transaction_data.unitCost:
                    dest_current_cost = float(destination_item.unitCost) if destination_item.unitCost else 0
                    dest_current_stock = float(destination_item.currentStock)
                    transfer_cost = float(transaction_data.unitCost)
                    
                    if dest_current_stock > 0 and dest_current_cost > 0:
                        total_value = (dest_current_stock * dest_current_cost) + (qty * transfer_cost)
                        total_stock = dest_current_stock + qty
                        new_destination_unit_cost = total_value / total_stock
                    else:
                        new_destination_unit_cost = transfer_cost
                
                await tx.inventoryitem.update(
                    where={"id": destination_item.id},
                    data={
                        "currentStock": Decimal(str(new_destination_stock)),
                        "unitCost": Decimal(str(new_destination_unit_cost)) if new_destination_unit_cost else None,
                    }
                )
                
                result_transaction = source_transaction
            else:
                # Create transaction record for non-transfer types
                transaction = await tx.inventorytransaction.create(
                    data={
                        "inventoryItemId": source_item.id,
                        "transactionType": transaction_data.transactionType,
                        "quantity": Decimal(str(qty)),
                        "unitCost": Decimal(str(transaction_data.unitCost)) if transaction_data.unitCost else None,
                        "reference": transaction_data.reference,
                        "notes": transaction_data.notes,
                        "actionBy": user_name,
                    }
                )
                
                # Update stock and unit cost (for IN transactions with cost)
                update_data = {
                    "currentStock": Decimal(str(new_source_stock)),
                }
                if transaction_data.transactionType == 'IN' and new_unit_cost is not None:
                    update_data["unitCost"] = Decimal(str(new_unit_cost))
                
                await tx.inventoryitem.update(
                    where={"id": source_item.id},
                    data=update_data
                )
                
                result_transaction = transaction
        
        # Fetch the created transaction with relations
        created_trans = await prisma.inventorytransaction.find_unique(
            where={"id": result_transaction.id}
        )
        
        related_trans = None
        related_item_info = None
        if created_trans.relatedTransactionId:
            related_trans_data = await prisma.inventorytransaction.find_unique(
                where={"id": created_trans.relatedTransactionId}
            )
            if related_trans_data:
                related_item = await prisma.inventoryitem.find_unique(
                    where={"id": related_trans_data.inventoryItemId}
                )
                if related_item:
                    related_item_info = {
                        "id": str(related_item.id),
                        "itemCode": str(related_item.itemCode),
                        "name": str(related_item.name)
                    }
                    related_trans = {
                        "id": str(related_trans_data.id),
                        "inventoryItemId": str(related_trans_data.inventoryItemId),
                        "transactionType": str(related_trans_data.transactionType),
                        "quantity": related_trans_data.quantity,
                        "unitCost": related_trans_data.unitCost,
                        "reference": related_trans_data.reference,
                        "notes": related_trans_data.notes,
                        "actionBy": str(related_trans_data.actionBy),
                        "transactionDate": related_trans_data.transactionDate,
                        "createdAt": related_trans_data.createdAt,
                        "updatedAt": related_trans_data.updatedAt,
                        "inventoryItem": related_item_info
                    }
        
        transaction = InventoryTransaction(
            id=str(created_trans.id),
            inventoryItemId=str(created_trans.inventoryItemId),
            transactionType=str(created_trans.transactionType),
            quantity=created_trans.quantity,
            unitCost=created_trans.unitCost if created_trans.unitCost else None,
            reference=created_trans.reference if created_trans.reference else None,
            notes=created_trans.notes if created_trans.notes else None,
            actionBy=str(created_trans.actionBy),
            transactionDate=created_trans.transactionDate,
            relatedTransactionId=str(created_trans.relatedTransactionId) if created_trans.relatedTransactionId else None,
            createdAt=created_trans.createdAt,
            updatedAt=created_trans.updatedAt,
            relatedTransaction=related_trans,
            inventoryItem=related_item_info
        )
        
        return InventoryTransactionResponse(transaction=transaction)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating transaction: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to create transaction")

@router.delete("/{item_id}/transactions/bulk-delete", response_model=BulkDeleteTransactionsResponse)
async def bulk_delete_transactions(
    item_id: str = Path(..., description="Inventory item ID or itemCode"),
    request: BulkDeleteTransactionsRequest = Body(...),
    auth: dict = Depends(verify_auth)
):
    """Bulk delete transactions for an inventory item"""
    try:
        # Check permission
        user_id = auth.get("user", {}).get("id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        if not request or not request.transactionIds or len(request.transactionIds) == 0:
            raise HTTPException(
                status_code=400,
                detail="Transaction IDs are required"
            )
        
        # Check if it's a UUID or itemCode
        is_id_uuid = is_uuid(item_id)
        
        # Get the inventory item to verify it exists
        where_clause = {}
        if is_id_uuid:
            where_clause["id"] = item_id
        else:
            where_clause["itemCode"] = item_id
        where_clause["isDeleted"] = False
        
        inventory_item = await prisma.inventoryitem.find_first(
            where=where_clause
        )
        
        if not inventory_item:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        
        # Verify all transactions belong to this inventory item
        transactions = await prisma.inventorytransaction.find_many(
            where={
                "id": {"in": request.transactionIds},
                "inventoryItemId": inventory_item.id,
            }
        )
        
        if len(transactions) != len(request.transactionIds):
            raise HTTPException(
                status_code=400,
                detail="Some transactions not found or do not belong to this inventory item"
            )
        
        # Delete transactions
        deleted_count = await prisma.inventorytransaction.delete_many(
            where={
                "id": {"in": request.transactionIds},
                "inventoryItemId": inventory_item.id,
            }
        )
        
        return BulkDeleteTransactionsResponse(
            success=True,
            deletedCount=deleted_count,
            message=f"Successfully deleted {deleted_count} transaction(s)"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error bulk deleting transactions: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to delete transactions")
