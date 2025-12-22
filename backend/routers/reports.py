"""
Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv
import asyncio

from models.reports import ReportDataResponse, ReportSummary, StatusGroup, CategoryGroup, LocationGroup, SiteGroup, RecentAsset, AuditReportResponse, AuditItem, PaginationInfo
from auth import verify_auth
from database import prisma

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/assets", tags=["reports"])

def format_number(value: Optional[float]) -> str:
    """Format number with commas and 2 decimal places"""
    if value is None or value == 0:
        return '0.00'
    return f"{float(value):,.2f}"

@router.get("/summary", response_model=ReportDataResponse)
async def get_assets_summary(
    status: Optional[str] = Query(None, description="Filter by status"),
    category: Optional[str] = Query(None, description="Filter by category ID"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeAllAssets: Optional[bool] = Query(False, description="Include all assets instead of just 10"),
    auth: dict = Depends(verify_auth)
):
    """Get assets summary report with optional filters"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        # Build where clause
        where_clause: Dict[str, Any] = {
            "isDeleted": False,
        }

        # Apply filters
        if status:
            where_clause["status"] = status

        if category:
            where_clause["categoryId"] = category

        if location:
            where_clause["location"] = location

        if site:
            where_clause["site"] = site

        if department:
            where_clause["department"] = department

        # Date range filter (on purchaseDate or createdAt)
        if startDate or endDate:
            date_filters: List[Dict[str, Any]] = []
            
            purchase_date_filter: Dict[str, Any] = {}
            if startDate:
                purchase_date_filter["gte"] = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
            if endDate:
                purchase_date_filter["lte"] = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
            if purchase_date_filter:
                date_filters.append({"purchaseDate": purchase_date_filter})
            
            created_at_filter: Dict[str, Any] = {}
            if startDate:
                created_at_filter["gte"] = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
            if endDate:
                created_at_filter["lte"] = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
            if created_at_filter:
                date_filters.append({"createdAt": created_at_filter})
            
            if date_filters:
                where_clause["OR"] = date_filters

        # Get total assets count and fetch assets for sum calculation
        # Prisma Client Python doesn't have aggregate, so we fetch and sum in Python
        total_assets, assets_for_sum = await asyncio.gather(
            prisma.assets.count(where=where_clause),
            prisma.assets.find_many(
                where={**where_clause, "cost": {"not": None}}
            )
        )

        # Calculate total value
        total_value = sum(
            float(asset.cost) if asset.cost is not None else 0.0
            for asset in assets_for_sum
        )

        # Get assets by status
        status_groups_raw = await prisma.assets.group_by(
            by=["status"],
            where=where_clause,
            count=True,
            sum={"cost": True}
        )

        by_status = [
            StatusGroup(
                status=row.get("status") or "Unknown",
                count=row.get("_count", {}).get("_all", 0),
                value=float(row.get("_sum", {}).get("cost", 0) or 0)
            )
            for row in status_groups_raw
        ]

        # Get assets by category
        category_groups_raw = await prisma.assets.group_by(
            by=["categoryId"],
            where=where_clause,
            count=True,
            sum={"cost": True}
        )

        # Fetch category names
        category_ids = [row.get("categoryId") for row in category_groups_raw if row.get("categoryId")]
        categories = await prisma.category.find_many(
            where={"id": {"in": category_ids}}
        )

        category_map = {cat.id: cat.name for cat in categories}

        by_category = [
            CategoryGroup(
                categoryId=row.get("categoryId"),
                categoryName=category_map.get(row.get("categoryId"), "Unknown"),
                count=row.get("_count", {}).get("_all", 0),
                value=float(row.get("_sum", {}).get("cost", 0) or 0)
            )
            for row in category_groups_raw
            if row.get("categoryId")
        ]
        by_category.sort(key=lambda x: x.count, reverse=True)

        # Get assets by location
        location_where = {**where_clause, "location": {"not": None}}
        location_groups_raw = await prisma.assets.group_by(
            by=["location"],
            where=location_where,
            count=True
        )

        by_location = [
            LocationGroup(
                location=row.get("location"),
                count=row.get("_count", {}).get("_all", 0)
            )
            for row in location_groups_raw
            if row.get("location")
        ]
        by_location.sort(key=lambda x: x.count, reverse=True)

        # Get assets by site
        site_where = {**where_clause, "site": {"not": None}}
        site_groups_raw = await prisma.assets.group_by(
            by=["site"],
            where=site_where,
            count=True
        )

        by_site = [
            SiteGroup(
                site=row.get("site"),
                count=row.get("_count", {}).get("_all", 0)
            )
            for row in site_groups_raw
            if row.get("site")
        ]
        by_site.sort(key=lambda x: x.count, reverse=True)

        # Get recent assets (last 10 or all if requested)
        recent_assets_raw = await prisma.assets.find_many(
            where=where_clause,
            take=None if includeAllAssets else 10,
            order={"createdAt": "desc"},
            include={"category": True}
        )

        recent_assets = [
            RecentAsset(
                id=asset.id,
                assetTagId=asset.assetTagId,
                description=asset.description,
                status=asset.status,
                cost=float(asset.cost) if asset.cost else None,
                category={"name": asset.category.name} if asset.category else None,
                location=asset.location,
                site=asset.site,
                department=asset.department
            )
            for asset in recent_assets_raw
        ]

        return ReportDataResponse(
            summary=ReportSummary(
                totalAssets=total_assets,
                totalValue=total_value,
                byStatus=by_status,
                byCategory=by_category,
                byLocation=by_location,
                bySite=by_site
            ),
            recentAssets=recent_assets,
            generatedAt=datetime.now().isoformat()
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating summary report: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to generate summary report")

@router.get("/export")
async def export_assets_report(
    format: str = Query("csv", description="Export format: csv or excel"),
    reportType: str = Query("summary", description="Report type: summary, status, or category"),
    status: Optional[str] = Query(None, description="Filter by status"),
    category: Optional[str] = Query(None, description="Filter by category ID"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeAssetList: Optional[bool] = Query(False, description="Include asset list in summary report"),
    auth: dict = Depends(verify_auth)
):
    """Export assets report to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        if format not in ["csv", "excel"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv or excel.")

        # Build where clause
        where_clause: Dict[str, Any] = {
            "isDeleted": False,
        }

        if status:
            where_clause["status"] = status
        if category:
            where_clause["categoryId"] = category
        if location:
            where_clause["location"] = location
        if site:
            where_clause["site"] = site
        if department:
            where_clause["department"] = department
        if startDate or endDate:
            created_at_filter: Dict[str, Any] = {}
            if startDate:
                created_at_filter["gte"] = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
            if endDate:
                created_at_filter["lte"] = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
            if created_at_filter:
                where_clause["createdAt"] = created_at_filter

        # Fetch assets
        assets_raw = await prisma.assets.find_many(
            where=where_clause,
            include={"category": True},
            order={"createdAt": "desc"}
        )

        # Convert to list of dicts for easier processing
        assets = [
            {
                "id": asset.id,
                "assetTagId": asset.assetTagId,
                "description": asset.description,
                "status": asset.status,
                "cost": float(asset.cost) if asset.cost else 0,
                "category": asset.category.name if asset.category else None,
                "purchasedFrom": asset.purchasedFrom,
                "purchaseDate": asset.purchaseDate.isoformat().split('T')[0] if asset.purchaseDate else None,
                "brand": asset.brand,
                "model": asset.model,
                "serialNo": asset.serialNo,
                "additionalInformation": asset.additionalInformation,
                "xeroAssetNo": asset.xeroAssetNo,
                "owner": asset.owner,
                "subCategory": asset.subCategory,
                "pbiNumber": asset.pbiNumber,
                "issuedTo": asset.issuedTo,
                "poNumber": asset.poNumber,
                "paymentVoucherNumber": asset.paymentVoucherNumber,
                "assetType": asset.assetType,
                "deliveryDate": asset.deliveryDate.isoformat().split('T')[0] if asset.deliveryDate else None,
                "unaccountedInventory": asset.unaccountedInventory,
                "remarks": asset.remarks,
                "qr": asset.qr,
                "oldAssetTag": asset.oldAssetTag,
                "depreciableAsset": asset.depreciableAsset,
                "depreciableCost": float(asset.depreciableCost) if asset.depreciableCost else None,
                "salvageValue": float(asset.salvageValue) if asset.salvageValue else None,
                "assetLifeMonths": asset.assetLifeMonths,
                "depreciationMethod": asset.depreciationMethod,
                "dateAcquired": asset.dateAcquired.isoformat().split('T')[0] if asset.dateAcquired else None,
                "department": asset.department,
                "site": asset.site,
                "location": asset.location,
                "createdAt": asset.createdAt.isoformat().split('T')[0],
            }
            for asset in assets_raw
        ]

        # Prepare export data based on report type
        export_data: Any = []
        report_type_label = "Summary" if reportType == "summary" else ("Status" if reportType == "status" else "Category")

        if reportType == "status":
            # Group by status
            status_groups: Dict[str, Dict[str, Any]] = {}
            for asset in assets:
                status_key = asset["status"] or "Unknown"
                if status_key not in status_groups:
                    status_groups[status_key] = {"count": 0, "totalValue": 0, "assets": []}
                status_groups[status_key]["count"] += 1
                status_groups[status_key]["totalValue"] += asset["cost"]
                status_groups[status_key]["assets"].append(asset)

            export_data = [
                {
                    "Status": status,
                    "Asset Count": str(data["count"]),
                    "Total Value": format_number(data["totalValue"]),
                    "Average Value": format_number(data["totalValue"] / data["count"] if data["count"] > 0 else 0),
                    "Percentage of Total": f"{(data['count'] / len(assets) * 100):.1f}%" if len(assets) > 0 else "0%",
                }
                for status, data in sorted(status_groups.items(), key=lambda x: x[1]["count"], reverse=True)
            ]

        elif reportType == "category":
            # Group by category
            category_groups: Dict[str, Dict[str, Any]] = {}
            for asset in assets:
                category_key = asset["category"] or "Uncategorized"
                if category_key not in category_groups:
                    category_groups[category_key] = {"count": 0, "totalValue": 0, "assets": []}
                category_groups[category_key]["count"] += 1
                category_groups[category_key]["totalValue"] += asset["cost"]
                category_groups[category_key]["assets"].append(asset)

            export_data = [
                {
                    "Category": category,
                    "Asset Count": str(data["count"]),
                    "Total Value": format_number(data["totalValue"]),
                    "Average Value": format_number(data["totalValue"] / data["count"] if data["count"] > 0 else 0),
                    "Percentage of Total": f"{(data['count'] / len(assets) * 100):.1f}%" if len(assets) > 0 else "0%",
                }
                for category, data in sorted(category_groups.items(), key=lambda x: x[1]["count"], reverse=True)
            ]

        else:  # summary
            # Calculate totals and groups
            total_assets = len(assets)
            total_value = sum(asset["cost"] for asset in assets)

            # Group by status
            status_groups: Dict[str, Dict[str, Any]] = {}
            for asset in assets:
                status_key = asset["status"] or "Unknown"
                if status_key not in status_groups:
                    status_groups[status_key] = {"count": 0, "totalValue": 0}
                status_groups[status_key]["count"] += 1
                status_groups[status_key]["totalValue"] += asset["cost"]

            # Group by category
            category_groups: Dict[str, Dict[str, Any]] = {}
            for asset in assets:
                category_key = asset["category"] or "Uncategorized"
                if category_key not in category_groups:
                    category_groups[category_key] = {"count": 0, "totalValue": 0}
                category_groups[category_key]["count"] += 1
                category_groups[category_key]["totalValue"] += asset["cost"]

            # Build summary statistics data
            summary_data = [
                {
                    "Metric": "Total Assets",
                    "Value": str(total_assets),
                    "Total Value": format_number(total_value),
                    "Average Value": format_number(total_value / total_assets if total_assets > 0 else 0),
                    "Percentage": "100%",
                },
                {
                    "Metric": "---",
                    "Value": "---",
                    "Total Value": "---",
                    "Average Value": "---",
                    "Percentage": "---",
                },
                {
                    "Metric": "ASSETS BY STATUS",
                    "Value": "",
                    "Total Value": "",
                    "Average Value": "",
                    "Percentage": "",
                },
                *[
                    {
                        "Metric": f"Status: {status}",
                        "Value": str(data["count"]),
                        "Total Value": format_number(data["totalValue"]),
                        "Average Value": format_number(data["totalValue"] / data["count"] if data["count"] > 0 else 0),
                        "Percentage": f"{(data['count'] / total_assets * 100):.1f}%" if total_assets > 0 else "0%",
                    }
                    for status, data in sorted(status_groups.items(), key=lambda x: x[1]["count"], reverse=True)
                ],
                {
                    "Metric": "---",
                    "Value": "---",
                    "Total Value": "---",
                    "Average Value": "---",
                    "Percentage": "---",
                },
                {
                    "Metric": "ASSETS BY CATEGORY",
                    "Value": "",
                    "Total Value": "",
                    "Average Value": "",
                    "Percentage": "",
                },
                *[
                    {
                        "Metric": f"Category: {category}",
                        "Value": str(data["count"]),
                        "Total Value": format_number(data["totalValue"]),
                        "Average Value": format_number(data["totalValue"] / data["count"] if data["count"] > 0 else 0),
                        "Percentage": f"{(data['count'] / total_assets * 100):.1f}%" if total_assets > 0 else "0%",
                    }
                    for category, data in sorted(category_groups.items(), key=lambda x: x[1]["count"], reverse=True)
                ],
            ]

            if includeAssetList:
                # Full asset export with all fields + summary
                asset_list_data = [
                    {
                        "Asset Tag ID": asset.get("assetTagId", ""),
                        "Description": asset.get("description", ""),
                        "Purchased From": asset.get("purchasedFrom", ""),
                        "Purchase Date": asset.get("purchaseDate", ""),
                        "Brand": asset.get("brand", ""),
                        "Cost": format_number(asset.get("cost")),
                        "Model": asset.get("model", ""),
                        "Serial No": asset.get("serialNo", ""),
                        "Additional Information": asset.get("additionalInformation", ""),
                        "Xero Asset No.": asset.get("xeroAssetNo", ""),
                        "Owner": asset.get("owner", ""),
                        "Sub Category": asset.get("subCategory", ""),
                        "PBI Number": asset.get("pbiNumber", ""),
                        "Status": asset.get("status", ""),
                        "Issued To": asset.get("issuedTo", ""),
                        "PO Number": asset.get("poNumber", ""),
                        "Payment Voucher Number": asset.get("paymentVoucherNumber", ""),
                        "Asset Type": asset.get("assetType", ""),
                        "Delivery Date": asset.get("deliveryDate", ""),
                        "Unaccounted Inventory": asset.get("unaccountedInventory", ""),
                        "Remarks": asset.get("remarks", ""),
                        "QR": asset.get("qr", ""),
                        "Old Asset Tag": asset.get("oldAssetTag", ""),
                        "Depreciable Asset": asset.get("depreciableAsset", ""),
                        "Depreciable Cost": format_number(asset.get("depreciableCost")),
                        "Salvage Value": format_number(asset.get("salvageValue")),
                        "Asset Life (months)": str(asset.get("assetLifeMonths", "")) if asset.get("assetLifeMonths") else "",
                        "Depreciation Method": asset.get("depreciationMethod", ""),
                        "Date Acquired": asset.get("dateAcquired", ""),
                        "Category": asset.get("category", ""),
                        "Department": asset.get("department", ""),
                        "Site": asset.get("site", ""),
                        "Location": asset.get("location", ""),
                        "Created At": asset.get("createdAt", ""),
                    }
                    for asset in assets
                ]

                export_data = {"summary": summary_data, "assetList": asset_list_data}
            else:
                export_data = summary_data

        # Generate file
        filename = f"asset-report-{reportType}-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            if isinstance(export_data, dict):
                # Summary report with asset list
                summary_data = export_data["summary"]
                asset_list_data = export_data["assetList"]

                # Create CSV content
                output = io.StringIO()
                writer = csv.writer(output)

                # Summary section
                writer.writerow(["=== SUMMARY STATISTICS ==="])
                if summary_data:
                    summary_headers = list(summary_data[0].keys())
                    writer.writerow(summary_headers)
                    for row in summary_data:
                        writer.writerow([row.get(header, "") for header in summary_headers])
                
                writer.writerow([])
                writer.writerow(["=== ASSET LIST ==="])
                
                # Asset list section
                if asset_list_data:
                    asset_headers = list(asset_list_data[0].keys())
                    writer.writerow(asset_headers)
                    for row in asset_list_data:
                        writer.writerow([row.get(header, "") for header in asset_headers])

                csv_content = output.getvalue()
                output.close()
            else:
                # Single section export
                if not export_data:
                    raise HTTPException(status_code=400, detail="No data to export")

                output = io.StringIO()
                writer = csv.writer(output)

                headers = list(export_data[0].keys())
                writer.writerow(headers)
                for row in export_data:
                    writer.writerow([row.get(header, "") for header in headers])

                csv_content = output.getvalue()
                output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # excel
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            if isinstance(export_data, dict):
                # Summary report with asset list - multiple sheets
                summary_data = export_data["summary"]
                asset_list_data = export_data["assetList"]

                # Summary sheet
                if summary_data:
                    summary_ws = wb.create_sheet("Summary")
                    summary_headers = list(summary_data[0].keys())
                    summary_ws.append(summary_headers)
                    for row in summary_data:
                        summary_ws.append([row.get(header, "") for header in summary_headers])

                # Assets by Status sheet
                status_data = [row for row in summary_data if row.get("Metric", "").startswith("Status:")]
                if status_data:
                    status_ws = wb.create_sheet("By Status")
                    status_headers = list(status_data[0].keys())
                    status_ws.append(status_headers)
                    for row in status_data:
                        status_ws.append([row.get(header, "") for header in status_headers])

                # Assets by Category sheet
                category_data = [row for row in summary_data if row.get("Metric", "").startswith("Category:")]
                if category_data:
                    category_ws = wb.create_sheet("By Category")
                    category_headers = list(category_data[0].keys())
                    category_ws.append(category_headers)
                    for row in category_data:
                        category_ws.append([row.get(header, "") for header in category_headers])

                # Asset List sheet
                if asset_list_data:
                    asset_list_ws = wb.create_sheet("Asset List")
                    asset_headers = list(asset_list_data[0].keys())
                    asset_list_ws.append(asset_headers)
                    for row in asset_list_data:
                        asset_list_ws.append([row.get(header, "") for header in asset_headers])
            else:
                # Single sheet export
                if not export_data:
                    raise HTTPException(status_code=400, detail="No data to export")

                ws = wb.create_sheet(f"Assets by {report_type_label}")
                headers = list(export_data[0].keys())
                ws.append(headers)
                for row in export_data:
                    ws.append([row.get(header, "") for header in headers])

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting report: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export report")

