"""
Audit Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv

from models.reports import AuditReportResponse, AuditItem, PaginationInfo
from auth import verify_auth
from database import prisma
from utils.pdf_generator import ReportPDF, PDF_AVAILABLE

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/audit", tags=["reports"])

async def check_permission(user_id: str, permission: str) -> bool:
    """Check if user has a specific permission. Admins have all permissions."""
    try:
        asset_user = await prisma.assetuser.find_unique(
            where={"userId": user_id}
        )
        if not asset_user or not asset_user.isActive:
            return False
        
        # Admins have all permissions
        if asset_user.role == "admin":
            return True
        
        return getattr(asset_user, permission, False)
    except Exception:
        return False

@router.get("", response_model=AuditReportResponse)
async def get_audit_reports(
    category: Optional[str] = Query(None, description="Filter by category name"),
    auditType: Optional[str] = Query(None, description="Filter by audit type"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    auditor: Optional[str] = Query(None, description="Filter by auditor name"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get audit reports with optional filters and pagination"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        skip = (page - 1) * pageSize

        # Build where clause for audit history
        where_clause: Dict[str, Any] = {
            "asset": {
                "isDeleted": False,
            }
        }

        # Category filter (by category name)
        if category:
            where_clause["asset"] = {
                **where_clause["asset"],
                "category": {
                    "name": category
                }
            }

        # Audit type filter
        if auditType:
            where_clause["auditType"] = auditType

        # Location filter
        if location:
            where_clause["asset"] = {
                **where_clause["asset"],
                "location": location
            }

        # Site filter
        if site:
            where_clause["asset"] = {
                **where_clause["asset"],
                "site": site
            }

        # Auditor filter (case-insensitive search)
        if auditor:
            where_clause["auditor"] = {
                "contains": auditor,
                "mode": "insensitive"
            }

        # Date range filter
        if startDate or endDate:
            audit_date_filter: Dict[str, Any] = {}
            if startDate:
                audit_date_filter["gte"] = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
            if endDate:
                audit_date_filter["lte"] = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
            if audit_date_filter:
                where_clause["auditDate"] = audit_date_filter

        # Get total count
        total = await prisma.assetsaudithistory.count(where=where_clause)

        # Get paginated audit records
        audits_raw = await prisma.assetsaudithistory.find_many(
            where=where_clause,
            include={"asset": {"include": {"category": True, "subCategory": True}}},
            order={"auditDate": "desc"},
            skip=skip,
            take=pageSize
        )

        # Format the response
        formatted_audits = [
            AuditItem(
                id=audit.id,
                assetTagId=audit.asset.assetTagId,
                category=audit.asset.category.name if audit.asset.category else None,
                subCategory=audit.asset.subCategory.name if audit.asset.subCategory else None,
                auditName=audit.auditType,
                auditedToSite=audit.asset.site,
                auditedToLocation=audit.asset.location,
                lastAuditDate=audit.auditDate.isoformat(),
                auditBy=audit.auditor
            )
            for audit in audits_raw
        ]

        total_pages = (total + pageSize - 1) // pageSize if total > 0 else 0

        return AuditReportResponse(
            audits=formatted_audits,
            pagination=PaginationInfo(
                total=total,
                page=page,
                pageSize=pageSize,
                totalPages=total_pages,
                hasNextPage=page < total_pages,
                hasPreviousPage=page > 1,
            )
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching audit reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch audit reports")

@router.get("/export")
async def export_audit_reports(
    format: str = Query("csv", description="Export format: csv or excel"),
    category: Optional[str] = Query(None, description="Filter by category name"),
    auditType: Optional[str] = Query(None, description="Filter by audit type"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    auditor: Optional[str] = Query(None, description="Filter by auditor name"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeAuditList: Optional[bool] = Query(False, description="Include audit list in export"),
    auth: dict = Depends(verify_auth)
):
    """Export audit reports to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Check permission - user must have canManageReports to export reports
        has_permission = await check_permission(user_id, "canManageReports")
        if not has_permission:
            raise HTTPException(
                status_code=403,
                detail="You do not have permission to export reports"
            )

        if format not in ["csv", "excel", "pdf"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv, excel, or pdf.")
        
        if format == "pdf" and not PDF_AVAILABLE:
            raise HTTPException(status_code=500, detail="PDF export not available - fpdf2 not installed")

        # Build where clause (same as main route)
        where_clause: Dict[str, Any] = {
            "asset": {
                "isDeleted": False,
            }
        }

        if category:
            where_clause["asset"] = {
                **where_clause["asset"],
                "category": {
                    "name": category
                }
            }

        if auditType:
            where_clause["auditType"] = auditType

        if location:
            where_clause["asset"] = {
                **where_clause["asset"],
                "location": location
            }

        if site:
            where_clause["asset"] = {
                **where_clause["asset"],
                "site": site
            }

        if auditor:
            where_clause["auditor"] = {
                "contains": auditor,
                "mode": "insensitive"
            }

        if startDate or endDate:
            audit_date_filter: Dict[str, Any] = {}
            if startDate:
                audit_date_filter["gte"] = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
            if endDate:
                audit_date_filter["lte"] = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
            if audit_date_filter:
                where_clause["auditDate"] = audit_date_filter

        # Fetch all audits (or limited if not including list)
        page_size = 10000 if includeAuditList else 1
        audits_raw = await prisma.assetsaudithistory.find_many(
            where=where_clause,
            include={"asset": {"include": {"category": True, "subCategory": True}}},
            order={"auditDate": "desc"},
            take=page_size
        )

        # Calculate summary statistics
        total_audits = len(audits_raw)
        unique_audit_types = set(audit.auditType for audit in audits_raw)
        audits_by_type = [
            {
                "auditType": audit_type,
                "count": sum(1 for audit in audits_raw if audit.auditType == audit_type)
            }
            for audit_type in unique_audit_types
        ]

        # Format audit data
        formatted_audits = [
            {
                "Asset Tag ID": audit.asset.assetTagId,
                "Category": audit.asset.category.name if audit.asset.category else "N/A",
                "Sub-Category": audit.asset.subCategory.name if audit.asset.subCategory else "N/A",
                "Audit Type": audit.auditType,
                "Audited to Site": audit.asset.site or "N/A",
                "Audited to Location": audit.asset.location or "N/A",
                "Last Audit Date": audit.auditDate.isoformat().split('T')[0],
                "Audit By": audit.auditor or "N/A",
            }
            for audit in audits_raw
        ]

        filename = f"audit-report-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)

            if includeAuditList:
                # Include summary and audit list
                writer.writerow(["AUDIT REPORT SUMMARY"])
                writer.writerow(["Total Audits", total_audits])
                writer.writerow(["Unique Audit Types", len(unique_audit_types)])
                writer.writerow([])
                
                writer.writerow(["AUDITS BY TYPE"])
                writer.writerow(["Audit Type", "Count"])
                for item in audits_by_type:
                    writer.writerow([item["auditType"], item["count"]])
                writer.writerow([])
                
                writer.writerow(["AUDIT RECORDS"])
                if formatted_audits:
                    headers = list(formatted_audits[0].keys())
                    writer.writerow(headers)
                    for audit in formatted_audits:
                        writer.writerow([audit.get(header, "") for header in headers])
            else:
                # Summary only
                writer.writerow(["AUDIT REPORT SUMMARY"])
                writer.writerow(["Total Audits", total_audits])
                writer.writerow(["Unique Audit Types", len(unique_audit_types)])
                writer.writerow([])
                writer.writerow(["AUDITS BY TYPE"])
                writer.writerow(["Audit Type", "Count"])
                for item in audits_by_type:
                    writer.writerow([item["auditType"], item["count"]])

            csv_content = output.getvalue()
            output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        elif format == "excel":
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Summary sheet
            summary_data = [
                ["AUDIT REPORT SUMMARY"],
                ["Total Audits", total_audits],
                ["Unique Audit Types", len(unique_audit_types)],
                [],
                ["AUDITS BY TYPE"],
                ["Audit Type", "Count"],
                *[[item["auditType"], item["count"]] for item in audits_by_type],
            ]
            summary_ws = wb.create_sheet("Summary")
            for row in summary_data:
                summary_ws.append(row)

            if includeAuditList:
                # Audit list sheet
                if formatted_audits:
                    audit_ws = wb.create_sheet("Audit List")
                    headers = list(formatted_audits[0].keys())
                    audit_ws.append(headers)
                    for audit in formatted_audits:
                        audit_ws.append([audit.get(header, "") for header in headers])

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # pdf
            pdf = ReportPDF("Audit Report", "Audit")
            pdf.add_page()

            pdf.add_section_title("Summary Statistics")
            summary_rows = [
                ["Total Audits", str(total_audits)],
                ["Unique Audit Types", str(len(unique_audit_types))],
            ]
            # Add audits by type
            for item in audits_by_type:
                summary_rows.append([f"Type: {item['auditType']}", str(item['count'])])
            
            if summary_rows:
                headers = ["Metric", "Value"]
                pdf.add_table(headers, summary_rows)
            
            pdf.ln(10)

            if formatted_audits and includeAuditList:
                pdf.add_section_title(f"Audit List ({len(formatted_audits)} audits)")
                simplified_headers = ["Asset Tag ID", "Category", "Sub-Category", "Audit Type", "Audited to Site", "Audited to Location", "Last Audit Date", "Audit By"]
                simplified_rows = [
                    [
                        str(a.get("Asset Tag ID", "")),
                        str(a.get("Category", "")),
                        str(a.get("Sub-Category", "")),
                        str(a.get("Audit Type", "")),
                        str(a.get("Audited to Site", "")),
                        str(a.get("Audited to Location", "")),
                        str(a.get("Last Audit Date", "")),
                        str(a.get("Audit By", "")),
                    ]
                    for a in formatted_audits
                ]
                pdf.add_table(simplified_headers, simplified_rows)

            pdf_content = bytes(pdf.output())
            filename += ".pdf"
            return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting audit reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export audit reports")

