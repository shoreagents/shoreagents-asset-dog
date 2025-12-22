"""
Checkout Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv

from models.reports import CheckoutReportResponse, CheckoutItem, CheckoutSummary, EmployeeGroup, DepartmentGroup, PaginationInfo
from auth import verify_auth
from database import prisma

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/checkout", tags=["reports"])

def format_number(value: Optional[float]) -> str:
    """Format number with commas and 2 decimal places"""
    if value is None or value == 0:
        return '0.00'
    return f"{float(value):,.2f}"

async def _fetch_checkout_data(
    employeeId: Optional[str] = Query(None, description="Filter by employee ID"),
    assetTagId: Optional[str] = Query(None, description="Filter by asset tag ID"),
    dueDate: Optional[str] = Query(None, description="Filter by due date (YYYY-MM-DD)"),
    isOverdue: Optional[bool] = Query(None, description="Filter by overdue status"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get checkout reports with optional filters and pagination"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        skip = (page - 1) * pageSize

        # Build where clause for checkouts
        where_clause: Dict[str, Any] = {}

        # Date range filter
        if startDate or endDate:
            checkout_date_filter: Dict[str, Any] = {}
            if startDate:
                checkout_date_filter["gte"] = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
            if endDate:
                checkout_date_filter["lte"] = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
            if checkout_date_filter:
                where_clause["checkoutDate"] = checkout_date_filter

        # Employee filter
        if employeeId:
            where_clause["employeeUserId"] = employeeId

        # Department filter (through employee)
        if department:
            where_clause["employeeUser"] = {
                "department": department
            }

        # Get all checkouts (including active and historical)
        checkouts_raw = await prisma.assetscheckout.find_many(
            where=where_clause,
            include={
                "asset": {
                    "include": {
                        "category": True,
                        "subCategory": True
                    }
                },
                "employeeUser": True,
                "checkins": True
            },
            order={"checkoutDate": "desc"}
        )
        
        # Sort checkins by date descending and take only the first one for each checkout
        for checkout in checkouts_raw:
            if checkout.checkins:
                checkout.checkins = sorted(
                    checkout.checkins,
                    key=lambda c: c.checkinDate if c.checkinDate else datetime.min,
                    reverse=True
                )[:1]

        # Filter active checkouts (those without checkin)
        active_checkouts = [c for c in checkouts_raw if len(c.checkins) == 0]

        # Filter overdue checkouts
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        overdue_checkouts = [
            c for c in active_checkouts
            if c.expectedReturnDate and datetime.fromisoformat(c.expectedReturnDate.isoformat()).replace(hour=0, minute=0, second=0, microsecond=0) < today
        ]

        # Group by employee
        by_employee_map: Dict[str, Dict[str, Any]] = {}
        for checkout in active_checkouts:
            employee_id = checkout.employeeUserId or 'unknown'
            employee_name = checkout.employeeUser.name if checkout.employeeUser else 'Unknown'
            employee_email = checkout.employeeUser.email if checkout.employeeUser else ''
            employee_dept = checkout.employeeUser.department if checkout.employeeUser else None

            if employee_id not in by_employee_map:
                by_employee_map[employee_id] = {
                    "employeeId": employee_id,
                    "employeeName": employee_name,
                    "employeeEmail": employee_email,
                    "department": employee_dept,
                    "count": 0,
                    "overdueCount": 0,
                }

            by_employee_map[employee_id]["count"] += 1

            # Check if overdue
            if checkout.expectedReturnDate:
                expected_return = datetime.fromisoformat(checkout.expectedReturnDate.isoformat()).replace(hour=0, minute=0, second=0, microsecond=0)
                if expected_return < today:
                    by_employee_map[employee_id]["overdueCount"] += 1

        by_employee = [EmployeeGroup(**data) for data in by_employee_map.values()]

        # Group by department
        by_department_map: Dict[str, Dict[str, Any]] = {}
        for checkout in active_checkouts:
            dept = checkout.employeeUser.department if checkout.employeeUser and checkout.employeeUser.department else 'Unassigned'
            employee_id = checkout.employeeUserId

            if dept not in by_department_map:
                by_department_map[dept] = {
                    "department": dept,
                    "count": 0,
                    "overdueCount": 0,
                    "employees": set(),
                }

            by_department_map[dept]["count"] += 1
            if employee_id:
                by_department_map[dept]["employees"].add(employee_id)

            # Check if overdue
            if checkout.expectedReturnDate:
                expected_return = datetime.fromisoformat(checkout.expectedReturnDate.isoformat()).replace(hour=0, minute=0, second=0, microsecond=0)
                if expected_return < today:
                    by_department_map[dept]["overdueCount"] += 1

        by_department = [
            DepartmentGroup(
                department=data["department"],
                count=data["count"],
                overdueCount=data["overdueCount"],
                employeeCount=len(data["employees"])
            )
            for data in by_department_map.values()
        ]

        # Apply filters to active checkouts
        filtered_checkouts = active_checkouts

        # Apply overdue filter
        if isOverdue:
            filtered_checkouts = overdue_checkouts

        # Apply department filter
        if department:
            filtered_checkouts = [
                c for c in filtered_checkouts
                if c.employeeUser and c.employeeUser.department == department
            ]

        # Apply asset tag filter
        if assetTagId:
            asset_tag_lower = assetTagId.lower()
            filtered_checkouts = [
                c for c in filtered_checkouts
                if c.asset.assetTagId.lower().find(asset_tag_lower) != -1
            ]

        # Apply due date filter
        if dueDate:
            due_date_obj = datetime.fromisoformat(dueDate.replace('Z', '+00:00')).replace(hour=0, minute=0, second=0, microsecond=0)
            filtered_checkouts = [
                c for c in filtered_checkouts
                if c.expectedReturnDate and datetime.fromisoformat(c.expectedReturnDate.isoformat()).replace(hour=0, minute=0, second=0, microsecond=0) == due_date_obj
            ]

        # Apply location filter
        if location:
            filtered_checkouts = [
                c for c in filtered_checkouts
                if c.asset.location == location
            ]

        # Apply site filter
        if site:
            filtered_checkouts = [
                c for c in filtered_checkouts
                if c.asset.site == site
            ]

        # Calculate pagination
        total = len(filtered_checkouts)
        total_pages = (total + pageSize - 1) // pageSize if total > 0 else 0

        # Apply pagination
        paginated_checkouts = filtered_checkouts[skip:skip + pageSize]

        # Format checkouts
        formatted_checkouts = []
        for checkout in paginated_checkouts:
            expected_return_date = None
            if checkout.expectedReturnDate:
                expected_return_date = checkout.expectedReturnDate.isoformat().split('T')[0]

            return_date = None
            if checkout.checkins and len(checkout.checkins) > 0 and checkout.checkins[0].checkinDate:
                return_date = checkout.checkins[0].checkinDate.isoformat().split('T')[0]

            is_overdue = False
            if checkout.expectedReturnDate:
                expected_return = datetime.fromisoformat(checkout.expectedReturnDate.isoformat()).replace(hour=0, minute=0, second=0, microsecond=0)
                is_overdue = expected_return < today

            formatted_checkouts.append(
                CheckoutItem(
                    id=checkout.id,
                    assetId=checkout.assetId,
                    assetTagId=checkout.asset.assetTagId,
                    assetDescription=checkout.asset.description,
                    assetStatus=checkout.asset.status,
                    assetCost=float(checkout.asset.cost) if checkout.asset.cost else None,
                    category=checkout.asset.category.name if checkout.asset.category else None,
                    subCategory=checkout.asset.subCategory.name if checkout.asset.subCategory else None,
                    checkoutDate=checkout.checkoutDate.isoformat().split('T')[0],
                    expectedReturnDate=expected_return_date,
                    returnDate=return_date,
                    isOverdue=is_overdue,
                    employeeId=checkout.employeeUserId,
                    employeeName=checkout.employeeUser.name if checkout.employeeUser else 'Unknown',
                    employeeEmail=checkout.employeeUser.email if checkout.employeeUser else '',
                    employeeDepartment=checkout.employeeUser.department if checkout.employeeUser else None,
                    location=checkout.asset.location,
                    site=checkout.asset.site,
                )
            )

        return {
            "summary": CheckoutSummary(
                totalActive=len(active_checkouts),
                totalOverdue=len(overdue_checkouts),
                totalHistorical=len(checkouts_raw) - len(active_checkouts),
                byEmployee=by_employee,
                byDepartment=by_department,
            ),
            "checkouts": formatted_checkouts,
            "all_checkouts": filtered_checkouts,  # Return all filtered checkouts for export
            "pagination": PaginationInfo(
                total=total,
                page=page,
                pageSize=pageSize,
                totalPages=total_pages,
                hasNextPage=page < total_pages,
                hasPreviousPage=page > 1,
            )
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching checkout reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch checkout reports")

@router.get("", response_model=CheckoutReportResponse)
async def get_checkout_reports(
    employeeId: Optional[str] = Query(None, description="Filter by employee ID"),
    assetTagId: Optional[str] = Query(None, description="Filter by asset tag ID"),
    dueDate: Optional[str] = Query(None, description="Filter by due date (YYYY-MM-DD)"),
    isOverdue: Optional[bool] = Query(None, description="Filter by overdue status"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get checkout reports with optional filters and pagination"""
    data = await _fetch_checkout_data(
        employeeId=employeeId,
        assetTagId=assetTagId,
        dueDate=dueDate,
        isOverdue=isOverdue,
        location=location,
        site=site,
        department=department,
        startDate=startDate,
        endDate=endDate,
        page=page,
        pageSize=pageSize,
        auth=auth
    )
    
    return CheckoutReportResponse(
        summary=data["summary"],
        checkouts=data["checkouts"],
        generatedAt=datetime.now().isoformat(),
        pagination=data["pagination"]
    )

@router.get("/export")
async def export_checkout_reports(
    format: str = Query("csv", description="Export format: csv or excel"),
    employeeId: Optional[str] = Query(None, description="Filter by employee ID"),
    assetTagId: Optional[str] = Query(None, description="Filter by asset tag ID"),
    dueDate: Optional[str] = Query(None, description="Filter by due date (YYYY-MM-DD)"),
    isOverdue: Optional[bool] = Query(None, description="Filter by overdue status"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeCheckoutList: Optional[bool] = Query(False, description="Include checkout list in export"),
    auth: dict = Depends(verify_auth)
):
    """Export checkout reports to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        if format not in ["csv", "excel"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv or excel.")

        # Fetch checkout report data (same logic as main endpoint)
        # We'll reuse the helper function but fetch all data if includeCheckoutList is True
        page_size = 10000 if includeCheckoutList else 1
        data = await _fetch_checkout_data(
            employeeId=employeeId,
            assetTagId=assetTagId,
            dueDate=dueDate,
            isOverdue=isOverdue,
            location=location,
            site=site,
            department=department,
            startDate=startDate,
            endDate=endDate,
            page=1,
            pageSize=page_size,
            auth=auth
        )

        summary = data["summary"]
        
        # For export, we need to format all checkouts, not just paginated ones
        all_filtered_checkouts = data["all_checkouts"]
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Format all checkouts for export
        all_formatted_checkouts = []
        for checkout in all_filtered_checkouts:
            expected_return_date = None
            if checkout.expectedReturnDate:
                expected_return_date = checkout.expectedReturnDate.isoformat().split('T')[0]

            return_date = None
            if checkout.checkins and len(checkout.checkins) > 0 and checkout.checkins[0].checkinDate:
                return_date = checkout.checkins[0].checkinDate.isoformat().split('T')[0]

            is_overdue = False
            if checkout.expectedReturnDate:
                expected_return = datetime.fromisoformat(checkout.expectedReturnDate.isoformat()).replace(hour=0, minute=0, second=0, microsecond=0)
                is_overdue = expected_return < today

            all_formatted_checkouts.append(
                CheckoutItem(
                    id=checkout.id,
                    assetId=checkout.assetId,
                    assetTagId=checkout.asset.assetTagId,
                    assetDescription=checkout.asset.description,
                    assetStatus=checkout.asset.status,
                    assetCost=float(checkout.asset.cost) if checkout.asset.cost else None,
                    category=checkout.asset.category.name if checkout.asset.category else None,
                    subCategory=checkout.asset.subCategory.name if checkout.asset.subCategory else None,
                    checkoutDate=checkout.checkoutDate.isoformat().split('T')[0],
                    expectedReturnDate=expected_return_date,
                    returnDate=return_date,
                    isOverdue=is_overdue,
                    employeeId=checkout.employeeUserId,
                    employeeName=checkout.employeeUser.name if checkout.employeeUser else 'Unknown',
                    employeeEmail=checkout.employeeUser.email if checkout.employeeUser else '',
                    employeeDepartment=checkout.employeeUser.department if checkout.employeeUser else None,
                    location=checkout.asset.location,
                    site=checkout.asset.site,
                )
            )
        
        checkouts = all_formatted_checkouts if includeCheckoutList else data["checkouts"]

        # Calculate total value
        total_value = sum(c.assetCost or 0 for c in checkouts)

        # Build summary data
        summary_data = [
            {
                "Metric": "Total Active Checkouts",
                "Value": str(summary.totalActive),
                "Overdue": str(summary.totalOverdue),
                "Historical": str(summary.totalHistorical),
                "Total Value": format_number(total_value),
            },
            {
                "Metric": "---",
                "Value": "---",
                "Overdue": "---",
                "Historical": "---",
                "Total Value": "---",
            },
            {
                "Metric": "CHECKOUTS BY EMPLOYEE",
                "Value": "",
                "Overdue": "",
                "Historical": "",
                "Total Value": "",
            },
            *[
                {
                    "Metric": f"Employee: {emp.employeeName or 'Unknown'}",
                    "Value": str(emp.count),
                    "Overdue": str(emp.overdueCount),
                    "Historical": str(emp.count - emp.overdueCount),
                    "Total Value": format_number(sum(c.assetCost or 0 for c in checkouts if c.employeeName == emp.employeeName)),
                }
                for emp in summary.byEmployee
            ],
            {
                "Metric": "---",
                "Value": "---",
                "Overdue": "---",
                "Historical": "---",
                "Total Value": "---",
            },
            {
                "Metric": "CHECKOUTS BY DEPARTMENT",
                "Value": "",
                "Overdue": "",
                "Historical": "",
                "Total Value": "",
            },
            *[
                {
                    "Metric": f"Department: {dept.department or 'Unknown'}",
                    "Value": str(dept.count),
                    "Overdue": str(dept.overdueCount),
                    "Historical": str(dept.count - dept.overdueCount),
                    "Total Value": format_number(sum(c.assetCost or 0 for c in checkouts if c.employeeDepartment == dept.department)),
                }
                for dept in summary.byDepartment
            ],
        ]

        # Prepare checkout list data
        checkout_list_data = [
            {
                "Asset Tag ID": c.assetTagId or "",
                "Description": c.assetDescription or "",
                "Category": c.category or "",
                "SUB-CATEGORY": c.subCategory or "",
                "Check-out Date": c.checkoutDate or "",
                "Due date": c.expectedReturnDate or "",
                "Return Date": c.returnDate or "",
                "Department": c.employeeDepartment or "",
                "Cost": format_number(c.assetCost) if c.assetCost else "",
            }
            for c in checkouts
        ]

        filename = f"checkout-report-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)

            if includeCheckoutList:
                # Include summary and checkout list
                writer.writerow(["=== SUMMARY STATISTICS ==="])
                if summary_data:
                    summary_headers = list(summary_data[0].keys())
                    writer.writerow(summary_headers)
                    for row in summary_data:
                        writer.writerow([row.get(header, "") for header in summary_headers])
                
                writer.writerow([])
                writer.writerow(["=== CHECKOUT LIST ==="])
                
                if checkout_list_data:
                    checkout_headers = list(checkout_list_data[0].keys())
                    writer.writerow(checkout_headers)
                    for row in checkout_list_data:
                        writer.writerow([row.get(header, "") for header in checkout_headers])
            else:
                # Summary only
                if summary_data:
                    summary_headers = list(summary_data[0].keys())
                    writer.writerow(summary_headers)
                    for row in summary_data:
                        writer.writerow([row.get(header, "") for header in summary_headers])

            csv_content = output.getvalue()
            output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # excel
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            if includeCheckoutList:
                # Multiple sheets: Summary, By Employee, By Department, Checkout List
                # Summary sheet
                if summary_data:
                    summary_ws = wb.create_sheet("Summary")
                    summary_headers = list(summary_data[0].keys())
                    summary_ws.append(summary_headers)
                    for row in summary_data:
                        summary_ws.append([row.get(header, "") for header in summary_headers])

                # By Employee sheet
                employee_data = [row for row in summary_data if row.get("Metric", "").startswith("Employee:")]
                if employee_data:
                    employee_ws = wb.create_sheet("By Employee")
                    employee_headers = list(employee_data[0].keys())
                    employee_ws.append(employee_headers)
                    for row in employee_data:
                        employee_ws.append([row.get(header, "") for header in employee_headers])

                # By Department sheet
                department_data = [row for row in summary_data if row.get("Metric", "").startswith("Department:")]
                if department_data:
                    department_ws = wb.create_sheet("By Department")
                    department_headers = list(department_data[0].keys())
                    department_ws.append(department_headers)
                    for row in department_data:
                        department_ws.append([row.get(header, "") for header in department_headers])

                # Checkout List sheet
                if checkout_list_data:
                    checkout_list_ws = wb.create_sheet("Checkout List")
                    checkout_headers = list(checkout_list_data[0].keys())
                    checkout_list_ws.append(checkout_headers)
                    for row in checkout_list_data:
                        checkout_list_ws.append([row.get(header, "") for header in checkout_headers])
            else:
                # Single sheet export
                if summary_data:
                    ws = wb.create_sheet("Checkout Report")
                    summary_headers = list(summary_data[0].keys())
                    ws.append(summary_headers)
                    for row in summary_data:
                        ws.append([row.get(header, "") for header in summary_headers])

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting checkout reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export checkout reports")

