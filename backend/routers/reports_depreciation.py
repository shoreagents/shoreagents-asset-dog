"""
Depreciation Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv

from models.reports import DepreciationReportResponse, DepreciationAsset, PaginationInfo
from auth import verify_auth
from database import prisma

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/depreciation", tags=["reports"])

def format_number(value: Optional[float]) -> str:
    """Format number with commas and 2 decimal places"""
    if value is None or value == 0:
        return '0.00'
    return f"{float(value):,.2f}"

def calculate_depreciation(
    depreciable_asset: bool,
    depreciable_cost: Optional[float],
    salvage_value: Optional[float],
    asset_life_months: Optional[int],
    depreciation_method: Optional[str],
    date_acquired: Optional[datetime]
) -> Dict[str, Any]:
    """Calculate depreciation values for an asset"""
    monthly_depreciation = 0.0
    annual_depreciation = 0.0
    accumulated_depreciation = 0.0
    current_value = 0.0
    depreciation_years = 0
    depreciation_months = 0

    if depreciable_asset and depreciable_cost and asset_life_months and date_acquired:
        depreciable_amount = float(depreciable_cost) - (float(salvage_value) if salvage_value else 0.0)
        
        # Calculate months elapsed
        now = datetime.now()
        months_elapsed = 0
        if asset_life_months > 0:
            time_diff = now - date_acquired
            days_elapsed = time_diff.days
            months_elapsed = min(
                int(days_elapsed / 30),
                asset_life_months
            )

        if depreciation_method == 'Straight-line' or not depreciation_method:
            # Straight-line depreciation
            monthly_depreciation = depreciable_amount / asset_life_months if asset_life_months > 0 else 0.0
            annual_depreciation = monthly_depreciation * 12
            accumulated_depreciation = monthly_depreciation * months_elapsed
            current_value = float(depreciable_cost) - accumulated_depreciation
            depreciation_years = months_elapsed // 12
            depreciation_months = months_elapsed % 12
        elif depreciation_method == 'Declining Balance':
            # Declining balance depreciation (simplified - using 200% declining balance)
            rate = 2.0 / asset_life_months if asset_life_months > 0 else 0.0
            remaining_value = float(depreciable_cost)
            accumulated_depreciation = 0.0
            
            for i in range(min(months_elapsed, asset_life_months)):
                monthly_dep = remaining_value * rate
                accumulated_depreciation += monthly_dep
                remaining_value -= monthly_dep
                if remaining_value < (float(salvage_value) if salvage_value else 0.0):
                    accumulated_depreciation = float(depreciable_cost) - (float(salvage_value) if salvage_value else 0.0)
                    break
            
            monthly_depreciation = accumulated_depreciation / months_elapsed if months_elapsed > 0 else 0.0
            annual_depreciation = monthly_depreciation * 12
            current_value = float(depreciable_cost) - accumulated_depreciation
            depreciation_years = months_elapsed // 12
            depreciation_months = months_elapsed % 12

    return {
        "monthlyDepreciation": monthly_depreciation,
        "annualDepreciation": annual_depreciation,
        "accumulatedDepreciation": accumulated_depreciation,
        "currentValue": current_value,
        "depreciationYears": depreciation_years,
        "depreciationMonths": depreciation_months,
    }

@router.get("", response_model=DepreciationReportResponse)
async def get_depreciation_reports(
    category: Optional[str] = Query(None, description="Filter by category name"),
    depreciationMethod: Optional[str] = Query(None, description="Filter by depreciation method"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    isDepreciable: Optional[bool] = Query(None, description="Filter by depreciable asset status"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get depreciation reports with optional filters and pagination"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        skip = (page - 1) * pageSize

        # Build where clause
        where_clause: Dict[str, Any] = {
            "isDeleted": False,
        }

        # Category filter
        if category:
            where_clause["category"] = {
                "name": category
            }

        # Depreciation method filter
        if depreciationMethod:
            where_clause["depreciationMethod"] = depreciationMethod

        # Location filter
        if location:
            where_clause["location"] = location

        # Site filter
        if site:
            where_clause["site"] = site

        # Depreciable asset filter
        if isDepreciable is not None:
            where_clause["depreciableAsset"] = isDepreciable

        # Date range filter (dateAcquired)
        if startDate or endDate:
            date_acquired_filter: Dict[str, Any] = {}
            if startDate:
                date_acquired_filter["gte"] = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
            if endDate:
                date_acquired_filter["lte"] = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
            if date_acquired_filter:
                where_clause["dateAcquired"] = date_acquired_filter

        # Get total count
        total = await prisma.assets.count(where=where_clause)

        # Get paginated assets
        assets_raw = await prisma.assets.find_many(
            where=where_clause,
            include={
                "category": True,
                "subCategory": True
            },
            order={"dateAcquired": "desc"},
            skip=skip,
            take=pageSize
        )

        # Calculate depreciation for each asset
        formatted_assets = []
        for asset in assets_raw:
            dep_values = calculate_depreciation(
                depreciable_asset=asset.depreciableAsset or False,
                depreciable_cost=float(asset.depreciableCost) if asset.depreciableCost else None,
                salvage_value=float(asset.salvageValue) if asset.salvageValue else None,
                asset_life_months=asset.assetLifeMonths,
                depreciation_method=asset.depreciationMethod,
                date_acquired=asset.dateAcquired
            )

            formatted_assets.append(
                DepreciationAsset(
                    id=asset.id,
                    assetTagId=asset.assetTagId,
                    description=asset.description,
                    category=asset.category.name if asset.category else None,
                    subCategory=asset.subCategory.name if asset.subCategory else None,
                    originalCost=float(asset.cost) if asset.cost else None,
                    depreciableCost=float(asset.depreciableCost) if asset.depreciableCost else None,
                    salvageValue=float(asset.salvageValue) if asset.salvageValue else None,
                    assetLifeMonths=asset.assetLifeMonths,
                    depreciationMethod=asset.depreciationMethod,
                    dateAcquired=asset.dateAcquired.isoformat() if asset.dateAcquired else None,
                    location=asset.location,
                    site=asset.site,
                    isDepreciable=asset.depreciableAsset or False,
                    monthlyDepreciation=dep_values["monthlyDepreciation"],
                    annualDepreciation=dep_values["annualDepreciation"],
                    accumulatedDepreciation=dep_values["accumulatedDepreciation"],
                    currentValue=dep_values["currentValue"],
                    depreciationYears=dep_values["depreciationYears"],
                    depreciationMonths=dep_values["depreciationMonths"],
                )
            )

        total_pages = (total + pageSize - 1) // pageSize if total > 0 else 0

        return DepreciationReportResponse(
            assets=formatted_assets,
            pagination=PaginationInfo(
                total=total,
                page=page,
                pageSize=pageSize,
                totalPages=total_pages,
                hasNextPage=page < total_pages,
                hasPreviousPage=page > 1,
            )
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching depreciation reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch depreciation reports")

@router.get("/export")
async def export_depreciation_reports(
    format: str = Query("csv", description="Export format: csv or excel"),
    category: Optional[str] = Query(None, description="Filter by category name"),
    depreciationMethod: Optional[str] = Query(None, description="Filter by depreciation method"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    isDepreciable: Optional[bool] = Query(None, description="Filter by depreciable asset status"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeAssetList: Optional[bool] = Query(False, description="Include asset list in export"),
    auth: dict = Depends(verify_auth)
):
    """Export depreciation reports to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        if format not in ["csv", "excel"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv or excel.")

        # Fetch all assets for export
        page_size = 10000 if includeAssetList else 1
        report_data = await get_depreciation_reports(
            category=category,
            depreciationMethod=depreciationMethod,
            location=location,
            site=site,
            isDepreciable=isDepreciable,
            startDate=startDate,
            endDate=endDate,
            page=1,
            pageSize=page_size,
            auth=auth
        )

        assets = report_data.assets

        # Calculate summary statistics
        depreciable_assets = [a for a in assets if a.isDepreciable]
        total_original_cost = sum(a.originalCost or 0 for a in assets)
        total_depreciable_cost = sum(a.depreciableCost or 0 for a in depreciable_assets)
        total_accumulated_depreciation = sum(a.accumulatedDepreciation for a in depreciable_assets)
        total_current_value = sum(a.currentValue for a in depreciable_assets)
        total_annual_depreciation = sum(a.annualDepreciation for a in depreciable_assets)

        # Group by method
        by_method: Dict[str, Dict[str, Any]] = {}
        for asset in depreciable_assets:
            method = asset.depreciationMethod or 'Not Specified'
            if method not in by_method:
                by_method[method] = {
                    "count": 0,
                    "totalCost": 0.0,
                    "totalDepreciation": 0.0,
                    "totalCurrentValue": 0.0,
                }
            by_method[method]["count"] += 1
            by_method[method]["totalCost"] += asset.depreciableCost or 0
            by_method[method]["totalDepreciation"] += asset.accumulatedDepreciation
            by_method[method]["totalCurrentValue"] += asset.currentValue

        filename = f"depreciation-report-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)

            if includeAssetList:
                # Include summary and asset list
                writer.writerow(["DEPRECIATION REPORT SUMMARY"])
                writer.writerow(["Total Assets", len(assets)])
                writer.writerow(["Depreciable Assets", len(depreciable_assets)])
                writer.writerow(["Total Original Cost", format_number(total_original_cost)])
                writer.writerow(["Total Depreciable Cost", format_number(total_depreciable_cost)])
                writer.writerow(["Accumulated Depreciation", format_number(total_accumulated_depreciation)])
                writer.writerow(["Total Current Value", format_number(total_current_value)])
                writer.writerow(["Total Annual Depreciation", format_number(total_annual_depreciation)])
                writer.writerow([])
                
                writer.writerow(["DEPRECIATION BY METHOD"])
                writer.writerow(["Method", "Asset Count", "Total Cost", "Accumulated Depreciation", "Current Value"])
                for method, stats in by_method.items():
                    writer.writerow([
                        method,
                        stats["count"],
                        format_number(stats["totalCost"]),
                        format_number(stats["totalDepreciation"]),
                        format_number(stats["totalCurrentValue"]),
                    ])
                writer.writerow([])
                
                writer.writerow(["ASSET DEPRECIATION DETAILS"])
                writer.writerow([
                    "Asset Tag ID",
                    "Description",
                    "Category",
                    "Depreciation Method",
                    "Original Cost",
                    "Depreciable Cost",
                    "Salvage Value",
                    "Asset Life (Months)",
                    "Date Acquired",
                    "Monthly Depreciation",
                    "Annual Depreciation",
                    "Accumulated Depreciation",
                    "Current Value",
                ])
                for asset in assets:
                    writer.writerow([
                        asset.assetTagId,
                        asset.description,
                        asset.category or "N/A",
                        asset.depreciationMethod or "N/A",
                        format_number(asset.originalCost),
                        format_number(asset.depreciableCost),
                        format_number(asset.salvageValue),
                        asset.assetLifeMonths or "N/A",
                        asset.dateAcquired or "N/A",
                        format_number(asset.monthlyDepreciation),
                        format_number(asset.annualDepreciation),
                        format_number(asset.accumulatedDepreciation),
                        format_number(asset.currentValue),
                    ])
            else:
                # Summary only
                writer.writerow(["DEPRECIATION REPORT SUMMARY"])
                writer.writerow(["Total Assets", len(assets)])
                writer.writerow(["Depreciable Assets", len(depreciable_assets)])
                writer.writerow(["Total Original Cost", format_number(total_original_cost)])
                writer.writerow(["Total Depreciable Cost", format_number(total_depreciable_cost)])
                writer.writerow(["Accumulated Depreciation", format_number(total_accumulated_depreciation)])
                writer.writerow(["Total Current Value", format_number(total_current_value)])
                writer.writerow(["Total Annual Depreciation", format_number(total_annual_depreciation)])
                writer.writerow([])
                writer.writerow(["DEPRECIATION BY METHOD"])
                writer.writerow(["Method", "Asset Count", "Total Cost", "Accumulated Depreciation", "Current Value"])
                for method, stats in by_method.items():
                    writer.writerow([
                        method,
                        stats["count"],
                        format_number(stats["totalCost"]),
                        format_number(stats["totalDepreciation"]),
                        format_number(stats["totalCurrentValue"]),
                    ])

            csv_content = output.getvalue()
            output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # excel
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Summary sheet
            summary_data = [
                ["DEPRECIATION REPORT SUMMARY"],
                ["Total Assets", len(assets)],
                ["Depreciable Assets", len(depreciable_assets)],
                ["Total Original Cost", format_number(total_original_cost)],
                ["Total Depreciable Cost", format_number(total_depreciable_cost)],
                ["Accumulated Depreciation", format_number(total_accumulated_depreciation)],
                ["Total Current Value", format_number(total_current_value)],
                ["Total Annual Depreciation", format_number(total_annual_depreciation)],
                [],
                ["DEPRECIATION BY METHOD"],
                ["Method", "Asset Count", "Total Cost", "Accumulated Depreciation", "Current Value"],
                *[[method, stats["count"], format_number(stats["totalCost"]), format_number(stats["totalDepreciation"]), format_number(stats["totalCurrentValue"])] for method, stats in by_method.items()],
            ]
            summary_ws = wb.create_sheet("Summary")
            for row in summary_data:
                summary_ws.append(row)

            if includeAssetList:
                # Asset list sheet
                asset_data = [
                    [
                        asset.assetTagId,
                        asset.description,
                        asset.category or "N/A",
                        asset.depreciationMethod or "N/A",
                        format_number(asset.originalCost),
                        format_number(asset.depreciableCost),
                        format_number(asset.salvageValue),
                        asset.assetLifeMonths or "N/A",
                        asset.dateAcquired or "N/A",
                        format_number(asset.monthlyDepreciation),
                        format_number(asset.annualDepreciation),
                        format_number(asset.accumulatedDepreciation),
                        format_number(asset.currentValue),
                    ]
                    for asset in assets
                ]
                asset_ws = wb.create_sheet("Asset List")
                asset_ws.append([
                    "Asset Tag ID",
                    "Description",
                    "Category",
                    "Depreciation Method",
                    "Original Cost",
                    "Depreciable Cost",
                    "Salvage Value",
                    "Asset Life (Months)",
                    "Date Acquired",
                    "Monthly Depreciation",
                    "Annual Depreciation",
                    "Accumulated Depreciation",
                    "Current Value",
                ])
                for row in asset_data:
                    asset_ws.append(row)

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting depreciation reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export depreciation reports")

