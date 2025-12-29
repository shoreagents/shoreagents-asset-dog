"""
Lease Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv

from models.reports import LeaseReportResponse, LeaseItem, PaginationInfo
from auth import verify_auth
from database import prisma
from utils.pdf_generator import ReportPDF, PDF_AVAILABLE

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/lease", tags=["reports"])

async def check_permission(user_id: str, permission: str) -> bool:
    """Check if user has a specific permission. Admins have all permissions."""
    try:
        asset_user = await prisma.assetuser.find_unique(
            where={"userId": user_id}
        )
        if not asset_user or not asset_user.isActive:
            return False
        
        # Admins have all permissions
        if asset_user.role == "admin":
            return True
        
        return getattr(asset_user, permission, False)
    except Exception:
        return False

def format_number(value: Optional[float]) -> str:
    """Format number with commas and 2 decimal places"""
    if value is None or value == 0:
        return '0.00'
    return f"{float(value):,.2f}"

@router.get("", response_model=LeaseReportResponse)
async def get_lease_reports(
    category: Optional[str] = Query(None, description="Filter by category name"),
    lessee: Optional[str] = Query(None, description="Filter by lessee name"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    status: Optional[str] = Query(None, description="Filter by status (active, expired, upcoming)"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get lease reports with optional filters and pagination"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        skip = (page - 1) * pageSize

        # Build where clause
        where_clause: Dict[str, Any] = {
            "asset": {
                "isDeleted": False,
            }
        }

        # Category filter
        if category:
            where_clause["asset"] = {
                **where_clause["asset"],
                "category": {
                    "name": category
                }
            }

        # Lessee filter (case-insensitive search)
        if lessee:
            where_clause["lessee"] = {
                "contains": lessee,
                "mode": "insensitive"
            }

        # Location filter
        if location:
            where_clause["asset"] = {
                **where_clause["asset"],
                "location": location
            }

        # Site filter
        if site:
            where_clause["asset"] = {
                **where_clause["asset"],
                "site": site
            }

        # Status filter (active, expired, upcoming)
        # Use timezone-naive datetime for Prisma queries
        now = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
        
        if status == 'active':
            where_clause["AND"] = [
                {"leaseStartDate": {"lte": now}},
                {
                    "OR": [
                        {"leaseEndDate": {"gte": now}},
                        {"leaseEndDate": None},
                    ]
                },
            ]
        elif status == 'expired':
            where_clause["leaseEndDate"] = {
                "lt": now,
                "not": None
            }
        elif status == 'upcoming':
            where_clause["leaseStartDate"] = {"gt": now}

        # Date range filter (lease start date)
        if startDate or endDate:
            lease_start_filter: Dict[str, Any] = {}
            if startDate:
                date_obj = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
                # Convert to timezone-naive if needed
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                lease_start_filter["gte"] = date_obj
            if endDate:
                date_obj = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
                # Convert to timezone-naive if needed
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                lease_start_filter["lte"] = date_obj
            if lease_start_filter:
                where_clause["leaseStartDate"] = lease_start_filter

        # Get total count
        total = await prisma.assetslease.count(where=where_clause)

        # Get paginated leases
        leases_raw = await prisma.assetslease.find_many(
            where=where_clause,
            include={
                "asset": {
                    "include": {
                        "category": True,
                        "subCategory": True
                    }
                },
                "returns": True
            },
            order={"leaseStartDate": "desc"},
            skip=skip,
            take=pageSize
        )

        # Sort returns by date descending and take only the first one for each lease
        for lease in leases_raw:
            if lease.returns:
                lease.returns = sorted(
                    lease.returns,
                    key=lambda r: r.returnDate if r.returnDate else datetime.min,
                    reverse=True
                )[:1]

        # Format lease data
        formatted_leases = []
        # Get timezone-naive current date for comparisons
        now_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
        
        for lease in leases_raw:
            # Convert to timezone-naive datetimes for comparison
            lease_start = lease.leaseStartDate
            if lease_start.tzinfo is not None:
                lease_start = lease_start.replace(tzinfo=None)
            lease_start_date = lease_start.replace(hour=0, minute=0, second=0, microsecond=0)
            
            lease_end = lease.leaseEndDate
            if lease_end:
                if lease_end.tzinfo is not None:
                    lease_end = lease_end.replace(tzinfo=None)
                end_date = lease_end.replace(hour=0, minute=0, second=0, microsecond=0)
            else:
                end_date = None
            
            # Calculate lease status
            lease_status = 'active'
            if end_date:
                if end_date < now_date:
                    lease_status = 'expired'
                elif lease_start_date > now_date:
                    lease_status = 'upcoming'
            elif lease_start_date > now_date:
                lease_status = 'upcoming'

            # Calculate days remaining or days expired
            days_remaining: Optional[int] = None
            if end_date:
                diff_time = (end_date - now_date).total_seconds()
                diff_days = int(diff_time / (60 * 60 * 24))
                days_remaining = diff_days

            formatted_leases.append(
                LeaseItem(
                    id=lease.id,
                    assetTagId=lease.asset.assetTagId,
                    description=lease.asset.description,
                    category=lease.asset.category.name if lease.asset.category else None,
                    subCategory=lease.asset.subCategory.name if lease.asset.subCategory else None,
                    lessee=lease.lessee,
                    leaseStartDate=lease.leaseStartDate.isoformat(),
                    leaseEndDate=lease.leaseEndDate.isoformat() if lease.leaseEndDate else None,
                    conditions=lease.conditions,
                    notes=lease.notes,
                    location=lease.asset.location,
                    site=lease.asset.site,
                    assetStatus=lease.asset.status,
                    assetCost=float(lease.asset.cost) if lease.asset.cost else None,
                    leaseStatus=lease_status,
                    daysRemaining=days_remaining,
                    lastReturnDate=lease.returns[0].returnDate.isoformat() if lease.returns and len(lease.returns) > 0 and lease.returns[0].returnDate else None,
                    returnCondition=lease.returns[0].condition if lease.returns and len(lease.returns) > 0 else None,
                    createdAt=lease.createdAt.isoformat(),
                )
            )

        total_pages = (total + pageSize - 1) // pageSize if total > 0 else 0

        return LeaseReportResponse(
            leases=formatted_leases,
            pagination=PaginationInfo(
                total=total,
                page=page,
                pageSize=pageSize,
                totalPages=total_pages,
                hasNextPage=page < total_pages,
                hasPreviousPage=page > 1,
            )
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching lease reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch lease reports")

@router.get("/export")
async def export_lease_reports(
    format: str = Query("csv", description="Export format: csv or excel"),
    category: Optional[str] = Query(None, description="Filter by category name"),
    lessee: Optional[str] = Query(None, description="Filter by lessee name"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    status: Optional[str] = Query(None, description="Filter by status (active, expired, upcoming)"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeLeaseList: Optional[bool] = Query(False, description="Include lease list in export"),
    auth: dict = Depends(verify_auth)
):
    """Export lease reports to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Check permission - user must have canManageReports to export reports
        has_permission = await check_permission(user_id, "canManageReports")
        if not has_permission:
            raise HTTPException(
                status_code=403,
                detail="You do not have permission to export reports"
            )

        if format not in ["csv", "excel", "pdf"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv, excel, or pdf.")
        
        if format == "pdf" and not PDF_AVAILABLE:
            raise HTTPException(status_code=500, detail="PDF export not available - fpdf2 not installed")

        # Fetch all leases for export
        page_size = 10000 if includeLeaseList else 1
        report_data = await get_lease_reports(
            category=category,
            lessee=lessee,
            location=location,
            site=site,
            status=status,
            startDate=startDate,
            endDate=endDate,
            page=1,
            pageSize=page_size,
            auth=auth
        )

        leases = report_data.leases

        # Calculate summary statistics
        active_leases = [l for l in leases if l.leaseStatus == 'active']
        expired_leases = [l for l in leases if l.leaseStatus == 'expired']
        upcoming_leases = [l for l in leases if l.leaseStatus == 'upcoming']
        total_asset_value = sum(l.assetCost or 0 for l in leases)

        # Group by lessee
        by_lessee: Dict[str, Dict[str, Any]] = {}
        for lease in leases:
            lessee = lease.lessee
            if lessee not in by_lessee:
                by_lessee[lessee] = {
                    "count": 0,
                    "totalValue": 0.0,
                }
            by_lessee[lessee]["count"] += 1
            by_lessee[lessee]["totalValue"] += lease.assetCost or 0

        filename = f"lease-report-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)

            if includeLeaseList:
                # Include summary and lease list
                writer.writerow(["LEASED ASSET REPORT SUMMARY"])
                writer.writerow(["Total Leases", len(leases)])
                writer.writerow(["Active Leases", len(active_leases)])
                writer.writerow(["Expired Leases", len(expired_leases)])
                writer.writerow(["Upcoming Leases", len(upcoming_leases)])
                writer.writerow(["Total Asset Value", format_number(total_asset_value)])
                writer.writerow([])
                
                writer.writerow(["LEASES BY LESSEE"])
                writer.writerow(["Lessee", "Lease Count", "Total Asset Value"])
                for lessee, stats in by_lessee.items():
                    writer.writerow([
                        lessee,
                        stats["count"],
                        format_number(stats["totalValue"]),
                    ])
                writer.writerow([])
                
                writer.writerow(["LEASE RECORDS"])
                writer.writerow([
                    "Asset Tag ID",
                    "Description",
                    "Category",
                    "Sub-Category",
                    "Lessee",
                    "Lease Start Date",
                    "Lease End Date",
                    "Status",
                    "Days Remaining",
                    "Location",
                    "Site",
                    "Asset Cost",
                ])
                for lease in leases:
                    writer.writerow([
                        lease.assetTagId,
                        lease.description,
                        lease.category or "N/A",
                        lease.subCategory or "N/A",
                        lease.lessee,
                        lease.leaseStartDate.split('T')[0] if lease.leaseStartDate else "N/A",
                        lease.leaseEndDate.split('T')[0] if lease.leaseEndDate else "N/A",
                        lease.leaseStatus,
                        str(lease.daysRemaining) if lease.daysRemaining is not None else "N/A",
                        lease.location or "N/A",
                        lease.site or "N/A",
                        format_number(lease.assetCost),
                    ])
            else:
                # Summary only
                writer.writerow(["LEASED ASSET REPORT SUMMARY"])
                writer.writerow(["Total Leases", len(leases)])
                writer.writerow(["Active Leases", len(active_leases)])
                writer.writerow(["Expired Leases", len(expired_leases)])
                writer.writerow(["Upcoming Leases", len(upcoming_leases)])
                writer.writerow(["Total Asset Value", format_number(total_asset_value)])
                writer.writerow([])
                writer.writerow(["LEASES BY LESSEE"])
                writer.writerow(["Lessee", "Lease Count", "Total Asset Value"])
                for lessee, stats in by_lessee.items():
                    writer.writerow([
                        lessee,
                        stats["count"],
                        format_number(stats["totalValue"]),
                    ])

            csv_content = output.getvalue()
            output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        elif format == "excel":
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Summary sheet
            summary_data = [
                ["LEASED ASSET REPORT SUMMARY"],
                ["Total Leases", len(leases)],
                ["Active Leases", len(active_leases)],
                ["Expired Leases", len(expired_leases)],
                ["Upcoming Leases", len(upcoming_leases)],
                ["Total Asset Value", format_number(total_asset_value)],
                [],
                ["LEASES BY LESSEE"],
                ["Lessee", "Lease Count", "Total Asset Value"],
                *[[lessee, stats["count"], format_number(stats["totalValue"])] for lessee, stats in by_lessee.items()],
            ]
            summary_ws = wb.create_sheet("Summary")
            for row in summary_data:
                summary_ws.append(row)

            if includeLeaseList:
                # Lease list sheet
                lease_data = [
                    [
                        lease.assetTagId,
                        lease.description,
                        lease.category or "N/A",
                        lease.subCategory or "N/A",
                        lease.lessee,
                        lease.leaseStartDate.split('T')[0] if lease.leaseStartDate else "N/A",
                        lease.leaseEndDate.split('T')[0] if lease.leaseEndDate else "N/A",
                        lease.leaseStatus,
                        lease.daysRemaining if lease.daysRemaining is not None else "N/A",
                        lease.location or "N/A",
                        lease.site or "N/A",
                        format_number(lease.assetCost),
                    ]
                    for lease in leases
                ]
                lease_ws = wb.create_sheet("Lease List")
                lease_ws.append([
                    "Asset Tag ID",
                    "Description",
                    "Category",
                    "Sub-Category",
                    "Lessee",
                    "Lease Start Date",
                    "Lease End Date",
                    "Status",
                    "Days Remaining",
                    "Location",
                    "Site",
                    "Asset Cost",
                ])
                for row in lease_data:
                    lease_ws.append(row)

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # pdf
            pdf = ReportPDF("Lease Report", "Lease")
            pdf.add_page()

            pdf.add_section_title("Summary Statistics")
            summary_rows = [
                ["Total Leases", str(len(leases))],
                ["Active Leases", str(len(active_leases))],
                ["Expired Leases", str(len(expired_leases))],
                ["Upcoming Leases", str(len(upcoming_leases))],
                ["Total Asset Value", format_number(total_asset_value)],
            ]
            if summary_rows:
                headers = ["Metric", "Value"]
                pdf.add_table(headers, summary_rows)
            
            pdf.ln(10)

            if leases and includeLeaseList:
                pdf.add_section_title(f"Lease List ({len(leases)} leases)")
                simplified_headers = ["Asset Tag ID", "Description", "Category", "Lessee", "Lease Start", "Lease End", "Status", "Days Remaining", "Asset Cost"]
                simplified_rows = [
                    [
                        str(l.assetTagId or ""),
                        str(l.description or "")[:50],
                        str(l.category or ""),
                        str(l.lessee or ""),
                        l.leaseStartDate.split('T')[0] if l.leaseStartDate else "",
                        l.leaseEndDate.split('T')[0] if l.leaseEndDate else "",
                        str(l.leaseStatus or ""),
                        str(l.daysRemaining) if l.daysRemaining is not None else "",
                        format_number(l.assetCost),
                    ]
                    for l in leases
                ]
                pdf.add_table(simplified_headers, simplified_rows)

            pdf_content = bytes(pdf.output())
            filename += ".pdf"
            return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting lease reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export lease reports")

