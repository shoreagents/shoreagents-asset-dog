"""
Location Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv
import asyncio

from models.reports import LocationReportResponse, LocationSummary, LocationGroup, SiteGroup, LocationAsset, MovementItem, PaginationInfo
from auth import verify_auth
from database import prisma

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/location", tags=["reports"])

def format_number(value: Optional[float]) -> str:
    """Format number with commas and 2 decimal places"""
    if value is None or value == 0:
        return '0.00'
    return f"{float(value):,.2f}"

@router.get("", response_model=LocationReportResponse)
async def get_location_reports(
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    category: Optional[str] = Query(None, description="Filter by category ID"),
    status: Optional[str] = Query(None, description="Filter by status"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get location reports with optional filters and pagination"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        skip = (page - 1) * pageSize

        # Build where clause
        where_clause: Dict[str, Any] = {
            "isDeleted": False,
        }

        # Apply filters
        if location:
            where_clause["location"] = location

        if site:
            where_clause["site"] = site

        if category:
            where_clause["categoryId"] = category

        if status:
            where_clause["status"] = status

        # Date range filter (on purchaseDate or createdAt)
        if startDate or endDate:
            date_filters: List[Dict[str, Any]] = []
            
            purchase_date_filter: Dict[str, Any] = {}
            if startDate:
                date_obj = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                purchase_date_filter["gte"] = date_obj
            if endDate:
                date_obj = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                purchase_date_filter["lte"] = date_obj
            if purchase_date_filter:
                date_filters.append({"purchaseDate": purchase_date_filter})
            
            created_at_filter: Dict[str, Any] = {}
            if startDate:
                date_obj = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                created_at_filter["gte"] = date_obj
            if endDate:
                date_obj = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                created_at_filter["lte"] = date_obj
            if created_at_filter:
                date_filters.append({"createdAt": created_at_filter})
            
            if date_filters:
                where_clause["OR"] = date_filters

        # Get total count and fetch all assets for summary (no pagination)
        total_assets, all_assets_raw, paginated_assets_raw = await asyncio.gather(
            prisma.assets.count(where=where_clause),
            prisma.assets.find_many(
                where=where_clause,
                include={"category": True}
            ),
            prisma.assets.find_many(
                where=where_clause,
                include={
                    "category": True,
                    "moves": True
                },
                order={"createdAt": "desc"},
                skip=skip,
                take=pageSize
            )
        )

        # Sort moves by date descending for each asset
        for asset in paginated_assets_raw:
            if asset.moves:
                asset.moves = sorted(
                    asset.moves,
                    key=lambda m: m.moveDate if m.moveDate else datetime.min,
                    reverse=True
                )[:10]

        # Group by location (using ALL assets for summary)
        by_location_map: Dict[str, Dict[str, Any]] = {}
        for asset in all_assets_raw:
            location_key = asset.location or 'Unassigned'
            if location_key not in by_location_map:
                by_location_map[location_key] = {
                    "location": location_key,
                    "count": 0,
                    "totalValue": 0.0,
                }
            by_location_map[location_key]["count"] += 1
            by_location_map[location_key]["totalValue"] += float(asset.cost) if asset.cost else 0.0

        # Group by site (using ALL assets for summary)
        by_site_map: Dict[str, Dict[str, Any]] = {}
        for asset in all_assets_raw:
            site_key = asset.site or 'Unassigned'
            if site_key not in by_site_map:
                by_site_map[site_key] = {
                    "site": site_key,
                    "count": 0,
                    "totalValue": 0.0,
                    "locations": set(),
                }
            by_site_map[site_key]["count"] += 1
            by_site_map[site_key]["totalValue"] += float(asset.cost) if asset.cost else 0.0
            if asset.location:
                by_site_map[site_key]["locations"].add(asset.location)

        # Calculate location utilization
        by_location = [
            LocationGroup(
                location=group["location"],
                assetCount=group["count"],
                totalValue=group["totalValue"],
                averageValue=group["totalValue"] / group["count"] if group["count"] > 0 else 0.0,
                utilizationPercentage=(group["count"] / total_assets * 100) if total_assets > 0 else 0.0,
            )
            for group in by_location_map.values()
        ]

        # Calculate site utilization
        by_site = [
            SiteGroup(
                site=group["site"],
                assetCount=group["count"],
                totalValue=group["totalValue"],
                locationCount=len(group["locations"]),
                averageValue=group["totalValue"] / group["count"] if group["count"] > 0 else 0.0,
                utilizationPercentage=(group["count"] / total_assets * 100) if total_assets > 0 else 0.0,
            )
            for group in by_site_map.values()
        ]

        # Get movement history
        movement_where: Dict[str, Any] = {}
        if startDate or endDate:
            move_date_filter: Dict[str, Any] = {}
            if startDate:
                date_obj = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                move_date_filter["gte"] = date_obj
            if endDate:
                date_obj = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                move_date_filter["lte"] = date_obj
            if move_date_filter:
                movement_where["moveDate"] = move_date_filter

        movements_raw = await prisma.assetsmove.find_many(
            where=movement_where,
            include={
                "asset": True,
                "employeeUser": True
            },
            order={"moveDate": "desc"},
            take=100
        )

        # Format assets
        formatted_assets = [
            LocationAsset(
                id=asset.id,
                assetTagId=asset.assetTagId,
                description=asset.description,
                status=asset.status,
                cost=float(asset.cost) if asset.cost else None,
                category=asset.category.name if asset.category else None,
                location=asset.location,
                site=asset.site,
                department=asset.department,
                lastMoveDate=asset.moves[0].moveDate.isoformat().split('T')[0] if asset.moves and len(asset.moves) > 0 and asset.moves[0].moveDate else None,
            )
            for asset in paginated_assets_raw
        ]

        # Format movements
        formatted_movements = [
            MovementItem(
                id=move.id,
                assetTagId=move.asset.assetTagId,
                assetDescription=move.asset.description,
                moveType=move.moveType,
                moveDate=move.moveDate.isoformat().split('T')[0] if move.moveDate else "",
                employeeName=move.employeeUser.name if move.employeeUser else None,
                reason=move.reason,
                notes=move.notes,
            )
            for move in movements_raw
        ]

        total_pages = (total_assets + pageSize - 1) // pageSize if total_assets > 0 else 0

        return LocationReportResponse(
            summary=LocationSummary(
                totalAssets=total_assets,
                totalLocations=len(by_location_map),
                totalSites=len(by_site_map),
                byLocation=by_location,
                bySite=by_site,
            ),
            assets=formatted_assets,
            movements=formatted_movements,
            generatedAt=datetime.now().isoformat(),
            pagination=PaginationInfo(
                total=total_assets,
                page=page,
                pageSize=pageSize,
                totalPages=total_pages,
                hasNextPage=page < total_pages,
                hasPreviousPage=page > 1,
            )
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching location reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch location reports")

@router.get("/export")
async def export_location_reports(
    format: str = Query("csv", description="Export format: csv or excel"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    category: Optional[str] = Query(None, description="Filter by category ID"),
    status: Optional[str] = Query(None, description="Filter by status"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeAssetList: Optional[bool] = Query(False, description="Include asset list in export"),
    auth: dict = Depends(verify_auth)
):
    """Export location reports to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        if format not in ["csv", "excel"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv or excel.")

        # Fetch all data for export
        page_size = 10000 if includeAssetList else 1
        report_data = await get_location_reports(
            location=location,
            site=site,
            category=category,
            status=status,
            startDate=startDate,
            endDate=endDate,
            page=1,
            pageSize=page_size,
            auth=auth
        )

        summary = report_data.summary
        assets = report_data.assets

        # Prepare summary statistics
        summary_data = [
            {
                "Metric": "Total Assets",
                "Value": str(summary.totalAssets),
                "Total Value": "",
                "Location Count": "",
                "Average Value": "",
                "Utilization %": "",
            },
            {
                "Metric": "Total Locations",
                "Value": str(summary.totalLocations),
                "Total Value": "",
                "Location Count": "",
                "Average Value": "",
                "Utilization %": "",
            },
            {
                "Metric": "Total Sites",
                "Value": str(summary.totalSites),
                "Total Value": "",
                "Location Count": "",
                "Average Value": "",
                "Utilization %": "",
            },
            {
                "Metric": "---",
                "Value": "---",
                "Total Value": "---",
                "Location Count": "---",
                "Average Value": "---",
                "Utilization %": "---",
            },
            {
                "Metric": "ASSETS BY LOCATION",
                "Value": "",
                "Total Value": "",
                "Location Count": "",
                "Average Value": "",
                "Utilization %": "",
            },
            *[
                {
                    "Metric": f"Location: {loc.location}",
                    "Value": str(loc.assetCount),
                    "Total Value": format_number(loc.totalValue),
                    "Location Count": "",
                    "Average Value": format_number(loc.averageValue),
                    "Utilization %": f"{loc.utilizationPercentage:.1f}%",
                }
                for loc in summary.byLocation
            ],
            {
                "Metric": "---",
                "Value": "---",
                "Total Value": "---",
                "Location Count": "---",
                "Average Value": "---",
                "Utilization %": "---",
            },
            {
                "Metric": "ASSETS BY SITE",
                "Value": "",
                "Total Value": "",
                "Location Count": "",
                "Average Value": "",
                "Utilization %": "",
            },
            *[
                {
                    "Metric": f"Site: {site.site}",
                    "Value": str(site.assetCount),
                    "Total Value": format_number(site.totalValue),
                    "Location Count": str(site.locationCount),
                    "Average Value": format_number(site.averageValue),
                    "Utilization %": f"{site.utilizationPercentage:.1f}%",
                }
                for site in summary.bySite
            ],
        ]

        # Prepare asset list data
        asset_list_data = [
            {
                "Asset Tag ID": asset.assetTagId or "",
                "Description": asset.description or "",
                "Status": asset.status or "",
                "Cost": format_number(asset.cost) if asset.cost else "",
                "Category": asset.category or "",
                "Location": asset.location or "",
                "Site": asset.site or "",
                "Department": asset.department or "",
                "Last Move Date": asset.lastMoveDate or "",
            }
            for asset in assets
        ]

        filename = f"location-report-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)

            if includeAssetList:
                # Include summary and asset list
                writer.writerow(["=== SUMMARY STATISTICS ==="])
                if summary_data:
                    summary_headers = list(summary_data[0].keys())
                    writer.writerow(summary_headers)
                    for row in summary_data:
                        writer.writerow([row.get(header, "") for header in summary_headers])
                
                writer.writerow([])
                writer.writerow(["=== ASSET LIST ==="])
                
                if asset_list_data:
                    asset_headers = list(asset_list_data[0].keys())
                    writer.writerow(asset_headers)
                    for row in asset_list_data:
                        writer.writerow([row.get(header, "") for header in asset_headers])
            else:
                # Summary only
                if summary_data:
                    summary_headers = list(summary_data[0].keys())
                    writer.writerow(summary_headers)
                    for row in summary_data:
                        writer.writerow([row.get(header, "") for header in summary_headers])

            csv_content = output.getvalue()
            output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # excel
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Summary sheet
            if summary_data:
                summary_ws = wb.create_sheet("Summary")
                summary_headers = list(summary_data[0].keys())
                summary_ws.append(summary_headers)
                for row in summary_data:
                    summary_ws.append([row.get(header, "") for header in summary_headers])

            if includeAssetList:
                # By Location sheet
                location_data = [row for row in summary_data if row.get("Metric", "").startswith("Location:")]
                if location_data:
                    location_ws = wb.create_sheet("By Location")
                    location_headers = list(location_data[0].keys())
                    location_ws.append(location_headers)
                    for row in location_data:
                        location_ws.append([row.get(header, "") for header in location_headers])

                # By Site sheet
                site_data = [row for row in summary_data if row.get("Metric", "").startswith("Site:")]
                if site_data:
                    site_ws = wb.create_sheet("By Site")
                    site_headers = list(site_data[0].keys())
                    site_ws.append(site_headers)
                    for row in site_data:
                        site_ws.append([row.get(header, "") for header in site_headers])

                # Asset List sheet
                if asset_list_data:
                    asset_list_ws = wb.create_sheet("Asset List")
                    asset_headers = list(asset_list_data[0].keys())
                    asset_list_ws.append(asset_headers)
                    for row in asset_list_data:
                        asset_list_ws.append([row.get(header, "") for header in asset_headers])
            else:
                # Single sheet export
                if not summary_data:
                    raise HTTPException(status_code=400, detail="No data to export")

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting location reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export location reports")

