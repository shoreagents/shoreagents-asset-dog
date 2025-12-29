"""
Maintenance Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv
import asyncio

from models.reports import MaintenanceReportResponse, MaintenanceSummary, MaintenanceItem, UpcomingMaintenance, MaintenanceStatusGroup, TotalCostByStatus, MaintenanceInventoryItem, PaginationInfo
from auth import verify_auth
from database import prisma
from utils.pdf_generator import ReportPDF, PDF_AVAILABLE

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/maintenance", tags=["reports"])

async def check_permission(user_id: str, permission: str) -> bool:
    """Check if user has a specific permission. Admins have all permissions."""
    try:
        asset_user = await prisma.assetuser.find_unique(
            where={"userId": user_id}
        )
        if not asset_user or not asset_user.isActive:
            return False
        
        # Admins have all permissions
        if asset_user.role == "admin":
            return True
        
        return getattr(asset_user, permission, False)
    except Exception:
        return False

def format_number(value: Optional[float]) -> str:
    """Format number with commas and 2 decimal places"""
    if value is None or value == 0:
        return '0.00'
    return f"{float(value):,.2f}"

@router.get("", response_model=MaintenanceReportResponse)
async def get_maintenance_reports(
    assetId: Optional[str] = Query(None, description="Filter by asset ID"),
    category: Optional[str] = Query(None, description="Filter by category name"),
    location: Optional[str] = Query(None, description="Filter by asset location"),
    site: Optional[str] = Query(None, description="Filter by asset site"),
    department: Optional[str] = Query(None, description="Filter by asset department"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get maintenance reports with optional filters and pagination"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        skip = (page - 1) * pageSize

        # Build where clause
        where_clause: Dict[str, Any] = {}

        # Apply filters
        if assetId:
            where_clause["assetId"] = assetId

        # Date range filter (on dueDate or createdAt)
        if startDate or endDate:
            date_filters: List[Dict[str, Any]] = []
            
            due_date_filter: Dict[str, Any] = {}
            if startDate:
                date_obj = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                due_date_filter["gte"] = date_obj
            if endDate:
                date_obj = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                due_date_filter["lte"] = date_obj
            if due_date_filter:
                date_filters.append({"dueDate": due_date_filter})
            
            created_at_filter: Dict[str, Any] = {}
            if startDate:
                date_obj = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                created_at_filter["gte"] = date_obj
            if endDate:
                date_obj = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                created_at_filter["lte"] = date_obj
            if created_at_filter:
                date_filters.append({"createdAt": created_at_filter})
            
            if date_filters:
                where_clause["OR"] = date_filters

        # Get total count and fetch all maintenances for summary
        total_maintenances, all_maintenances_raw, paginated_maintenances_raw = await asyncio.gather(
            prisma.assetsmaintenance.count(where=where_clause),
            prisma.assetsmaintenance.find_many(
                where=where_clause,
                include={
                    "asset": {
                        "include": {
                            "category": True
                        }
                    },
                    "inventoryItems": {
                        "include": {
                            "inventoryItem": True
                        }
                    }
                }
            ),
            prisma.assetsmaintenance.find_many(
                where=where_clause,
                include={
                    "asset": {
                        "include": {
                            "category": True
                        }
                    },
                    "inventoryItems": {
                        "include": {
                            "inventoryItem": True
                        }
                    }
                },
                order={"createdAt": "desc"},
                skip=skip,
                take=pageSize
            )
        )

        # Filter by asset properties (category, location, site, department)
        filtered_all_maintenances = all_maintenances_raw
        
        if category:
            # Filter by category name (case-insensitive)
            filtered_all_maintenances = [
                m for m in filtered_all_maintenances
                if m.asset and m.asset.category and m.asset.category.name and 
                   m.asset.category.name.lower() == category.lower()
            ]
        
        if location:
            filtered_all_maintenances = [
                m for m in filtered_all_maintenances
                if m.asset and m.asset.location and m.asset.location.lower() == location.lower()
            ]
        
        if site:
            filtered_all_maintenances = [
                m for m in filtered_all_maintenances
                if m.asset and m.asset.site and m.asset.site.lower() == site.lower()
            ]
        
        if department:
            filtered_all_maintenances = [
                m for m in filtered_all_maintenances
                if m.asset and m.asset.department and m.asset.department.lower() == department.lower()
            ]

        # Get today's date for calculations
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)

        # Get assets under repair (status: In progress)
        under_repair = [
            m for m in filtered_all_maintenances
            if m.status == 'In progress'
        ]

        # Get upcoming maintenance (Scheduled with dueDate in future)
        upcoming_maintenance = [
            m for m in filtered_all_maintenances
            if m.status == 'Scheduled' and m.dueDate and (
                m.dueDate.replace(tzinfo=None) if m.dueDate.tzinfo else m.dueDate
            ).replace(hour=0, minute=0, second=0, microsecond=0) >= today
        ]

        # Get maintenance history (completed)
        completed_maintenances = [
            m for m in filtered_all_maintenances
            if m.status == 'Completed'
        ]

        # Get maintenances by status for cost calculations
        scheduled_maintenances = [m for m in filtered_all_maintenances if m.status == 'Scheduled']
        cancelled_maintenances = [m for m in filtered_all_maintenances if m.status == 'Cancelled']
        in_progress_maintenances = [m for m in filtered_all_maintenances if m.status == 'In progress']

        # Calculate total costs by status
        total_cost_completed = sum(float(m.cost) if m.cost else 0.0 for m in completed_maintenances)
        total_cost_scheduled = sum(float(m.cost) if m.cost else 0.0 for m in scheduled_maintenances)
        total_cost_cancelled = sum(float(m.cost) if m.cost else 0.0 for m in cancelled_maintenances)
        total_cost_in_progress = sum(float(m.cost) if m.cost else 0.0 for m in in_progress_maintenances)

        # Group by status
        by_status_map: Dict[str, Dict[str, Any]] = {}
        for maintenance in filtered_all_maintenances:
            status_key = maintenance.status or 'Unknown'
            if status_key not in by_status_map:
                by_status_map[status_key] = {
                    "status": status_key,
                    "count": 0,
                    "totalCost": 0.0,
                }
            by_status_map[status_key]["count"] += 1
            by_status_map[status_key]["totalCost"] += float(maintenance.cost) if maintenance.cost else 0.0

        by_status = [
            MaintenanceStatusGroup(
                status=group["status"],
                count=group["count"],
                totalCost=group["totalCost"],
                averageCost=group["totalCost"] / group["count"] if group["count"] > 0 else 0.0,
            )
            for group in by_status_map.values()
        ]

        # Calculate total maintenance costs - only from COMPLETED maintenances
        total_cost = total_cost_completed

        # Calculate average cost per maintenance - only from COMPLETED maintenances
        average_cost = total_cost / len(completed_maintenances) if len(completed_maintenances) > 0 else 0.0

        # Sort upcoming by due date
        sorted_upcoming = sorted(
            upcoming_maintenance,
            key=lambda m: m.dueDate if m.dueDate else datetime.max
        )[:20]

        # Filter paginated maintenances by category if needed
        filtered_maintenances = paginated_maintenances_raw
        # Apply asset property filters to paginated results
        if category:
            filtered_maintenances = [
                m for m in filtered_maintenances
                if m.asset and m.asset.category and m.asset.category.name and 
                   m.asset.category.name.lower() == category.lower()
            ]
        
        if location:
            filtered_maintenances = [
                m for m in filtered_maintenances
                if m.asset and m.asset.location and m.asset.location.lower() == location.lower()
            ]
        
        if site:
            filtered_maintenances = [
                m for m in filtered_maintenances
                if m.asset and m.asset.site and m.asset.site.lower() == site.lower()
            ]
        
        if department:
            filtered_maintenances = [
                m for m in filtered_maintenances
                if m.asset and m.asset.department and m.asset.department.lower() == department.lower()
            ]

        # Format maintenances
        formatted_maintenances = []
        for maintenance in filtered_maintenances:
            due_date = maintenance.dueDate
            due_date_naive = None
            if due_date:
                due_date_naive = due_date.replace(tzinfo=None) if due_date.tzinfo else due_date
                due_date_naive = due_date_naive.replace(hour=0, minute=0, second=0, microsecond=0)

            is_overdue = False
            is_upcoming = False
            if due_date_naive:
                is_overdue = due_date_naive < today and maintenance.status == 'Scheduled'
                is_upcoming = due_date_naive >= today and maintenance.status == 'Scheduled'

            formatted_maintenances.append(
                MaintenanceItem(
                    id=maintenance.id,
                    assetId=maintenance.assetId,
                    assetTagId=maintenance.asset.assetTagId,
                    assetDescription=maintenance.asset.description,
                    assetStatus=maintenance.asset.status,
                    assetCost=float(maintenance.asset.cost) if maintenance.asset.cost else None,
                    category=maintenance.asset.category.name if maintenance.asset.category else None,
                    title=maintenance.title,
                    details=maintenance.details,
                    status=maintenance.status,
                    dueDate=maintenance.dueDate.isoformat().split('T')[0] if maintenance.dueDate else None,
                    dateCompleted=maintenance.dateCompleted.isoformat().split('T')[0] if maintenance.dateCompleted else None,
                    dateCancelled=maintenance.dateCancelled.isoformat().split('T')[0] if maintenance.dateCancelled else None,
                    maintenanceBy=maintenance.maintenanceBy,
                    cost=float(maintenance.cost) if maintenance.cost else None,
                    isRepeating=maintenance.isRepeating,
                    isOverdue=is_overdue,
                    isUpcoming=is_upcoming,
                    inventoryItems=[
                        MaintenanceInventoryItem(
                            id=item.id,
                            inventoryItemId=item.inventoryItemId,
                            quantity=int(item.quantity),
                            unitCost=float(item.unitCost) if item.unitCost else None,
                            inventoryItem={
                                "id": item.inventoryItem.id,
                                "itemCode": item.inventoryItem.itemCode,
                                "name": item.inventoryItem.name,
                                "unit": item.inventoryItem.unit,
                                "unitCost": float(item.inventoryItem.unitCost) if item.inventoryItem.unitCost else None,
                            }
                        )
                        for item in maintenance.inventoryItems
                    ] if maintenance.inventoryItems else None,
                )
            )

        # Format upcoming maintenances
        formatted_upcoming = []
        for maintenance in sorted_upcoming:
            due_date = maintenance.dueDate
            due_date_naive = None
            days_until_due = None
            if due_date:
                due_date_naive = due_date.replace(tzinfo=None) if due_date.tzinfo else due_date
                due_date_naive = due_date_naive.replace(hour=0, minute=0, second=0, microsecond=0)
                if due_date_naive:
                    diff_time = (due_date_naive - today).total_seconds()
                    days_until_due = int(diff_time / (60 * 60 * 24))

            formatted_upcoming.append(
                UpcomingMaintenance(
                    id=maintenance.id,
                    assetId=maintenance.assetId,
                    assetTagId=maintenance.asset.assetTagId,
                    assetDescription=maintenance.asset.description,
                    title=maintenance.title,
                    dueDate=maintenance.dueDate.isoformat().split('T')[0] if maintenance.dueDate else None,
                    maintenanceBy=maintenance.maintenanceBy,
                    daysUntilDue=days_until_due,
                )
            )

        total_pages = (total_maintenances + pageSize - 1) // pageSize if total_maintenances > 0 else 0

        return MaintenanceReportResponse(
            summary=MaintenanceSummary(
                totalMaintenances=len(filtered_all_maintenances),
                underRepair=len(under_repair),
                upcoming=len(upcoming_maintenance),
                completed=len(completed_maintenances),
                totalCost=total_cost,
                averageCost=average_cost,
                totalCostByStatus=TotalCostByStatus(
                    completed=total_cost_completed,
                    scheduled=total_cost_scheduled,
                    cancelled=total_cost_cancelled,
                    inProgress=total_cost_in_progress,
                ),
                byStatus=by_status,
            ),
            maintenances=formatted_maintenances,
            upcoming=formatted_upcoming,
            generatedAt=datetime.now().isoformat(),
            pagination=PaginationInfo(
                total=total_maintenances,
                page=page,
                pageSize=pageSize,
                totalPages=total_pages,
                hasNextPage=page < total_pages,
                hasPreviousPage=page > 1,
            )
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching maintenance reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch maintenance reports")

@router.get("/export")
async def export_maintenance_reports(
    format: str = Query("csv", description="Export format: csv, excel, or pdf"),
    assetId: Optional[str] = Query(None, description="Filter by asset ID"),
    category: Optional[str] = Query(None, description="Filter by category name"),
    location: Optional[str] = Query(None, description="Filter by asset location"),
    site: Optional[str] = Query(None, description="Filter by asset site"),
    department: Optional[str] = Query(None, description="Filter by asset department"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeMaintenanceList: Optional[bool] = Query(False, description="Include maintenance list in export"),
    auth: dict = Depends(verify_auth)
):
    """Export maintenance reports to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Check permission - user must have canManageReports to export reports
        has_permission = await check_permission(user_id, "canManageReports")
        if not has_permission:
            raise HTTPException(
                status_code=403,
                detail="You do not have permission to export reports"
            )

        if format not in ["csv", "excel", "pdf"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv, excel, or pdf.")
        
        if format == "pdf" and not PDF_AVAILABLE:
            raise HTTPException(status_code=500, detail="PDF export not available - fpdf2 not installed")

        # Fetch all data for export
        page_size = 10000 if includeMaintenanceList else 1
        report_data = await get_maintenance_reports(
            assetId=assetId,
            category=category,
            location=location,
            site=site,
            department=department,
            startDate=startDate,
            endDate=endDate,
            page=1,
            pageSize=page_size,
            auth=auth
        )

        summary = report_data.summary
        maintenances = report_data.maintenances

        # Prepare summary statistics
        summary_data = [
            {
                "Metric": "Total Maintenances",
                "Value": str(summary.totalMaintenances),
                "Under Repair": str(summary.underRepair),
                "Upcoming": str(summary.upcoming),
                "Completed": str(summary.completed),
            },
            {
                "Metric": "---",
                "Value": "---",
                "Under Repair": "---",
                "Upcoming": "---",
                "Completed": "---",
            },
            {
                "Metric": "TOTAL COST BY STATUS",
                "Value": "",
                "Under Repair": "",
                "Upcoming": "",
                "Completed": "",
            },
            {
                "Metric": "Total Cost - Completed",
                "Value": format_number(summary.totalCostByStatus.completed),
                "Under Repair": "",
                "Upcoming": "",
                "Completed": "",
            },
            {
                "Metric": "Total Cost - Scheduled",
                "Value": format_number(summary.totalCostByStatus.scheduled),
                "Under Repair": "",
                "Upcoming": "",
                "Completed": "",
            },
            {
                "Metric": "Total Cost - In Progress",
                "Value": format_number(summary.totalCostByStatus.inProgress),
                "Under Repair": "",
                "Upcoming": "",
                "Completed": "",
            },
            {
                "Metric": "Total Cost - Cancelled",
                "Value": format_number(summary.totalCostByStatus.cancelled),
                "Under Repair": "",
                "Upcoming": "",
                "Completed": "",
            },
            {
                "Metric": "---",
                "Value": "---",
                "Under Repair": "---",
                "Upcoming": "---",
                "Completed": "---",
            },
            {
                "Metric": "MAINTENANCES BY STATUS",
                "Value": "",
                "Under Repair": "",
                "Upcoming": "",
                "Completed": "",
            },
            *[
                {
                    "Metric": f"Status: {status_item.status}",
                    "Value": str(status_item.count),
                    "Under Repair": "",
                    "Upcoming": "",
                    "Completed": "",
                    "Total Cost": format_number(status_item.totalCost) if status_item.totalCost > 0 else "-",
                    "Average Cost": format_number(status_item.averageCost) if status_item.totalCost > 0 else "-",
                }
                for status_item in summary.byStatus
            ],
        ]

        # Prepare maintenance list data
        maintenance_list_data = []
        for maintenance in maintenances:
            # Format inventory items as a string
            inventory_items_str = ""
            if maintenance.inventoryItems:
                inventory_items_str = "; ".join([
                    f"{item.inventoryItem.get('itemCode', '')} ({item.quantity} {item.inventoryItem.get('unit', 'units')})"
                    for item in maintenance.inventoryItems
                ])
            
            # Format inventory items count
            inventory_items_count = len(maintenance.inventoryItems) if maintenance.inventoryItems else 0
            
            # Calculate total inventory cost
            total_inventory_cost = 0.0
            if maintenance.inventoryItems:
                for item in maintenance.inventoryItems:
                    item_cost = (item.unitCost or item.inventoryItem.get('unitCost') or 0.0) * item.quantity
                    total_inventory_cost += item_cost

            maintenance_list_data.append({
                "Asset Tag ID": maintenance.assetTagId or "",
                "Asset Description": maintenance.assetDescription or "",
                "Category": maintenance.category or "",
                "Asset Status": maintenance.assetStatus or "",
                "Asset Cost": format_number(maintenance.assetCost) if maintenance.assetCost else "",
                "Title": maintenance.title or "",
                "Details": maintenance.details or "",
                "Status": maintenance.status or "",
                "Due Date": maintenance.dueDate or "",
                "Date Completed": maintenance.dateCompleted or "",
                "Date Cancelled": maintenance.dateCancelled or "",
                "Cost": format_number(maintenance.cost) if maintenance.cost else "",
                "Inventory Items Count": str(inventory_items_count),
                "Inventory Items": inventory_items_str,
                "Total Inventory Cost": format_number(total_inventory_cost),
                "Is Repeating": "Yes" if maintenance.isRepeating else "No",
                "Is Overdue": "Yes" if maintenance.isOverdue else "No",
                "Is Upcoming": "Yes" if maintenance.isUpcoming else "No",
            })

        filename = f"maintenance-report-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)

            if includeMaintenanceList:
                # Include summary and maintenance list
                writer.writerow(["=== SUMMARY STATISTICS ==="])
                if summary_data:
                    summary_headers = list(summary_data[0].keys())
                    writer.writerow(summary_headers)
                    for row in summary_data:
                        writer.writerow([row.get(header, "") for header in summary_headers])
                
                writer.writerow([])
                writer.writerow(["=== MAINTENANCE LIST ==="])
                
                if maintenance_list_data:
                    maintenance_headers = list(maintenance_list_data[0].keys())
                    writer.writerow(maintenance_headers)
                    for row in maintenance_list_data:
                        writer.writerow([row.get(header, "") for header in maintenance_headers])
            else:
                # Summary only
                if summary_data:
                    summary_headers = list(summary_data[0].keys())
                    writer.writerow(summary_headers)
                    for row in summary_data:
                        writer.writerow([row.get(header, "") for header in summary_headers])

            csv_content = output.getvalue()
            output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        elif format == "excel":
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Summary sheet
            if summary_data:
                summary_ws = wb.create_sheet("Summary")
                summary_headers = list(summary_data[0].keys())
                summary_ws.append(summary_headers)
                for row in summary_data:
                    summary_ws.append([row.get(header, "") for header in summary_headers])

            if includeMaintenanceList:
                # By Status sheet
                status_data = [row for row in summary_data if row.get("Metric", "").startswith("Status:")]
                if status_data:
                    status_ws = wb.create_sheet("By Status")
                    status_headers = list(status_data[0].keys())
                    status_ws.append(status_headers)
                    for row in status_data:
                        status_ws.append([row.get(header, "") for header in status_headers])

                # Maintenance List sheet
                if maintenance_list_data:
                    maintenance_list_ws = wb.create_sheet("Maintenance List")
                    maintenance_headers = list(maintenance_list_data[0].keys())
                    maintenance_list_ws.append(maintenance_headers)
                    for row in maintenance_list_data:
                        maintenance_list_ws.append([row.get(header, "") for header in maintenance_headers])
            else:
                # Single sheet export
                if not summary_data:
                    raise HTTPException(status_code=400, detail="No data to export")

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # pdf
            pdf = ReportPDF("Maintenance Report", "Maintenance")
            pdf.add_page()

            pdf.add_section_title("Summary Statistics")
            if summary_data:
                headers = list(summary_data[0].keys())
                rows = [[str(row.get(h, '')) for h in headers] for row in summary_data]
                pdf.add_table(headers, rows)
            
            pdf.ln(10)

            if maintenance_list_data and includeMaintenanceList:
                pdf.add_section_title(f"Maintenance List ({len(maintenance_list_data)} records)")
                simplified_headers = ["Asset Tag", "Description", "Title", "Status", "Due Date", "Completed", "Cost", "Inventory Items"]
                simplified_rows = [
                    [
                        str(m.get("Asset Tag ID", "")),
                        str(m.get("Asset Description", ""))[:50],
                        str(m.get("Title", "")),
                        str(m.get("Status", "")),
                        str(m.get("Due Date", "")),
                        str(m.get("Date Completed", "")),
                        str(m.get("Cost", "")),
                        str(m.get("Inventory Items", ""))[:40],
                    ]
                    for m in maintenance_list_data
                ]
                pdf.add_table(simplified_headers, simplified_rows)

            pdf_content = bytes(pdf.output())
            filename += ".pdf"
            return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting maintenance reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export maintenance reports")

