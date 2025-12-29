"""
Reservation Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv

from models.reports import ReservationReportResponse, ReservationItem, PaginationInfo
from auth import verify_auth
from database import prisma
from utils.pdf_generator import ReportPDF, PDF_AVAILABLE

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/reservation", tags=["reports"])

async def check_permission(user_id: str, permission: str) -> bool:
    """Check if user has a specific permission. Admins have all permissions."""
    try:
        asset_user = await prisma.assetuser.find_unique(
            where={"userId": user_id}
        )
        if not asset_user or not asset_user.isActive:
            return False
        
        # Admins have all permissions
        if asset_user.role == "admin":
            return True
        
        return getattr(asset_user, permission, False)
    except Exception:
        return False

def format_number(value: Optional[float]) -> str:
    """Format number with commas and 2 decimal places"""
    if value is None or value == 0:
        return '0.00'
    return f"{float(value):,.2f}"

@router.get("", response_model=ReservationReportResponse)
async def get_reservation_reports(
    category: Optional[str] = Query(None, description="Filter by category name"),
    reservationType: Optional[str] = Query(None, description="Filter by reservation type (Employee or Department)"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    employeeId: Optional[str] = Query(None, description="Filter by employee ID"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get reservation reports with optional filters and pagination"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        skip = (page - 1) * pageSize

        # Build where clause
        where_clause: Dict[str, Any] = {
            "asset": {
                "isDeleted": False,
            }
        }

        # Category filter
        if category:
            where_clause["asset"] = {
                **where_clause["asset"],
                "category": {
                    "name": category
                }
            }

        # Reservation type filter
        if reservationType:
            where_clause["reservationType"] = reservationType

        # Location filter
        if location:
            where_clause["asset"] = {
                **where_clause["asset"],
                "location": location
            }

        # Site filter
        if site:
            where_clause["asset"] = {
                **where_clause["asset"],
                "site": site
            }

        # Department filter
        if department:
            where_clause["department"] = department

        # Employee filter
        if employeeId:
            where_clause["employeeUserId"] = employeeId

        # Date range filter (reservation date)
        if startDate or endDate:
            reservation_date_filter: Dict[str, Any] = {}
            if startDate:
                date_obj = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                reservation_date_filter["gte"] = date_obj
            if endDate:
                date_obj = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                reservation_date_filter["lte"] = date_obj
            if reservation_date_filter:
                where_clause["reservationDate"] = reservation_date_filter

        # Get total count
        total = await prisma.assetsreserve.count(where=where_clause)

        # Get paginated reservations
        reservations_raw = await prisma.assetsreserve.find_many(
            where=where_clause,
            include={
                "asset": {
                    "include": {
                        "category": True,
                        "subCategory": True
                    }
                },
                "employeeUser": True
            },
            order={"reservationDate": "desc"},
            skip=skip,
            take=pageSize
        )

        # Format reservation data
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
        
        formatted_reservations = []
        for reservation in reservations_raw:
            reservation_date = reservation.reservationDate
            reservation_date_naive = None
            if reservation_date:
                reservation_date_naive = reservation_date.replace(tzinfo=None) if reservation_date.tzinfo else reservation_date
                reservation_date_naive = reservation_date_naive.replace(hour=0, minute=0, second=0, microsecond=0)

            # Calculate reservation status
            reservation_status = 'upcoming'
            if reservation_date_naive:
                if reservation_date_naive < today:
                    reservation_status = 'past'
                elif reservation_date_naive == today:
                    reservation_status = 'today'

            # Calculate days until/from reservation
            days_until = 0
            if reservation_date_naive:
                diff_time = (reservation_date_naive - today).total_seconds()
                days_until = int(diff_time / (60 * 60 * 24))

            formatted_reservations.append(
                ReservationItem(
                    id=reservation.id,
                    assetTagId=reservation.asset.assetTagId,
                    description=reservation.asset.description,
                    category=reservation.asset.category.name if reservation.asset.category else None,
                    subCategory=reservation.asset.subCategory.name if reservation.asset.subCategory else None,
                    reservationType=reservation.reservationType,
                    reservationDate=reservation.reservationDate.isoformat(),
                    purpose=reservation.purpose,
                    notes=reservation.notes,
                    location=reservation.asset.location,
                    site=reservation.asset.site,
                    assetStatus=reservation.asset.status,
                    assetCost=float(reservation.asset.cost) if reservation.asset.cost else None,
                    department=reservation.department,
                    employeeName=reservation.employeeUser.name if reservation.employeeUser else None,
                    employeeEmail=reservation.employeeUser.email if reservation.employeeUser else None,
                    reservationStatus=reservation_status,
                    daysUntil=days_until,
                    createdAt=reservation.createdAt.isoformat(),
                )
            )

        total_pages = (total + pageSize - 1) // pageSize if total > 0 else 0

        return ReservationReportResponse(
            reservations=formatted_reservations,
            pagination=PaginationInfo(
                total=total,
                page=page,
                pageSize=pageSize,
                totalPages=total_pages,
                hasNextPage=page < total_pages,
                hasPreviousPage=page > 1,
            )
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching reservation reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch reservation reports")

@router.get("/export")
async def export_reservation_reports(
    format: str = Query("csv", description="Export format: csv or excel"),
    category: Optional[str] = Query(None, description="Filter by category name"),
    reservationType: Optional[str] = Query(None, description="Filter by reservation type (Employee or Department)"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    employeeId: Optional[str] = Query(None, description="Filter by employee ID"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeReservationList: Optional[bool] = Query(False, description="Include reservation list in export"),
    auth: dict = Depends(verify_auth)
):
    """Export reservation reports to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Check permission - user must have canManageReports to export reports
        has_permission = await check_permission(user_id, "canManageReports")
        if not has_permission:
            raise HTTPException(
                status_code=403,
                detail="You do not have permission to export reports"
            )

        if format not in ["csv", "excel", "pdf"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv, excel, or pdf.")
        
        if format == "pdf" and not PDF_AVAILABLE:
            raise HTTPException(status_code=500, detail="PDF export not available - fpdf2 not installed")

        # Fetch all data for export
        page_size = 10000 if includeReservationList else 1
        report_data = await get_reservation_reports(
            category=category,
            reservationType=reservationType,
            location=location,
            site=site,
            department=department,
            employeeId=employeeId,
            startDate=startDate,
            endDate=endDate,
            page=1,
            pageSize=page_size,
            auth=auth
        )

        reservations = report_data.reservations

        # Calculate summary statistics
        upcoming_reservations = [r for r in reservations if r.reservationStatus == 'upcoming']
        today_reservations = [r for r in reservations if r.reservationStatus == 'today']
        past_reservations = [r for r in reservations if r.reservationStatus == 'past']
        employee_reservations = [r for r in reservations if r.reservationType == 'Employee']
        department_reservations = [r for r in reservations if r.reservationType == 'Department']
        total_asset_value = sum(r.assetCost or 0 for r in reservations)

        # Group by type
        by_type: Dict[str, Dict[str, Any]] = {}
        for reservation in reservations:
            res_type = reservation.reservationType
            if res_type not in by_type:
                by_type[res_type] = {
                    "count": 0,
                    "totalValue": 0.0,
                }
            by_type[res_type]["count"] += 1
            by_type[res_type]["totalValue"] += reservation.assetCost or 0

        filename = f"reservation-report-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)

            if includeReservationList:
                # Include summary and reservation list
                writer.writerow(["RESERVATION REPORT SUMMARY"])
                writer.writerow(["Total Reservations", len(reservations)])
                writer.writerow(["Upcoming", len(upcoming_reservations)])
                writer.writerow(["Today", len(today_reservations)])
                writer.writerow(["Past", len(past_reservations)])
                writer.writerow(["Employee Reservations", len(employee_reservations)])
                writer.writerow(["Department Reservations", len(department_reservations)])
                writer.writerow(["Total Asset Value", format_number(total_asset_value)])
                writer.writerow([])
                
                writer.writerow(["RESERVATIONS BY TYPE"])
                writer.writerow(["Reservation Type", "Count", "Total Asset Value"])
                for res_type, stats in by_type.items():
                    writer.writerow([
                        res_type,
                        stats["count"],
                        format_number(stats["totalValue"]),
                    ])
                writer.writerow([])
                
                writer.writerow(["RESERVATION RECORDS"])
                writer.writerow([
                    "Asset Tag ID",
                    "Description",
                    "Category",
                    "Sub-Category",
                    "Reservation Type",
                    "Reserved By",
                    "Reservation Date",
                    "Purpose",
                    "Status",
                    "Days Until/From",
                    "Location",
                    "Site",
                    "Asset Cost",
                ])
                for reservation in reservations:
                    reserved_by = reservation.employeeName if reservation.reservationType == 'Employee' else reservation.department
                    reserved_by = reserved_by or 'N/A'
                    days_text = (
                        f"{reservation.daysUntil} days until" if reservation.daysUntil > 0
                        else "Today" if reservation.daysUntil == 0
                        else f"{abs(reservation.daysUntil)} days ago"
                    )
                    writer.writerow([
                        reservation.assetTagId,
                        reservation.description,
                        reservation.category or "N/A",
                        reservation.subCategory or "N/A",
                        reservation.reservationType,
                        reserved_by,
                        reservation.reservationDate.split('T')[0] if reservation.reservationDate else "N/A",
                        reservation.purpose or "N/A",
                        reservation.reservationStatus,
                        days_text,
                        reservation.location or "N/A",
                        reservation.site or "N/A",
                        format_number(reservation.assetCost),
                    ])
            else:
                # Summary only
                writer.writerow(["RESERVATION REPORT SUMMARY"])
                writer.writerow(["Total Reservations", len(reservations)])
                writer.writerow(["Upcoming", len(upcoming_reservations)])
                writer.writerow(["Today", len(today_reservations)])
                writer.writerow(["Past", len(past_reservations)])
                writer.writerow(["Employee Reservations", len(employee_reservations)])
                writer.writerow(["Department Reservations", len(department_reservations)])
                writer.writerow(["Total Asset Value", format_number(total_asset_value)])
                writer.writerow([])
                writer.writerow(["RESERVATIONS BY TYPE"])
                writer.writerow(["Reservation Type", "Count", "Total Asset Value"])
                for res_type, stats in by_type.items():
                    writer.writerow([
                        res_type,
                        stats["count"],
                        format_number(stats["totalValue"]),
                    ])

            csv_content = output.getvalue()
            output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        elif format == "excel":
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Summary sheet
            summary_data = [
                ["RESERVATION REPORT SUMMARY"],
                ["Total Reservations", len(reservations)],
                ["Upcoming", len(upcoming_reservations)],
                ["Today", len(today_reservations)],
                ["Past", len(past_reservations)],
                ["Employee Reservations", len(employee_reservations)],
                ["Department Reservations", len(department_reservations)],
                ["Total Asset Value", format_number(total_asset_value)],
                [],
                ["RESERVATIONS BY TYPE"],
                ["Reservation Type", "Count", "Total Asset Value"],
                *[[res_type, stats["count"], format_number(stats["totalValue"])] for res_type, stats in by_type.items()],
            ]
            summary_ws = wb.create_sheet("Summary")
            for row in summary_data:
                summary_ws.append(row)

            if includeReservationList:
                # Reservation list sheet
                reservation_data = []
                for reservation in reservations:
                    reserved_by = reservation.employeeName if reservation.reservationType == 'Employee' else reservation.department
                    reserved_by = reserved_by or 'N/A'
                    days_text = (
                        f"{reservation.daysUntil} days until" if reservation.daysUntil > 0
                        else "Today" if reservation.daysUntil == 0
                        else f"{abs(reservation.daysUntil)} days ago"
                    )
                    reservation_data.append([
                        reservation.assetTagId,
                        reservation.description,
                        reservation.category or "N/A",
                        reservation.subCategory or "N/A",
                        reservation.reservationType,
                        reserved_by,
                        reservation.reservationDate.split('T')[0] if reservation.reservationDate else "N/A",
                        reservation.purpose or "N/A",
                        reservation.reservationStatus,
                        days_text,
                        reservation.location or "N/A",
                        reservation.site or "N/A",
                        format_number(reservation.assetCost),
                    ])
                
                reservation_ws = wb.create_sheet("Reservation List")
                reservation_ws.append([
                    "Asset Tag ID",
                    "Description",
                    "Category",
                    "Sub-Category",
                    "Reservation Type",
                    "Reserved By",
                    "Reservation Date",
                    "Purpose",
                    "Status",
                    "Days Until/From",
                    "Location",
                    "Site",
                    "Asset Cost",
                ])
                for row in reservation_data:
                    reservation_ws.append(row)

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # pdf
            pdf = ReportPDF("Reservation Report", "Reservation")
            pdf.add_page()

            pdf.add_section_title("Summary Statistics")
            summary_rows = [
                ["Total Reservations", str(len(reservations))],
                ["Upcoming", str(len(upcoming_reservations))],
                ["Today", str(len(today_reservations))],
                ["Past", str(len(past_reservations))],
                ["Employee Reservations", str(len(employee_reservations))],
                ["Department Reservations", str(len(department_reservations))],
                ["Total Asset Value", format_number(total_asset_value)],
            ]
            if summary_rows:
                headers = ["Metric", "Value"]
                pdf.add_table(headers, summary_rows)
            
            pdf.ln(10)

            if reservations and includeReservationList:
                pdf.add_section_title(f"Reservation List ({len(reservations)} reservations)")
                simplified_headers = ["Asset Tag ID", "Description", "Category", "Reservation Type", "Reserved By", "Reservation Date", "Status", "Days Until/From", "Asset Cost"]
                simplified_rows = [
                    [
                        str(r.assetTagId or ""),
                        str(r.description or "")[:50],
                        str(r.category or ""),
                        str(r.reservationType or ""),
                        str(r.employeeName or r.department or ""),
                        r.reservationDate.split('T')[0] if r.reservationDate else "",
                        str(r.reservationStatus or ""),
                        f"{r.daysUntil} days" if r.daysUntil != 0 else "Today",
                        format_number(r.assetCost),
                    ]
                    for r in reservations
                ]
                pdf.add_table(simplified_headers, simplified_rows)

            pdf_content = bytes(pdf.output())
            filename += ".pdf"
            return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting reservation reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export reservation reports")

