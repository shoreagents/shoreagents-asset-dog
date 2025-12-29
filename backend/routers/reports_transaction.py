"""
Transaction Reports API router
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional, Dict, Any, List
from datetime import datetime
import logging
import io
import csv
import asyncio

from models.reports import TransactionReportResponse, TransactionSummary, TransactionTypeGroup, TransactionItem, PaginationInfo
from auth import verify_auth
from database import prisma
from utils.pdf_generator import ReportPDF, PDF_AVAILABLE

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/transaction", tags=["reports"])

async def check_permission(user_id: str, permission: str) -> bool:
    """Check if user has a specific permission. Admins have all permissions."""
    try:
        asset_user = await prisma.assetuser.find_unique(
            where={"userId": user_id}
        )
        if not asset_user or not asset_user.isActive:
            return False
        
        # Admins have all permissions
        if asset_user.role == "admin":
            return True
        
        return getattr(asset_user, permission, False)
    except Exception:
        return False

def format_number(value: Optional[float]) -> str:
    """Format number with commas and 2 decimal places"""
    if value is None or value == 0:
        return '0.00'
    return f"{float(value):,.2f}"

async def _fetch_add_asset_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Add Asset transactions"""
    if transaction_type and transaction_type != 'Add Asset':
        return []
    
    where_clause: Dict[str, Any] = {
        "isDeleted": False,
    }
    
    if date_filter:
        where_clause["createdAt"] = date_filter
    if category:
        where_clause["category"] = {"name": category}
    if location:
        where_clause["location"] = location
    if site:
        where_clause["site"] = site
    if department:
        where_clause["department"] = department

    added_assets = await prisma.assets.find_many(
        where=where_clause,
        include={"category": True, "subCategory": True},
        order={"createdAt": "desc"},
        take=5000
    )

    # Get history logs for actionBy
    asset_ids = [a.id for a in added_assets]
    history_logs = []
    if asset_ids:
        history_logs = await prisma.assetshistorylogs.find_many(
            where={
                "assetId": {"in": asset_ids},
                "eventType": "added",
            }
        )

    history_map = {h.assetId: h.actionBy for h in history_logs}

    return [
        TransactionItem(
            id=f"add-{asset.id}",
            transactionType="Add Asset",
            assetTagId=asset.assetTagId,
            assetDescription=asset.description,
            category=asset.category.name if asset.category else None,
            subCategory=asset.subCategory.name if asset.subCategory else None,
            transactionDate=asset.createdAt.isoformat(),
            actionBy=history_map.get(asset.id),
            details="Asset added to system",
            location=asset.location,
            site=asset.site,
            department=asset.department,
            assetCost=float(asset.cost) if asset.cost else None,
        )
        for asset in added_assets
    ]

async def _fetch_edit_asset_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    action_by: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Edit Asset transactions"""
    if transaction_type and transaction_type != 'Edit Asset':
        return []
    
    where_clause: Dict[str, Any] = {
        "eventType": "edited",
        "asset": {"isDeleted": False}
    }
    
    if date_filter:
        where_clause["eventDate"] = date_filter
    if action_by:
        where_clause["actionBy"] = {"contains": action_by, "mode": "insensitive"}
    if category:
        where_clause["asset"]["category"] = {"name": category}
    if location:
        where_clause["asset"]["location"] = location
    if site:
        where_clause["asset"]["site"] = site
    if department:
        where_clause["asset"]["department"] = department

    edit_logs = await prisma.assetshistorylogs.find_many(
        where=where_clause,
        include={
            "asset": {
                "include": {"category": True, "subCategory": True}
            }
        },
        order={"eventDate": "desc"},
        take=5000
    )

    return [
        TransactionItem(
            id=f"edit-{log.id}",
            transactionType="Edit Asset",
            assetTagId=log.asset.assetTagId,
            assetDescription=log.asset.description,
            category=log.asset.category.name if log.asset.category else None,
            subCategory=log.asset.subCategory.name if log.asset.subCategory else None,
            transactionDate=log.eventDate.isoformat(),
            actionBy=log.actionBy,
            details=f'Field "{log.field}" changed from "{log.changeFrom or "N/A"}" to "{log.changeTo or "N/A"}"' if log.field else "Asset edited",
            fieldChanged=log.field,
            oldValue=log.changeFrom,
            newValue=log.changeTo,
            location=log.asset.location,
            site=log.asset.site,
            department=log.asset.department,
            assetCost=float(log.asset.cost) if log.asset.cost else None,
        )
        for log in edit_logs
    ]

async def _fetch_delete_asset_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Delete Asset transactions"""
    if transaction_type and transaction_type != 'Delete Asset':
        return []
    
    where_clause: Dict[str, Any] = {
        "isDeleted": True,
    }
    
    if category:
        where_clause["category"] = {"name": category}
    if location:
        where_clause["location"] = location
    if site:
        where_clause["site"] = site
    if department:
        where_clause["department"] = department
    if date_filter:
        where_clause["deletedAt"] = date_filter

    deleted_assets = await prisma.assets.find_many(
        where=where_clause,
        include={"category": True, "subCategory": True},
        order={"deletedAt": "desc"},
        take=5000
    )

    # Get history logs for deleted assets
    asset_ids = [a.id for a in deleted_assets]
    delete_logs = []
    if asset_ids:
        delete_logs = await prisma.assetshistorylogs.find_many(
            where={
                "assetId": {"in": asset_ids},
                "eventType": "deleted",
            }
        )

    log_map = {log.assetId: log for log in delete_logs}

    results = []
    for asset in deleted_assets:
        log = log_map.get(asset.id)
        delete_date = asset.deletedAt
        if not delete_date and log:
            delete_date = log.eventDate
        if not delete_date:
            delete_date = datetime.now()
        
        # Convert to timezone-naive if needed
        if delete_date and hasattr(delete_date, 'tzinfo') and delete_date.tzinfo:
            delete_date = delete_date.replace(tzinfo=None)
        
        results.append(
            TransactionItem(
                id=f"delete-{asset.id}",
                transactionType="Delete Asset",
                assetTagId=asset.assetTagId,
                assetDescription=asset.description,
                category=asset.category.name if asset.category else None,
                subCategory=asset.subCategory.name if asset.subCategory else None,
                transactionDate=delete_date.isoformat() if isinstance(delete_date, datetime) else str(delete_date),
                actionBy=log.actionBy if log else None,
                details="Asset deleted",
                location=asset.location,
                site=asset.site,
                department=asset.department,
                assetCost=float(asset.cost) if asset.cost else None,
            )
        )
    
    return results

async def _fetch_disposal_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch disposal transactions (Sold, Donated, Scrapped, Lost/Missing, Destroyed)"""
    disposal_types = {
        'Sold': 'Sold Asset',
        'Donated': 'Donated Asset',
        'Scrapped': 'Scrapped Asset',
        'Lost/Missing': 'Lost/Missing Asset',
        'Destroyed': 'Destroyed Asset',
    }

    results: List[TransactionItem] = []

    for method, trans_type in disposal_types.items():
        if transaction_type and transaction_type != trans_type:
            continue

        where_clause: Dict[str, Any] = {
            "disposalMethod": method,
            "asset": {"isDeleted": False}
        }
        
        if date_filter:
            where_clause["disposeDate"] = date_filter
        if category:
            where_clause["asset"]["category"] = {"name": category}
        if location:
            where_clause["asset"]["location"] = location
        if site:
            where_clause["asset"]["site"] = site
        if department:
            where_clause["asset"]["department"] = department

        disposals = await prisma.assetsdispose.find_many(
            where=where_clause,
            include={
                "asset": {
                    "include": {"category": True, "subCategory": True}
                }
            },
            order={"disposeDate": "desc"},
            take=1000
        )

        for disposal in disposals:
            results.append(
                TransactionItem(
                    id=f"dispose-{disposal.id}",
                    transactionType=trans_type,
                    assetTagId=disposal.asset.assetTagId,
                    assetDescription=disposal.asset.description,
                    category=disposal.asset.category.name if disposal.asset.category else None,
                    subCategory=disposal.asset.subCategory.name if disposal.asset.subCategory else None,
                    transactionDate=disposal.disposeDate.isoformat(),
                    actionBy=None,
                    details=disposal.disposeReason or f"Asset {method.lower()}",
                    location=disposal.asset.location,
                    site=disposal.asset.site,
                    department=disposal.asset.department,
                    assetCost=float(disposal.asset.cost) if disposal.asset.cost else None,
                    disposeDate=disposal.disposeDate.isoformat(),
                    disposeReason=disposal.disposeReason,
                    disposeValue=float(disposal.disposeValue) if disposal.disposeValue else None,
                )
            )

    return results

async def _fetch_lease_out_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Lease Out transactions"""
    if transaction_type and transaction_type != 'Lease Out':
        return []
    
    where_clause: Dict[str, Any] = {
        "asset": {"isDeleted": False}
    }
    
    if date_filter:
        where_clause["leaseStartDate"] = date_filter
    if category:
        where_clause["asset"]["category"] = {"name": category}
    if location:
        where_clause["asset"]["location"] = location
    if site:
        where_clause["asset"]["site"] = site
    if department:
        where_clause["asset"]["department"] = department

    leases = await prisma.assetslease.find_many(
        where=where_clause,
        include={
            "asset": {
                "include": {"category": True, "subCategory": True}
            }
        },
        order={"leaseStartDate": "desc"},
        take=5000
    )

    return [
        TransactionItem(
            id=f"lease-{lease.id}",
            transactionType="Lease Out",
            assetTagId=lease.asset.assetTagId,
            assetDescription=lease.asset.description,
            category=lease.asset.category.name if lease.asset.category else None,
            subCategory=lease.asset.subCategory.name if lease.asset.subCategory else None,
            transactionDate=lease.leaseStartDate.isoformat(),
            actionBy=None,
            details=f"Leased to {lease.lessee}",
            location=lease.asset.location,
            site=lease.asset.site,
            department=lease.asset.department,
            assetCost=float(lease.asset.cost) if lease.asset.cost else None,
            lessee=lease.lessee,
            leaseStartDate=lease.leaseStartDate.isoformat(),
            leaseEndDate=lease.leaseEndDate.isoformat() if lease.leaseEndDate else None,
            conditions=lease.conditions,
        )
        for lease in leases
    ]

async def _fetch_lease_return_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Lease Return transactions"""
    if transaction_type and transaction_type != 'Lease Return':
        return []
    
    where_clause: Dict[str, Any] = {
        "asset": {"isDeleted": False}
    }
    
    if date_filter:
        where_clause["returnDate"] = date_filter
    if category:
        where_clause["asset"]["category"] = {"name": category}
    if location:
        where_clause["asset"]["location"] = location
    if site:
        where_clause["asset"]["site"] = site
    if department:
        where_clause["asset"]["department"] = department

    returns = await prisma.assetsleasereturn.find_many(
        where=where_clause,
        include={
            "asset": {
                "include": {"category": True, "subCategory": True}
            },
            "lease": True
        },
        order={"returnDate": "desc"},
        take=5000
    )

    return [
        TransactionItem(
            id=f"lease-return-{return_record.id}",
            transactionType="Lease Return",
            assetTagId=return_record.asset.assetTagId,
            assetDescription=return_record.asset.description,
            category=return_record.asset.category.name if return_record.asset.category else None,
            subCategory=return_record.asset.subCategory.name if return_record.asset.subCategory else None,
            transactionDate=return_record.returnDate.isoformat(),
            actionBy=None,
            details=f"Returned from {return_record.lease.lessee}",
            location=return_record.asset.location,
            site=return_record.asset.site,
            department=return_record.asset.department,
            assetCost=float(return_record.asset.cost) if return_record.asset.cost else None,
            lessee=return_record.lease.lessee,
            returnDate=return_record.returnDate.isoformat(),
            condition=return_record.condition,
            notes=return_record.notes,
        )
        for return_record in returns
    ]

async def _fetch_repair_asset_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Repair Asset transactions"""
    if transaction_type and transaction_type != 'Repair Asset':
        return []
    
    where_clause: Dict[str, Any] = {
        "asset": {"isDeleted": False}
    }
    
    if date_filter:
        where_clause["createdAt"] = date_filter
    if category:
        where_clause["asset"]["category"] = {"name": category}
    if location:
        where_clause["asset"]["location"] = location
    if site:
        where_clause["asset"]["site"] = site
    if department:
        where_clause["asset"]["department"] = department

    maintenances = await prisma.assetsmaintenance.find_many(
        where=where_clause,
        include={
            "asset": {
                "include": {"category": True, "subCategory": True}
            }
        },
        order={"createdAt": "desc"},
        take=5000
    )

    return [
        TransactionItem(
            id=f"maintenance-{maintenance.id}",
            transactionType="Repair Asset",
            assetTagId=maintenance.asset.assetTagId,
            assetDescription=maintenance.asset.description,
            category=maintenance.asset.category.name if maintenance.asset.category else None,
            subCategory=maintenance.asset.subCategory.name if maintenance.asset.subCategory else None,
            transactionDate=maintenance.createdAt.isoformat(),
            actionBy=maintenance.maintenanceBy,
            details=maintenance.title,
            location=maintenance.asset.location,
            site=maintenance.asset.site,
            department=maintenance.asset.department,
            assetCost=float(maintenance.asset.cost) if maintenance.asset.cost else None,
            title=maintenance.title,
            maintenanceBy=maintenance.maintenanceBy,
            dueDate=maintenance.dueDate.isoformat() if maintenance.dueDate else None,
            status=maintenance.status,
            cost=float(maintenance.cost) if maintenance.cost else None,
            dateCompleted=maintenance.dateCompleted.isoformat() if maintenance.dateCompleted else None,
        )
        for maintenance in maintenances
    ]

async def _fetch_move_asset_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Move Asset transactions"""
    if transaction_type and transaction_type != 'Move Asset':
        return []
    
    where_clause: Dict[str, Any] = {
        "asset": {"isDeleted": False}
    }
    
    if date_filter:
        where_clause["moveDate"] = date_filter
    if category:
        where_clause["asset"]["category"] = {"name": category}
    if location:
        where_clause["asset"]["location"] = location
    if site:
        where_clause["asset"]["site"] = site
    if department:
        where_clause["asset"]["department"] = department

    moves = await prisma.assetsmove.find_many(
        where=where_clause,
        include={
            "asset": {
                "include": {"category": True, "subCategory": True}
            },
            "employeeUser": True
        },
        order={"moveDate": "desc"},
        take=5000
    )

    # Get history logs for moves to find actionBy and fromLocation
    asset_ids = [m.assetId for m in moves]
    move_history_logs = []
    if asset_ids:
        move_history_logs = await prisma.assetshistorylogs.find_many(
            where={
                "assetId": {"in": asset_ids},
                "eventType": "edited",
                "field": {"in": ["location", "department"]},
            },
            order={"eventDate": "desc"}
        )

    # Create maps for quick lookup
    action_by_map: Dict[str, str] = {}
    from_location_map: Dict[str, str] = {}
    
    # Match history logs to moves by assetId and date proximity (within 1 day)
    for move in moves:
        move_date_naive = move.moveDate.replace(tzinfo=None) if move.moveDate.tzinfo else move.moveDate
        relevant_logs = [
            log for log in move_history_logs
            if log.assetId == move.assetId
        ]
        # Filter by date proximity
        relevant_logs = [
            log for log in relevant_logs
            if abs((log.eventDate.replace(tzinfo=None) if log.eventDate.tzinfo else log.eventDate) - move_date_naive).total_seconds() < 86400
        ]
        
        if relevant_logs:
            matching_log = next(
                (log for log in relevant_logs if 
                 (move.moveType == 'Location Transfer' and log.field == 'location') or
                 (move.moveType == 'Department Transfer' and log.field == 'department')),
                relevant_logs[0]
            )
            
            if matching_log:
                action_by_map[move.id] = matching_log.actionBy
                if move.moveType == 'Location Transfer' and matching_log.field == 'location':
                    from_location_map[move.id] = matching_log.changeFrom or ''

    results = []
    for move in moves:
        to_location = None
        if move.moveType == 'Location Transfer':
            to_location = move.asset.location
        elif move.moveType == 'Department Transfer':
            to_location = move.asset.department

        results.append(
            TransactionItem(
                id=f"move-{move.id}",
                transactionType="Move Asset",
                assetTagId=move.asset.assetTagId,
                assetDescription=move.asset.description,
                category=move.asset.category.name if move.asset.category else None,
                subCategory=move.asset.subCategory.name if move.asset.subCategory else None,
                transactionDate=move.moveDate.isoformat(),
                actionBy=action_by_map.get(move.id),
                details=f"{move.moveType}: {move.reason or 'No reason provided'}",
                location=move.asset.location,
                site=move.asset.site,
                department=move.asset.department,
                assetCost=float(move.asset.cost) if move.asset.cost else None,
                moveType=move.moveType,
                moveDate=move.moveDate.isoformat(),
                employeeName=move.employeeUser.name if move.employeeUser else None,
                reason=move.reason,
                fromLocation=from_location_map.get(move.id),
                toLocation=to_location,
            )
        )

    return results

async def _fetch_checkout_asset_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Checkout Asset transactions"""
    if transaction_type and transaction_type != 'Checkout Asset':
        return []
    
    where_clause: Dict[str, Any] = {
        "asset": {"isDeleted": False}
    }
    
    if date_filter:
        where_clause["checkoutDate"] = date_filter
    if category:
        where_clause["asset"]["category"] = {"name": category}
    if location:
        where_clause["asset"]["location"] = location
    if site:
        where_clause["asset"]["site"] = site
    if department:
        where_clause["asset"]["department"] = department

    checkouts = await prisma.assetscheckout.find_many(
        where=where_clause,
        include={
            "asset": {
                "include": {"category": True, "subCategory": True}
            },
            "employeeUser": True,
            "checkins": True
        },
        order={"checkoutDate": "desc"},
        take=5000
    )

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
    
    results = []
    for checkout in checkouts:
        expected_return = None
        is_overdue = False
        if checkout.expectedReturnDate:
            expected_return = checkout.expectedReturnDate.replace(tzinfo=None) if checkout.expectedReturnDate.tzinfo else checkout.expectedReturnDate
            expected_return = expected_return.replace(hour=0, minute=0, second=0, microsecond=0)
            is_overdue = expected_return < today and len(checkout.checkins) == 0

        results.append(
            TransactionItem(
                id=f"checkout-{checkout.id}",
                transactionType="Checkout Asset",
                assetTagId=checkout.asset.assetTagId,
                assetDescription=checkout.asset.description,
                category=checkout.asset.category.name if checkout.asset.category else None,
                subCategory=checkout.asset.subCategory.name if checkout.asset.subCategory else None,
                transactionDate=checkout.checkoutDate.isoformat(),
                actionBy=checkout.employeeUser.name if checkout.employeeUser else None,
                details=f"Checked out to {checkout.employeeUser.name if checkout.employeeUser else 'Unknown'}",
                location=checkout.asset.location,
                site=checkout.asset.site,
                department=checkout.asset.department,
                assetCost=float(checkout.asset.cost) if checkout.asset.cost else None,
                employeeName=checkout.employeeUser.name if checkout.employeeUser else None,
                checkoutDate=checkout.checkoutDate.isoformat(),
                expectedReturnDate=checkout.expectedReturnDate.isoformat() if checkout.expectedReturnDate else None,
                isOverdue=is_overdue,
            )
        )

    return results

async def _fetch_checkin_asset_transactions(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Checkin Asset transactions"""
    if transaction_type and transaction_type != 'Checkin Asset':
        return []
    
    where_clause: Dict[str, Any] = {
        "asset": {"isDeleted": False}
    }
    
    if date_filter:
        where_clause["checkinDate"] = date_filter
    if category:
        where_clause["asset"]["category"] = {"name": category}
    if location:
        where_clause["asset"]["location"] = location
    if site:
        where_clause["asset"]["site"] = site
    if department:
        where_clause["asset"]["department"] = department

    checkins = await prisma.assetscheckin.find_many(
        where=where_clause,
        include={
            "asset": {
                "include": {"category": True, "subCategory": True}
            },
            "employeeUser": True
        },
        order={"checkinDate": "desc"},
        take=5000
    )

    return [
        TransactionItem(
            id=f"checkin-{checkin.id}",
            transactionType="Checkin Asset",
            assetTagId=checkin.asset.assetTagId,
            assetDescription=checkin.asset.description,
            category=checkin.asset.category.name if checkin.asset.category else None,
            subCategory=checkin.asset.subCategory.name if checkin.asset.subCategory else None,
            transactionDate=checkin.checkinDate.isoformat(),
            actionBy=checkin.employeeUser.name if checkin.employeeUser else None,
            details=f"Checked in from {checkin.employeeUser.name if checkin.employeeUser else 'Unknown'}",
            location=checkin.asset.location,
            site=checkin.asset.site,
            department=checkin.asset.department,
            assetCost=float(checkin.asset.cost) if checkin.asset.cost else None,
            employeeName=checkin.employeeUser.name if checkin.employeeUser else None,
            checkinDate=checkin.checkinDate.isoformat(),
            condition=checkin.condition,
            notes=checkin.notes,
        )
        for checkin in checkins
    ]

async def _fetch_actions_by_users(
    transaction_type: Optional[str],
    category: Optional[str],
    location: Optional[str],
    site: Optional[str],
    department: Optional[str],
    action_by: Optional[str],
    date_filter: Optional[Dict[str, Any]]
) -> List[TransactionItem]:
    """Fetch Actions By Users (all history logs grouped by user)"""
    if transaction_type and transaction_type != 'Actions By Users':
        return []
    
    where_clause: Dict[str, Any] = {}
    
    if date_filter:
        where_clause["eventDate"] = date_filter
    if action_by:
        where_clause["actionBy"] = {"contains": action_by, "mode": "insensitive"}
    if category:
        where_clause["asset"] = {"category": {"name": category}}
    if location:
        where_clause["asset"] = {**where_clause.get("asset", {}), "location": location}
    if site:
        where_clause["asset"] = {**where_clause.get("asset", {}), "site": site}
    if department:
        where_clause["asset"] = {**where_clause.get("asset", {}), "department": department}

    history_logs = await prisma.assetshistorylogs.find_many(
        where=where_clause,
        include={
            "asset": {
                "include": {"category": True, "subCategory": True}
            }
        },
        order={"eventDate": "desc"},
        take=10000
    )

    event_type_map = {
        'added': 'Add Asset',
        'edited': 'Edit Asset',
        'deleted': 'Delete Asset',
    }

    return [
        TransactionItem(
            id=f"action-{log.id}",
            transactionType=event_type_map.get(log.eventType, 'Edit Asset'),
            assetTagId=log.asset.assetTagId,
            assetDescription=log.asset.description,
            category=log.asset.category.name if log.asset.category else None,
            subCategory=log.asset.subCategory.name if log.asset.subCategory else None,
            transactionDate=log.eventDate.isoformat(),
            actionBy=log.actionBy,
            details=(
                f'Field "{log.field}" changed from "{log.changeFrom or "N/A"}" to "{log.changeTo or "N/A"}"'
                if log.eventType == 'edited' and log.field
                else 'Asset added to system' if log.eventType == 'added'
                else 'Asset deleted' if log.eventType == 'deleted'
                else 'Asset action'
            ),
            fieldChanged=log.field,
            oldValue=log.changeFrom,
            newValue=log.changeTo,
            location=log.asset.location,
            site=log.asset.site,
            department=log.asset.department,
            assetCost=float(log.asset.cost) if log.asset.cost else None,
        )
        for log in history_logs
    ]

@router.get("", response_model=TransactionReportResponse)
async def get_transaction_reports(
    transactionType: Optional[str] = Query(None, description="Filter by transaction type"),
    category: Optional[str] = Query(None, description="Filter by category name"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    actionBy: Optional[str] = Query(None, description="Filter by action by (user name)"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=1000),
    auth: dict = Depends(verify_auth)
):
    """Get transaction reports with optional filters and pagination"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")

        skip = (page - 1) * pageSize

        # Build date filter
        date_filter: Optional[Dict[str, Any]] = None
        if startDate or endDate:
            date_filter = {}
            if startDate:
                date_obj = datetime.fromisoformat(startDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                date_filter["gte"] = date_obj
            if endDate:
                date_obj = datetime.fromisoformat(endDate.replace('Z', '+00:00'))
                if date_obj.tzinfo is not None:
                    date_obj = date_obj.replace(tzinfo=None)
                date_filter["lte"] = date_obj

        # Execute all queries in parallel when fetching all transaction types
        # If "Actions By Users" is selected, only fetch that
        if transactionType == 'Actions By Users':
            transactions = await _fetch_actions_by_users(
                transactionType, category, location, site, department, actionBy, date_filter
            )
        else:
            transactions_lists = await asyncio.gather(
                _fetch_add_asset_transactions(transactionType, category, location, site, department, date_filter),
                _fetch_edit_asset_transactions(transactionType, category, location, site, department, actionBy, date_filter),
                _fetch_delete_asset_transactions(transactionType, category, location, site, department, date_filter),
                _fetch_disposal_transactions(transactionType, category, location, site, department, date_filter),
                _fetch_lease_out_transactions(transactionType, category, location, site, department, date_filter),
                _fetch_lease_return_transactions(transactionType, category, location, site, department, date_filter),
                _fetch_repair_asset_transactions(transactionType, category, location, site, department, date_filter),
                _fetch_move_asset_transactions(transactionType, category, location, site, department, date_filter),
                _fetch_checkout_asset_transactions(transactionType, category, location, site, department, date_filter),
                _fetch_checkin_asset_transactions(transactionType, category, location, site, department, date_filter),
            )
            
            transactions = []
            for tx_list in transactions_lists:
                transactions.extend(tx_list)

        # Filter by actionBy if specified (for non-Actions By Users queries)
        if actionBy and transactionType != 'Actions By Users':
            transactions = [
                t for t in transactions
                if t.actionBy and actionBy.lower() in t.actionBy.lower()
            ]

        # Sort by transaction date (newest first)
        transactions.sort(key=lambda t: t.transactionDate, reverse=True)

        # Calculate summary statistics
        total_transactions = len(transactions)
        by_type_map: Dict[str, Dict[str, Any]] = {}
        for trans in transactions:
            trans_type = trans.transactionType
            if trans_type not in by_type_map:
                by_type_map[trans_type] = {"count": 0, "totalValue": 0.0}
            by_type_map[trans_type]["count"] += 1
            by_type_map[trans_type]["totalValue"] += trans.assetCost or 0.0

        by_type = [
            TransactionTypeGroup(
                type=trans_type,
                count=stats["count"],
                totalValue=stats["totalValue"],
            )
            for trans_type, stats in by_type_map.items()
        ]

        # Paginate
        paginated_transactions = transactions[skip:skip + pageSize]
        total_pages = (total_transactions + pageSize - 1) // pageSize if total_transactions > 0 else 0

        return TransactionReportResponse(
            transactions=paginated_transactions,
            summary=TransactionSummary(
                totalTransactions=total_transactions,
                byType=by_type,
            ),
            generatedAt=datetime.now().isoformat(),
            pagination=PaginationInfo(
                total=total_transactions,
                page=page,
                pageSize=pageSize,
                totalPages=total_pages,
                hasNextPage=page < total_pages,
                hasPreviousPage=page > 1,
            )
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching transaction reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch transaction reports")

@router.get("/export")
async def export_transaction_reports(
    format: str = Query("csv", description="Export format: csv or excel"),
    transactionType: Optional[str] = Query(None, description="Filter by transaction type"),
    category: Optional[str] = Query(None, description="Filter by category name"),
    location: Optional[str] = Query(None, description="Filter by location"),
    site: Optional[str] = Query(None, description="Filter by site"),
    department: Optional[str] = Query(None, description="Filter by department"),
    actionBy: Optional[str] = Query(None, description="Filter by action by (user name)"),
    startDate: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    endDate: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    includeTransactionList: Optional[bool] = Query(False, description="Include transaction list in export"),
    auth: dict = Depends(verify_auth)
):
    """Export transaction reports to CSV or Excel"""
    try:
        user_id = auth.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Unauthorized")
        
        # Check permission - user must have canManageReports to export reports
        has_permission = await check_permission(user_id, "canManageReports")
        if not has_permission:
            raise HTTPException(
                status_code=403,
                detail="You do not have permission to export reports"
            )

        if format not in ["csv", "excel", "pdf"]:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv, excel, or pdf.")
        
        if format == "pdf" and not PDF_AVAILABLE:
            raise HTTPException(status_code=500, detail="PDF export not available - fpdf2 not installed")

        # Fetch all data for export
        page_size = 10000 if includeTransactionList else 1
        report_data = await get_transaction_reports(
            transactionType=transactionType,
            category=category,
            location=location,
            site=site,
            department=department,
            actionBy=actionBy,
            startDate=startDate,
            endDate=endDate,
            page=1,
            pageSize=page_size,
            auth=auth
        )

        transactions = report_data.transactions
        summary = report_data.summary

        filename = f"transaction-report-{datetime.now().strftime('%Y-%m-%d')}"

        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)

            if includeTransactionList:
                # Include summary and transaction list
                writer.writerow(["TRANSACTION REPORT SUMMARY"])
                writer.writerow(["Total Transactions", summary.totalTransactions])
                writer.writerow([])
                
                writer.writerow(["TRANSACTIONS BY TYPE"])
                writer.writerow(["Transaction Type", "Count", "Total Asset Value"])
                for item in summary.byType:
                    writer.writerow([
                        item.type,
                        item.count,
                        format_number(item.totalValue),
                    ])
                writer.writerow([])
                
                writer.writerow(["TRANSACTION RECORDS"])
                writer.writerow([
                    "Transaction Type",
                    "Asset Tag ID",
                    "Description",
                    "Category",
                    "Sub-Category",
                    "Transaction Date",
                    "Action By",
                    "Details",
                    "Location",
                    "Site",
                    "Department",
                    "Asset Cost",
                ])
                for transaction in transactions:
                    writer.writerow([
                        transaction.transactionType,
                        transaction.assetTagId,
                        transaction.assetDescription,
                        transaction.category or "N/A",
                        transaction.subCategory or "N/A",
                        transaction.transactionDate.split('T')[0] if transaction.transactionDate else "N/A",
                        transaction.actionBy or "N/A",
                        transaction.details or "N/A",
                        transaction.location or "N/A",
                        transaction.site or "N/A",
                        transaction.department or "N/A",
                        format_number(transaction.assetCost),
                    ])
            else:
                # Summary only
                writer.writerow(["TRANSACTION REPORT SUMMARY"])
                writer.writerow(["Total Transactions", summary.totalTransactions])
                writer.writerow([])
                writer.writerow(["TRANSACTIONS BY TYPE"])
                writer.writerow(["Transaction Type", "Count", "Total Asset Value"])
                for item in summary.byType:
                    writer.writerow([
                        item.type,
                        item.count,
                        format_number(item.totalValue),
                    ])

            csv_content = output.getvalue()
            output.close()

            filename += ".csv"
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        elif format == "excel":
            # Import openpyxl for Excel generation
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="Excel export not available - openpyxl not installed"
                )

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Summary sheet
            summary_data = [
                ["TRANSACTION REPORT SUMMARY"],
                ["Total Transactions", summary.totalTransactions],
                [],
                ["TRANSACTIONS BY TYPE"],
                ["Transaction Type", "Count", "Total Asset Value"],
                *[[item.type, item.count, format_number(item.totalValue)] for item in summary.byType],
            ]
            summary_ws = wb.create_sheet("Summary")
            for row in summary_data:
                summary_ws.append(row)

            if includeTransactionList:
                # Transaction list sheet
                transaction_data = [
                    [
                        transaction.transactionType,
                        transaction.assetTagId,
                        transaction.assetDescription,
                        transaction.category or "N/A",
                        transaction.subCategory or "N/A",
                        transaction.transactionDate.split('T')[0] if transaction.transactionDate else "N/A",
                        transaction.actionBy or "N/A",
                        transaction.details or "N/A",
                        transaction.location or "N/A",
                        transaction.site or "N/A",
                        transaction.department or "N/A",
                        format_number(transaction.assetCost),
                    ]
                    for transaction in transactions
                ]
                
                transaction_ws = wb.create_sheet("Transactions")
                transaction_ws.append([
                    "Transaction Type",
                    "Asset Tag ID",
                    "Description",
                    "Category",
                    "Sub-Category",
                    "Transaction Date",
                    "Action By",
                    "Details",
                    "Location",
                    "Site",
                    "Department",
                    "Asset Cost",
                ])
                for row in transaction_data:
                    transaction_ws.append(row)

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename += ".xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )

        else:  # pdf
            pdf = ReportPDF("Transaction Report", "Transaction")
            pdf.add_page()

            pdf.add_section_title("Summary Statistics")
            # Overall summary
            overall_rows = [
                ["Total Transactions", str(summary.totalTransactions)],
            ]
            pdf.add_table(["Metric", "Value"], overall_rows)
            
            pdf.ln(5)
            
            # Breakdown by type
            if summary.byType:
                pdf.add_section_title("Transactions By Type")
                type_rows = [
                    [item.type, str(item.count), format_number(item.totalValue)]
                    for item in summary.byType
                ]
                pdf.add_table(["Transaction Type", "Count", "Total Value"], type_rows)
            
            pdf.ln(10)

            if transactions and includeTransactionList:
                pdf.add_section_title(f"Transaction List ({len(transactions)} transactions)")
                
                # Different columns based on transaction type filter
                trans_type = transactionType or ""
                
                if trans_type == "" or trans_type.lower() == "all":
                    # All Transactions
                    simplified_headers = ["Transaction Type", "Asset Tag ID", "Description", "Category", "Date", "Action By", "Details", "Location", "Asset Cost"]
                    simplified_rows = [
                        [
                            str(t.transactionType or ""),
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:35],
                            str(t.category or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.actionBy or ""),
                            str(t.details or "")[:25],
                            str(t.location or ""),
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Add Asset":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Date", "Added By", "Location", "Site", "Department", "Asset Cost"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:40],
                            str(t.category or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.actionBy or ""),
                            str(t.location or ""),
                            str(t.site or ""),
                            str(t.department or ""),
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                elif trans_type in ["Sold Asset", "Donated Asset", "Scrapped Asset", "Lost/Missing Asset", "Destroyed Asset"]:
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Disposal Date", "Reason", "Disposal Value", "Location", "Original Cost"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:40],
                            str(t.category or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.details or "")[:30],
                            format_number(t.disposeValue),
                            str(t.location or ""),
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Edit Asset":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Date", "Edited By", "Field Changed", "Old Value", "New Value"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:35],
                            str(t.category or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.actionBy or ""),
                            str(t.fieldChanged or ""),
                            str(t.oldValue or "")[:20],
                            str(t.newValue or "")[:20],
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Lease Out":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Lessee", "Lease Start", "Lease End", "Conditions", "Asset Cost"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:35],
                            str(t.category or ""),
                            str(t.lessee or ""),
                            t.leaseStartDate.split('T')[0] if t.leaseStartDate else "",
                            t.leaseEndDate.split('T')[0] if t.leaseEndDate else "",
                            str(t.conditions or "")[:25],
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Lease Return":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Lessee", "Return Date", "Condition", "Notes", "Asset Cost"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:35],
                            str(t.category or ""),
                            str(t.lessee or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.condition or ""),
                            str(t.notes or "")[:25],
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Repair Asset":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Title", "Maintained By", "Due Date", "Status", "Cost", "Completed"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:30],
                            str(t.category or ""),
                            str(t.title or "")[:25],
                            str(t.maintenanceBy or ""),
                            t.dueDate.split('T')[0] if t.dueDate else "",
                            str(t.status or ""),
                            format_number(t.cost),
                            t.dateCompleted.split('T')[0] if t.dateCompleted else "",
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Move Asset":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Move Type", "Move Date", "Assigned To", "Reason", "From", "To"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:30],
                            str(t.category or ""),
                            str(t.moveType or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.employeeName or ""),
                            str(t.reason or "")[:20],
                            str(t.fromLocation or ""),
                            str(t.toLocation or ""),
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Checkout Asset":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Checked Out To", "Checkout Date", "Expected Return", "Status", "Location", "Asset Cost"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:30],
                            str(t.category or ""),
                            str(t.employeeName or ""),
                            t.checkoutDate.split('T')[0] if t.checkoutDate else "",
                            t.expectedReturnDate.split('T')[0] if t.expectedReturnDate else "",
                            "Overdue" if t.isOverdue else "Active",
                            str(t.location or ""),
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Checkin Asset":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Checked In From", "Checkin Date", "Condition", "Notes", "Location", "Asset Cost"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:30],
                            str(t.category or ""),
                            str(t.employeeName or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.condition or ""),
                            str(t.notes or "")[:20],
                            str(t.location or ""),
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Delete Asset":
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Deleted By", "Deleted Date", "Reason", "Location", "Asset Cost"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:35],
                            str(t.category or ""),
                            str(t.actionBy or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.details or "")[:25],
                            str(t.location or ""),
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                elif trans_type == "Actions By Users":
                    simplified_headers = ["Action By", "Action Type", "Asset Tag ID", "Description", "Date", "Details"]
                    simplified_rows = [
                        [
                            str(t.actionBy or ""),
                            str(t.transactionType or ""),
                            str(t.assetTagId or ""),
                            str(t.assetDescription or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.details or ""),
                        ]
                        for t in transactions
                    ]
                else:
                    # Default fallback
                    simplified_headers = ["Asset Tag ID", "Description", "Category", "Date", "Action By", "Details", "Location", "Asset Cost"]
                    simplified_rows = [
                        [
                            str(t.assetTagId or ""),
                            str(t.assetDescription or "")[:40],
                            str(t.category or ""),
                            t.transactionDate.split('T')[0] if t.transactionDate else "",
                            str(t.actionBy or ""),
                            str(t.details or "")[:30],
                            str(t.location or ""),
                            format_number(t.assetCost),
                        ]
                        for t in transactions
                    ]
                pdf.add_table(simplified_headers, simplified_rows)

            pdf_content = bytes(pdf.output())
            filename += ".pdf"
            return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting transaction reports: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export transaction reports")

